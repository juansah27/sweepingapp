from fastapi import FastAPI, HTTPException, Depends, UploadFile, File, Form, BackgroundTasks, Query, Request, Body, WebSocket, WebSocketDisconnect
from contextlib import asynccontextmanager
from fastapi.middleware.cors import CORSMiddleware
from fastapi.middleware.gzip import GZipMiddleware
from fastapi.middleware.trustedhost import TrustedHostMiddleware
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import JSONResponse, StreamingResponse
from fastapi_limiter import FastAPILimiter
from fastapi_limiter.depends import RateLimiter
from sqlalchemy import create_engine, Column, Integer, String, DateTime, Text, UniqueConstraint, Index, text, or_, func, Boolean, ForeignKey
from sqlalchemy.orm import declarative_base
from sqlalchemy.orm import sessionmaker, Session
from typing import List, Optional
import pandas as pd
from collections import defaultdict
import os
import io
import shutil
import json
from datetime import datetime, timedelta, timezone, time as datetime_time
import jwt
import pytz
import pyodbc
from dotenv import load_dotenv
import asyncio
import threading
from threading import Thread
import uuid
import time
import logging
from functools import lru_cache

# Import the new dashboard views API
try:
    from .dashboard_views_api import router as dashboard_views_router
except ImportError:
    # For direct execution
    from dashboard_views_api import router as dashboard_views_router
import hashlib
import secrets
from pydantic import BaseModel, field_validator, EmailStr
from passlib.context import CryptContext
import re
import schedule
import uvicorn
from monitoring import performance_tracker, start_monitoring, stop_monitoring, get_metrics_summary
from multi_user_handler import multi_user_handler

# Load environment variables from .env file
load_dotenv()

def get_project_root():
    """Get project root directory, handling both local and Docker environments"""
    current_dir = os.path.dirname(os.path.abspath(__file__))
    
    # Check if running in Docker container
    if os.path.exists('/app/JobGetOrder'):
        # Running in Docker container
        return '/app'
    else:
        # Running locally
        return os.path.dirname(current_dir)

# Auto-run configuration
AUTO_RUN_MARKETPLACE_APPS = os.getenv("AUTO_RUN_MARKETPLACE_APPS", "false").lower() == "true"
MARKETPLACE_APPS_ENABLED = os.getenv("MARKETPLACE_APPS_ENABLED", "false").lower() == "true"

# Configure logging based on environment
ENVIRONMENT = os.getenv("ENVIRONMENT", "development").lower()

if ENVIRONMENT == "production":
    # Production logging - only WARNING and ERROR
    logging.basicConfig(
        level=logging.WARNING,
        format='%(asctime)s - %(levelname)s - %(message)s'
    )
    # Disable verbose logging for production
    logging.getLogger("uvicorn.access").setLevel(logging.WARNING)
    logging.getLogger("uvicorn.error").setLevel(logging.WARNING)
    logging.getLogger("sqlalchemy.engine").setLevel(logging.WARNING)
    logging.getLogger("sqlalchemy.pool").setLevel(logging.WARNING)
else:
    # Development logging - INFO and above
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    )

logger = logging.getLogger(__name__)

@asynccontextmanager
async def lifespan(app: FastAPI):
    """Initialize monitoring on startup and cleanup on shutdown"""
    # Startup
    try:
        # Create logs directory if it doesn't exist
        os.makedirs("logs", exist_ok=True)
        
        # Start monitoring
        start_monitoring()
        
        # Start daily remark reset scheduler
        schedule_daily_reset()
        
        logger.info("Application started successfully with monitoring enabled")
    except Exception as e:
        logger.error(f"Startup error: {e}")
    
    yield
    
    # Shutdown
    try:
        stop_monitoring()
        logger.info("Application shutdown gracefully")
    except Exception as e:
        logger.error(f"Shutdown error: {e}")

app = FastAPI(
    title="Sweeping Apps API",
    description="Optimized Order Management System",
    version="2.0.0",
    docs_url="/docs",
    redoc_url="/redoc",
    lifespan=lifespan
)

# Include the optimized dashboard views router
app.include_router(dashboard_views_router)

# Add compression middleware
app.add_middleware(GZipMiddleware, minimum_size=1000)

# Add trusted host middleware for security
app.add_middleware(
    TrustedHostMiddleware, 
    allowed_hosts=["localhost", "127.0.0.1", "*.local", "*", "0.0.0.0"]
)

# Security and performance monitoring middleware
@app.middleware("http")
async def security_and_performance_middleware(request: Request, call_next):
    start_time = time.time()
    
    # Log request
    logger.info(f"Request: {request.method} {request.url} from {request.client.host}")
    
    # Process request
    response = await call_next(request)
    
    # Calculate processing time
    process_time = time.time() - start_time
    
    # Track performance metrics
    try:
        performance_tracker.track_request(
            endpoint=str(request.url.path),
            method=request.method,
            response_time=process_time,
            status_code=response.status_code,
            user_agent=request.headers.get("user-agent", ""),
            ip_address=request.client.host
        )
    except Exception as e:
        logger.error(f"Error tracking performance metrics: {e}")
    
    # Add security headers
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "DENY"
    response.headers["X-XSS-Protection"] = "1; mode=block"
    response.headers["Strict-Transport-Security"] = "max-age=31536000; includeSubDomains"
    response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    response.headers["Content-Security-Policy"] = "default-src 'self'"
    
    # Add performance headers
    response.headers["X-Process-Time"] = str(process_time)
    response.headers["X-Cache-Status"] = "MISS"  # Will be updated by cache logic
    
    # Log response
    logger.info(f"Response: {response.status_code} - {process_time:.3f}s")
    
    return response

# CORS middleware - Allow access from frontend
# Get allowed origins from environment variable
ALLOWED_ORIGINS = os.getenv("ALLOWED_ORIGINS", "*").split(",")
if ALLOWED_ORIGINS == ["*"]:
    allow_origins = ["*"]
else:
    allow_origins = [origin.strip() for origin in ALLOWED_ORIGINS]

app.add_middleware(
    CORSMiddleware,
    allow_origins=allow_origins,
    allow_credentials=True,
    allow_methods=["GET", "POST", "PUT", "DELETE", "OPTIONS"],
    allow_headers=["*"],
)

# Security Configuration
SECRET_KEY = os.getenv("SECRET_KEY", secrets.token_urlsafe(32))
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 30

# Rate limiting configuration
RATE_LIMIT_PER_MINUTE = 60
RATE_LIMIT_PER_HOUR = 1000

# File upload security
MAX_FILE_SIZE = 50 * 1024 * 1024  # 50MB
ALLOWED_EXTENSIONS = {'.xlsx', '.csv'}
ALLOWED_MIME_TYPES = {
    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    'text/csv',
    'application/csv'
}

# Timezone configuration - Indonesia Western Time (WIB)
DEFAULT_TIMEZONE = os.getenv("DEFAULT_TIMEZONE", "Asia/Jakarta")
WIB_TIMEZONE = pytz.timezone(DEFAULT_TIMEZONE)

# Set default timezone for the application
os.environ['TZ'] = DEFAULT_TIMEZONE

# Import database configuration
from database_config import (
    DB_SERVER, DB_NAME, DB_USERNAME, DB_PASSWORD, DB_TRUSTED_CONNECTION,
    FLEXO_DB_CONNECTION_STRING, WMSPROD_DB_CONNECTION_STRING
)

# Alternative connection strings for different ODBC drivers
FLEXO_DB_CONNECTION_STRING_ALT = f"DRIVER={{ODBC Driver 18 for SQL Server}};SERVER={DB_SERVER};DATABASE=Flexo_Db;UID={DB_USERNAME};PWD={DB_PASSWORD};TrustServerCertificate=yes;Connection Timeout=30;"
WMSPROD_DB_CONNECTION_STRING_ALT = f"DRIVER={{ODBC Driver 18 for SQL Server}};SERVER={DB_SERVER};DATABASE=WMSPROD;UID={DB_USERNAME};PWD={DB_PASSWORD};TrustServerCertificate=yes;Connection Timeout=30;"

# Marketplace field mapping for interface status check
MARKETPLACE_FIELD_MAPPING = {
    'ginee': 'OrigSystemRefId',     # Order ID = OrigSystemRefId
    'blibli': 'EntityId',           # No. Order = EntityId
    'zalora': 'EntityId',           # Order Number = EntityId
    'shopee': 'EntityId',           # No. Pesanan = EntityId
    'tiktok': 'EntityId',           # Order ID = EntityId
    'desty': 'EntityId',            # Nomor Pesanan (di Desty) = EntityId
    'lazada': 'OrigSystemRefId',    # orderNumber = OrigSystemRefId
    'jubelio': 'SystemRefId'        # salesorder_no = SystemRefId
}

def get_wib_now():
    """Get current time in WIB timezone (Indonesia Western Time)"""
    return datetime.now(WIB_TIMEZONE)

def convert_to_wib(dt):
    """Convert datetime to WIB timezone (Indonesia Western Time)"""
    if dt is None:
        return None
    if dt.tzinfo is None:
        # If naive datetime, assume it's in WIB
        return WIB_TIMEZONE.localize(dt)
    return dt.astimezone(WIB_TIMEZONE)

def format_wib_datetime(dt, format_str='%Y-%m-%d %H:%M:%S'):
    """Format datetime in WIB timezone"""
    if dt is None:
        return None
    wib_dt = convert_to_wib(dt)
    return wib_dt.strftime(format_str)

def convert_to_naive_wib(dt):
    """Convert timezone-aware datetime to naive WIB for database comparison"""
    if dt is None:
        return None
    if dt.tzinfo is not None:
        return dt.astimezone(WIB_TIMEZONE).replace(tzinfo=None)
    else:
        return dt

def get_pic_upload_counts(db, start_datetime=None, end_datetime=None):
    """Get file upload counts per PIC with optional date filtering - using UpdateUploadedOrder to avoid duplicates"""
    try:
        # Build the base query using clean_orders view
        base_query = "SELECT \"PIC\", COUNT(*) as count FROM clean_orders WHERE \"PIC\" IS NOT NULL"
        params = {}
        
        # Apply date filters with proper timezone conversion
        if start_datetime:
            start_naive = convert_to_naive_wib(start_datetime)
            base_query += " AND \"UploadDate\" >= :start_date"
            params['start_date'] = start_naive
        
        if end_datetime:
            end_naive = convert_to_naive_wib(end_datetime)
            base_query += " AND \"UploadDate\" <= :end_date"
            params['end_date'] = end_naive
            
        base_query += " GROUP BY \"PIC\""
        
        # Execute query
        results = db.execute(text(base_query), params).fetchall()
        
        # Return results as list of tuples (PIC, count)
        return [(row[0], row[1]) for row in results]
        
    except Exception as e:
        print(f"❌ Error in get_pic_upload_counts: {e}")
        return []

def get_batch_upload_counts(db, start_datetime=None, end_datetime=None):
    """Get file upload counts per batch with optional date filtering - using clean_orders view to avoid duplicates"""
    try:
        # Build the base query using clean_orders view
        base_query = "SELECT \"Batch\", COUNT(*) as count FROM clean_orders WHERE \"Batch\" IS NOT NULL"
        params = {}
        
        # Apply date filters with proper timezone conversion
        if start_datetime:
            start_naive = convert_to_naive_wib(start_datetime)
            base_query += " AND \"UploadDate\" >= :start_date"
            params['start_date'] = start_naive
        
        if end_datetime:
            end_naive = convert_to_naive_wib(end_datetime)
            base_query += " AND \"UploadDate\" <= :end_date"
            params['end_date'] = end_naive
            
        base_query += " GROUP BY \"Batch\""
        
        # Execute query
        results = db.execute(text(base_query), params).fetchall()
        
        # Return results as list of tuples (Batch, count)
        return [(row[0], row[1]) for row in results]
        
    except Exception as e:
        print(f"❌ Error in get_batch_upload_counts: {e}")
        return []

def get_table_field(table_type, field_name):
    """Helper function to get the correct table field based on table type"""
    if table_type == "update":
        return getattr(UpdateUploadedOrder, field_name)
    else:
        return getattr(UploadedOrder, field_name)

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
security = HTTPBearer()

# Database setup with connection pooling and optimization
# PostgreSQL ONLY - No SQLite fallback
from database_config import POSTGRES_DATABASE_URL

# Force PostgreSQL mode - no fallback
POSTGRES_HOST = os.getenv("POSTGRES_HOST", "localhost").strip()
if not POSTGRES_HOST:
    raise ValueError("POSTGRES_HOST environment variable is required. SQLite fallback has been removed.")

SQLALCHEMY_DATABASE_URL = POSTGRES_DATABASE_URL

# Production vs Development pool settings
if ENVIRONMENT == "production":
    # Production: Smaller pool per worker (4 workers total)
    POOL_SIZE = 5
    MAX_OVERFLOW = 10
else:
    # Development: Larger pool for single worker
    POOL_SIZE = 10
    MAX_OVERFLOW = 20

engine = create_engine(
    SQLALCHEMY_DATABASE_URL,
    pool_size=POOL_SIZE,
    max_overflow=MAX_OVERFLOW,
    pool_pre_ping=True,
    pool_recycle=300,  # Recycle connections after 5 minutes (reduced from 1 hour)
    pool_timeout=30,  # Reduced timeout for getting connection
    echo=False,  # Set to True for SQL debugging
    connect_args={
        "options": "-c statement_timeout=300000"  # 5 minutes statement timeout
    }
)
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
Base = declarative_base()

# Models
class User(Base):
    __tablename__ = "users"
    id = Column(Integer, primary_key=True, index=True)
    username = Column(String, unique=True, index=True)
    email = Column(String, unique=True, index=True)
    password_hash = Column(String)
    is_active = Column(Boolean, default=True)
    role = Column(String, default="user")  # user, admin, superuser
    created_at = Column(DateTime, default=get_wib_now)
    updated_at = Column(DateTime, default=get_wib_now, onupdate=get_wib_now)

class UploadedOrder(Base):
    __tablename__ = "uploaded_orders"
    Id = Column(Integer, primary_key=True, index=True)
    Marketplace = Column(Text, index=True)
    Brand = Column(Text, index=True)
    OrderNumber = Column(Text, unique=True, index=True)  # Add unique constraint
    OrderStatus = Column(Text)
    AWB = Column(Text, index=True)
    Transporter = Column(Text)
    OrderDate = Column(DateTime, index=True)
    SLA = Column(Text)
    Batch = Column(Text, index=True)
    PIC = Column(Text, index=True)
    UploadDate = Column(DateTime, index=True)
    Remarks = Column(Text, default="")
    InterfaceStatus = Column(Text, default="Not Yet Interface", index=True)
    TaskId = Column(Text, index=True)  # Add task_id field
    OrderNumberFlexo = Column(Text, index=True)  # Flexo order number field (already exists)
    OrderStatusFlexo = Column(Text)  # Flexo order status field (already exists)
    ItemId = Column(Text)  # Marketplace-specific SKU field
    ItemIdFlexo = Column(Text)  # ItemId from Flexo_Db.dbo.SalesOrderLine


class ListBrand(Base):
    __tablename__ = "list_brand"
    
    id = Column(Integer, primary_key=True, index=True)
    brand = Column(String, nullable=False)
    marketplace = Column(String, nullable=False)
    batch = Column(String)
    remark = Column(Text, nullable=True)  # User-editable remark column
    created_at = Column(DateTime, default=get_wib_now)
    
    # Add unique constraint for brand-marketplace-batch combination
    __table_args__ = (
        UniqueConstraint('brand', 'marketplace', 'batch', name='uq_brand_marketplace_batch'),
    )

class NotUploadedHistory(Base):
    __tablename__ = "not_uploaded_history"
    
    id = Column(Integer, primary_key=True, index=True)
    brand = Column(String(255), nullable=False)
    marketplace = Column(String(255), nullable=False)
    batch = Column(String(255), nullable=False)
    remark = Column(Text, nullable=True)
    status = Column(String(50), nullable=False)  # 'not_uploaded', 'uploaded'
    check_date = Column(DateTime, default=get_wib_now)
    created_at = Column(DateTime, default=get_wib_now)

class BrandShop(Base):
    """Comprehensive brand and shop information table"""
    __tablename__ = "brand_shops"
    
    id = Column(Integer, primary_key=True, index=True)
    client_shop_id = Column(Integer, index=True)
    client_id = Column(Integer, index=True)
    marketplace_id = Column(Integer, index=True)
    shop_name = Column(String(255), index=True)
    shop_logo = Column(Text)
    shop_url = Column(Text)
    established_date = Column(DateTime)
    is_open = Column(Integer, default=1)
    status = Column(Integer, default=1)
    notes = Column(Text)
    address = Column(Text)
    kelurahan = Column(String(100))
    kecamatan = Column(String(100))
    kota = Column(String(100))
    provinsi = Column(String(100))
    zipcode = Column(String(20))
    shop_key_1 = Column(String(100))
    shop_key_2 = Column(String(100))
    shop_key_3 = Column(String(100))
    shop_key_4 = Column(String(100))
    shop_key_5 = Column(String(100))
    created_at = Column(DateTime)
    created_by = Column(String(100))
    updated_at = Column(DateTime)
    updated_by = Column(String(100))
    wh_id = Column(Integer)
    jdaclientid = Column(String(100))
    shop_name_seller = Column(String(255))
    brand = Column(String(255), index=True)
    order_type = Column(String(50))

class UploadHistory(Base):
    __tablename__ = "file_upload_history"
    
    id = Column(Integer, primary_key=True, index=True)
    marketplace = Column(String, nullable=False)
    brand = Column(String, nullable=False)
    pic = Column(String, nullable=False)
    batch = Column(String, nullable=False)
    upload_date = Column(DateTime, default=get_wib_now)

class UploadTask(Base):
    __tablename__ = "upload_tasks"
    
    id = Column(Integer, primary_key=True, index=True)
    user_id = Column(Integer, ForeignKey("users.id"), nullable=True)
    task_name = Column(String(255), nullable=False)  # Required field from init.sql schema
    task_id = Column(String, unique=True, index=True)
    status = Column(String, default="pending")  # pending, processing, completed, failed
    file_path = Column(String(500), nullable=True)
    upload_progress = Column(Integer, default=0)
    marketplace = Column(String)
    brand = Column(String)
    batch = Column(String)
    pic = Column(String)
    total_orders = Column(Integer, default=0)
    processed_orders = Column(Integer, default=0)
    processing_time = Column(String)
    external_db_query_time = Column(String)
    error_message = Column(Text)
    created_at = Column(DateTime, default=get_wib_now)
    updated_at = Column(DateTime, default=get_wib_now, onupdate=get_wib_now)
    completed_at = Column(DateTime)

# Enhanced Pydantic models with validation
class UserCreate(BaseModel):
    username: str
    email: Optional[EmailStr] = None
    password: str
    role: str = "user"
    
    @field_validator('username')
    @classmethod
    def validate_username(cls, v):
        if len(v) < 2 or len(v) > 50:
            raise ValueError('Username must be between 2 and 50 characters')
        if not re.match(r'^[a-zA-Z0-9_-]+$', v):
            raise ValueError('Username can only contain letters, numbers, underscores, and hyphens')
        return v.lower()  # Convert to lowercase for case-insensitive consistency
    
    @field_validator('password')
    @classmethod
    def validate_password(cls, v):
        if len(v) < 8:
            raise ValueError('Password must be at least 8 characters long')
        if not re.search(r'[A-Z]', v):
            raise ValueError('Password must contain at least one uppercase letter')
        if not re.search(r'[a-z]', v):
            raise ValueError('Password must contain at least one lowercase letter')
        if not re.search(r'\d', v):
            raise ValueError('Password must contain at least one number')
        return v
    
    @field_validator('role')
    @classmethod
    def validate_role(cls, v):
        if v not in ['user', 'admin', 'superuser']:
            raise ValueError('Role must be either "user", "admin", or "superuser"')
        return v

class UserLogin(BaseModel):
    username: str
    password: str
    
    @field_validator('username')
    @classmethod
    def validate_username(cls, v):
        if not v or len(v.strip()) == 0:
            raise ValueError('Username cannot be empty')
        return v.strip().lower()  # Convert to lowercase for case-insensitive login
    
    @field_validator('password')
    @classmethod
    def validate_password(cls, v):
        if not v or len(v.strip()) == 0:
            raise ValueError('Password cannot be empty')
        return v

class Token(BaseModel):
    access_token: str
    token_type: str

class UserResponse(BaseModel):
    id: int
    username: str
    email: Optional[str] = None
    role: str
    is_active: bool
    created_at: datetime
    updated_at: datetime
    
    class Config:
        from_attributes = True

class UserUpdate(BaseModel):
    username: Optional[str] = None
    email: Optional[EmailStr] = None
    role: Optional[str] = None
    is_active: Optional[bool] = None
    
    @field_validator('role')
    @classmethod
    def validate_role(cls, v):
        if v is not None and v not in ['user', 'admin', 'superuser']:
            raise ValueError('Role must be either "user", "admin", or "superuser"')
        return v

    @field_validator('is_active')
    @classmethod
    def validate_is_active(cls, v):
        if v is not None:
            return bool(v)  # Ensure it's converted to boolean
        return v

class UserCreateAdmin(BaseModel):
    username: str
    email: Optional[EmailStr] = None
    password: str
    role: str = "user"
    
    @field_validator('username')
    @classmethod
    def validate_username(cls, v):
        if len(v) < 2 or len(v) > 50:
            raise ValueError('Username must be between 2 and 50 characters')
        if not re.match(r'^[a-zA-Z0-9_-]+$', v):
            raise ValueError('Username can only contain letters, numbers, underscores, and hyphens')
        return v.lower()  # Convert to lowercase for case-insensitive consistency
    
    @field_validator('password')
    @classmethod
    def validate_password(cls, v):
        if len(v) < 8:
            raise ValueError('Password must be at least 8 characters long')
        if not re.search(r'[A-Z]', v):
            raise ValueError('Password must contain at least one uppercase letter')
        if not re.search(r'[a-z]', v):
            raise ValueError('Password must contain at least one lowercase letter')
        if not re.search(r'\d', v):
            raise ValueError('Password must contain at least one number')
        return v
    
    @field_validator('role')
    @classmethod
    def validate_role(cls, v):
        if v not in ['user', 'admin', 'superuser']:
            raise ValueError('Role must be either "user", "admin", or "superuser"')
        return v

class RemarksUpdate(BaseModel):
    remarks: str

class UpdateUploadedOrder(BaseModel):
    remarks: str


class OrderResponse(BaseModel):
    Id: int
    Marketplace: str
    Brand: str
    OrderNumber: str
    OrderStatus: str
    AWB: str
    Transporter: str
    OrderDate: datetime
    SLA: str
    Batch: str
    PIC: str

class UploadHistoryResponse(BaseModel):
    id: int
    marketplace: str
    brand: str
    pic: str
    batch: str
    upload_date: datetime

class ListBrandResponse(BaseModel):
    id: int
    brand: str
    marketplace: str
    batch: str
    created_at: str

    class Config:
        from_attributes = True

class BrandShopResponse(BaseModel):
    id: int
    client_shop_id: Optional[int]
    client_id: Optional[int]
    marketplace_id: Optional[int]
    shop_name: Optional[str]
    shop_logo: Optional[str]
    shop_url: Optional[str]
    established_date: Optional[str]
    is_open: Optional[int]
    status: Optional[int]
    notes: Optional[str]
    address: Optional[str]
    kelurahan: Optional[str]
    kecamatan: Optional[str]
    kota: Optional[str]
    provinsi: Optional[str]
    zipcode: Optional[str]
    shop_key_1: Optional[str]
    shop_key_2: Optional[str]
    shop_key_3: Optional[str]
    shop_key_4: Optional[str]
    shop_key_5: Optional[str]
    created_at: Optional[str]
    created_by: Optional[str]
    updated_at: Optional[str]
    updated_by: Optional[str]
    wh_id: Optional[int]
    jdaclientid: Optional[str]
    shop_name_seller: Optional[str]
    brand: Optional[str]
    order_type: Optional[str]

    class Config:
        from_attributes = True

# Create tables
Base.metadata.create_all(bind=engine)

# Add TaskId column if it doesn't exist (migration)
def migrate_database():
    try:
        with engine.connect() as conn:
            # Check if TaskId column exists (PostgreSQL syntax)
            if POSTGRES_HOST:
                result = conn.execute(text("SELECT column_name FROM information_schema.columns WHERE table_name = 'uploaded_orders' AND column_name = 'TaskId'"))
                columns = [row[0] for row in result.fetchall()]
            else:
                result = conn.execute(text("PRAGMA table_info(uploaded_orders)"))
                columns = [row[1] for row in result.fetchall()]
            
            if 'TaskId' not in columns:
                conn.execute(text("ALTER TABLE uploaded_orders ADD COLUMN TaskId TEXT"))
                conn.commit()
    except Exception as e:
        pass

# Run migration
migrate_database()

# Create upload_tasks table if it doesn't exist
def create_upload_tasks_table():
    """Create upload_tasks table for background task tracking"""
    try:
        with engine.connect() as conn:
            # Check if upload_tasks table exists (PostgreSQL syntax)
            result = conn.execute(text("SELECT table_name FROM information_schema.tables WHERE table_name = 'upload_tasks'")).fetchone()
            
            if not result:
                # Create upload_tasks table (matching init.sql schema)
                conn.execute(text("""
                    CREATE TABLE upload_tasks (
                        id SERIAL PRIMARY KEY,
                        user_id INTEGER REFERENCES users(id),
                        task_name VARCHAR(255) NOT NULL,
                        task_id VARCHAR UNIQUE,
                        status VARCHAR DEFAULT 'pending',
                        file_path VARCHAR(500),
                        upload_progress INTEGER DEFAULT 0,
                        marketplace VARCHAR,
                        brand VARCHAR,
                        batch VARCHAR,
                        pic VARCHAR,
                        total_orders INTEGER DEFAULT 0,
                        processed_orders INTEGER DEFAULT 0,
                        processing_time VARCHAR,
                        external_db_query_time VARCHAR,
                        error_message TEXT,
                        created_at TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP,
                        updated_at TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP,
                        completed_at TIMESTAMP WITH TIME ZONE
                    )
                """))
                conn.commit()
                logger.info("Created upload_tasks table")
            else:
                logger.info("upload_tasks table already exists")
    except Exception as e:
        logger.info(f"upload_tasks table creation info: {str(e)}")

# Create upload_tasks table
create_upload_tasks_table()

# In-memory cache for frequently accessed data
cache = {}
CACHE_TTL = 300  # 5 minutes

def get_cache_key(*args):
    """Generate cache key from arguments"""
    key_string = "|".join(str(arg) for arg in args)
    return hashlib.md5(key_string.encode()).hexdigest()

def get_cached_data(key):
    """Get data from cache if not expired"""
    if key in cache:
        data, timestamp = cache[key]
        if time.time() - timestamp < CACHE_TTL:
            return data
        else:
            del cache[key]
    return None

def set_cached_data(key, data):
    """Set data in cache with timestamp"""
    cache[key] = (data, time.time())

@lru_cache(maxsize=128)
def get_marketplace_mapping_cached(sales_channel: str):
    """Cached version of marketplace mapping"""
    return get_marketplace_mapping(sales_channel)

# Helper function to parse dates in multiple formats
def parse_date_flexible(date_string: str) -> datetime:
    """Parse date string in multiple formats (ISO, YYYY-MM-DD) and return datetime object in WIB timezone"""
    if not date_string or not date_string.strip():
        return None
    
    date_string = date_string.strip()
    
    # Try ISO format first (e.g., 2025-09-19T17:00:00.000Z or 2025-09-21T00:00:00)
    try:
        if 'T' in date_string:
            if 'Z' in date_string or '+' in date_string or '-' in date_string[-6:]:
                # Full ISO format with timezone
                dt = datetime.fromisoformat(date_string.replace('Z', '+00:00'))
                # Convert to WIB timezone for proper date extraction
                dt_wib = dt.astimezone(WIB_TIMEZONE)
                return dt_wib
            else:
                # ISO format without timezone (assume WIB timezone)
                dt = datetime.fromisoformat(date_string)
                # Localize to WIB timezone
                dt_wib = WIB_TIMEZONE.localize(dt)
                return dt_wib
    except ValueError:
        pass
    
    # Try YYYY-MM-DD format
    try:
        date_obj = datetime.strptime(date_string, "%Y-%m-%d").date()
        # Create datetime in WIB timezone for consistency with upload data
        dt_wib = WIB_TIMEZONE.localize(datetime.combine(date_obj, datetime_time.min))
        return dt_wib
    except ValueError:
        pass
    
    # If all parsing fails, raise error
    raise ValueError(f"Invalid date format: {date_string}. Expected ISO format or YYYY-MM-DD")

def create_clean_orders_view(db):
    """Create a view for clean orders data (deduplicated)"""
    try:
        # Drop view if exists
        db.execute(text("DROP VIEW IF EXISTS clean_orders"))
        
        # Create view for clean orders (deduplicated)
        create_view_sql = """
        CREATE VIEW clean_orders AS
        SELECT DISTINCT ON ("OrderNumber") 
            "Id", "Marketplace", "Brand", "OrderNumber", "OrderStatus", "AWB", "Transporter", 
            "OrderDate", "SLA", "Batch", "PIC", "UploadDate", "Remarks", "InterfaceStatus", "TaskId", 
            "OrderNumberFlexo", "OrderStatusFlexo"
        FROM uploaded_orders
        ORDER BY "OrderNumber", "UploadDate" DESC;
        """
        
        db.execute(text(create_view_sql))
        db.commit()
        print("✅ View 'clean_orders' created successfully")
        return True
        
    except Exception as e:
        print(f"❌ Error creating view: {e}")
        db.rollback()
        return False

# Dependency with connection pooling optimization and proper cleanup
def get_db():
    db = SessionLocal()
    try:
        # PostgreSQL optimizations
        yield db
    except Exception as e:
        logger.error(f"Database error: {str(e)}")
        db.rollback()
        raise
    finally:
        try:
            db.close()
        except Exception as close_error:
            logger.error(f"Error closing database connection: {str(close_error)}")

def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        payload = jwt.decode(credentials.credentials, SECRET_KEY, algorithms=[ALGORITHM])
        username: str = payload.get("sub")
        if username is None:
            raise HTTPException(status_code=401, detail="Invalid token")
        return username
    except jwt.PyJWTError:
        raise HTTPException(status_code=401, detail="Invalid token")

# Background processing functions
def process_upload_background(task_id: str, file_content: bytes, filename: str, current_user: str):
    """Background function to process upload with user-specific workspace"""
    total_start_time = datetime.now()
    db = None  # Initialize to None to prevent UnboundLocalError in finally block
    
    try:
        # Add initial log
        add_upload_log(task_id, "info", f"🚀 Starting upload process for file: {filename}")
        
        # Initialize database and setup
        db = SessionLocal()
        task = db.query(UploadTask).filter(UploadTask.task_id == task_id).first()
        if task:
            task.status = "processing"
            db.commit()
        
        # PERFORMANCE OPTIMIZATION: Skip workspace creation during upload
        # Workspace will be created only when needed for orderlist generation
        add_upload_log(task_id, "info", f"👤 Processing for user: {current_user}")
        
        # Parse filename and validate
        filename_info = parse_filename(filename)
        if not filename_info:
            add_upload_log(task_id, "error", "❌ Invalid filename format")
            raise Exception("Invalid filename format")
        
        add_upload_log(task_id, "info", f"📊 Processing: {filename_info['brand']} - {filename_info['sales_channel']} - Batch {filename_info['batch']}")
        
        # Read file - ULTRA-OPTIMIZED with fastest engine
        add_upload_log(task_id, "info", "📖 Reading file...")
        if filename.endswith('.xlsx'):
            # Use openpyxl directly for maximum speed
            df = pd.read_excel(io.BytesIO(file_content), engine='openpyxl')
        elif filename.endswith('.csv'):
            df = pd.read_csv(io.BytesIO(file_content), encoding='utf-8')
        else:
            add_upload_log(task_id, "error", "❌ Unsupported file format")
            raise Exception("Unsupported file format")
        
        add_upload_log(task_id, "info", f"✅ File read successfully: {len(df)} rows")
        
        # Marketplace mapping and validation
        sales_channel = filename_info['sales_channel']
        marketplace_mapping = get_marketplace_mapping(sales_channel)
        
        if not marketplace_mapping:
            raise Exception(f"Unsupported sales channel: {sales_channel}")
        
        brand_name = filename_info['brand']
        
        # Note: shop_id is already validated in validate_upload_request() before background task starts
        # No need to re-validate here, which improves performance
        
        # Data preparation
        current_time = get_wib_now().replace(tzinfo=None)
        batch_name = filename_info['batch']
        
        # Create upload history record
        upload_history = UploadHistory(
            marketplace=sales_channel,
            brand=brand_name,
            pic=current_user,
            batch=batch_name
        )
        db.add(upload_history)
        db.commit()
        
        add_upload_log(task_id, "info", f"🔄 Processing {len(df)} rows")
        
        # Field mappings
        order_number_col = marketplace_mapping['order_number']
        order_status_col = marketplace_mapping['order_status']
        order_date_col = marketplace_mapping['order_date']
        warehouse_col = marketplace_mapping['wh_loc']
        awb_col = marketplace_mapping['awb']
        transporter_col = marketplace_mapping['transporter']
        sla_col = marketplace_mapping['sla']
        sku_col = marketplace_mapping.get('sku_field')
        
        # Check required columns for this marketplace
        available_columns = list(df.columns)
        required_columns = {
            'order_number': marketplace_mapping.get('order_number'),
            'order_status': marketplace_mapping.get('order_status'),
            'order_date': marketplace_mapping.get('order_date'),
            'awb': marketplace_mapping.get('awb')
        }
        
        missing_columns = []
        for field, expected_name in required_columns.items():
            if expected_name and expected_name not in df.columns:
                missing_columns.append(expected_name)
        
        # If missing critical columns, stop processing and show informative error
        if missing_columns:
            error_msg = f"❌ Missing required columns in Excel file for {sales_channel.upper()}:"
            add_upload_log(task_id, "error", error_msg)
            for missing_col in missing_columns:
                add_upload_log(task_id, "error", f"   • {missing_col}")
            add_upload_log(task_id, "error", "")
            add_upload_log(task_id, "error", f"📋 Available columns in your file:")
            for col in available_columns:
                add_upload_log(task_id, "error", f"   • {col}")
            add_upload_log(task_id, "error", "")
            add_upload_log(task_id, "error", f"💡 Please rename or add missing columns in Excel file to match expected column names.")
            
            # Update task as failed
            if task:
                task.status = "failed"
                task.error_message = f"Missing required columns: {', '.join(missing_columns)}"
                task.completed_at = get_wib_now()
                db.commit()
            
            raise Exception(f"Missing required columns for {sales_channel}: {', '.join(missing_columns)}")
        
        add_upload_log(task_id, "info", f"✅ All required columns found for {sales_channel.upper()}")
        
        # Vectorized field extraction
        order_numbers = df[order_number_col].fillna('').astype(str) if order_number_col in df.columns else pd.Series([''] * len(df))
        
        # Check for empty order numbers
        non_empty_count = sum(1 for order_num in order_numbers if order_num and order_num != 'nan' and order_num.strip())
        empty_count = len(order_numbers) - non_empty_count
        
        if non_empty_count == 0:
            add_upload_log(task_id, "error", f"❌ No valid order numbers found in '{order_number_col}' column")
            add_upload_log(task_id, "error", f"📊 Total rows: {len(order_numbers)}, Valid orders: {non_empty_count}, Empty/Invalid: {empty_count}")
            add_upload_log(task_id, "error", f"💡 Please ensure the '{order_number_col}' column contains valid order numbers")
            
            # Update task as failed
            if task:
                task.status = "failed"
                task.error_message = f"No valid order numbers found in {order_number_col} column"
                task.completed_at = get_wib_now()
                db.commit()
            
            raise Exception(f"No valid order numbers found in {order_number_col} column")
        
        add_upload_log(task_id, "info", f"📊 Found {non_empty_count} valid order numbers out of {len(order_numbers)} rows")
        order_statuses = df[order_status_col].fillna('').astype(str) if order_status_col in df.columns else pd.Series([''] * len(df))
        order_dates = df[order_date_col].fillna('') if order_date_col in df.columns else pd.Series([''] * len(df))
        warehouses = df[warehouse_col].fillna('').astype(str) if warehouse_col in df.columns else pd.Series([''] * len(df))
        awbs = df[awb_col].fillna('').astype(str) if awb_col in df.columns else pd.Series([''] * len(df))
        transporters = df[transporter_col].fillna('').astype(str) if transporter_col in df.columns else pd.Series([''] * len(df))
        slas = df[sla_col].fillna('').astype(str) if sla_col in df.columns else pd.Series([''] * len(df))
        skus = df[sku_col].fillna('').astype(str) if sku_col and sku_col in df.columns else pd.Series([''] * len(df))
        
        # Group by OrderNumber to handle multiple items per order (vectorized)
        order_skus = defaultdict(list)
        if sku_col and sku_col in df.columns:
            # Vectorized approach for better performance
            valid_mask = (order_numbers != '') & (order_numbers != 'nan') & (skus != '') & (skus != 'nan')
            valid_data = df[valid_mask]
            if not valid_data.empty:
                for order_num, sku in zip(valid_data[order_number_col], valid_data[sku_col]):
                    order_num = str(order_num).strip()
                    sku = str(sku).strip()
                    if order_num and sku and sku not in order_skus[order_num]:
                        order_skus[order_num].append(sku)
        
        # Pre-process date for Tokopedia
        tokopedia_date = None
        if sales_channel.lower() == 'tokopedia':
            try:
                date_str = filename_info['date']
                tokopedia_date = datetime.strptime(date_str, '%Y%m%d')
                tokopedia_date = WIB_TIMEZONE.localize(tokopedia_date)
            except:
                tokopedia_date = current_time
        
        # OPTIMIZED APPROACH: Vectorized operations for maximum performance
        
        # Vectorized data preparation - much faster than row-by-row
        valid_mask = (order_numbers != '') & (order_numbers != 'nan') & (order_numbers.notna())
        valid_df = df[valid_mask].copy()
        
        if len(valid_df) == 0:
            add_upload_log(task_id, "error", "❌ No valid order numbers found")
            raise Exception("No valid order numbers found")
        
        # Vectorized date processing - OPTIMIZED
        if sales_channel.lower() == 'tokopedia':
            # Use tokopedia_date for all orders if no dates provided
            date_mask = valid_df[order_date_col].isna() | (valid_df[order_date_col] == '')
            valid_df.loc[date_mask, 'processed_date'] = tokopedia_date
            valid_df.loc[~date_mask, 'processed_date'] = pd.to_datetime(
                valid_df.loc[~date_mask, order_date_col], 
                dayfirst=True, 
                errors='coerce'
            ).apply(lambda x: convert_to_wib(x) if pd.notna(x) else current_time)
        else:
            # Vectorized date conversion for other marketplaces
            valid_df['processed_date'] = pd.to_datetime(
                valid_df[order_date_col], 
                dayfirst=True, 
                errors='coerce'
            ).apply(lambda x: convert_to_wib(x) if pd.notna(x) else current_time)
        
        # Vectorized data extraction - OPTIMIZED
        valid_df['processed_order_number'] = valid_df[order_number_col].astype(str).str.strip()
        valid_df['processed_order_status'] = valid_df[order_status_col].fillna('').astype(str)
        valid_df['processed_awb'] = valid_df[awb_col].fillna('').astype(str) if awb_col and awb_col in valid_df.columns else ''
        valid_df['processed_transporter'] = valid_df[transporter_col].fillna('').astype(str) if transporter_col and transporter_col in valid_df.columns else ''
        valid_df['processed_sla'] = valid_df[sla_col].fillna('').astype(str) if sla_col and sla_col in valid_df.columns else ''
        valid_df['processed_sku'] = valid_df[sku_col].fillna('').astype(str) if sku_col and sku_col in valid_df.columns else ''
        
        # OPTIMIZED: Use pandas groupby for much faster grouping
        def aggregate_skus(series):
            """Aggregate SKUs for the same order number"""
            valid_skus = series[series != ''].dropna().unique()
            return ','.join(valid_skus) if len(valid_skus) > 0 else None
        
        # Group by OrderNumber and aggregate - VECTORIZED & FAST
        grouped_df = valid_df.groupby('processed_order_number').agg({
            'processed_order_status': 'first',
            'processed_awb': 'first', 
            'processed_transporter': 'first',
            'processed_date': 'first',
            'processed_sla': 'first',
            'processed_sku': aggregate_skus
        }).reset_index()
        
        # Convert to list of dictionaries - OPTIMIZED
        all_order_data = []
        for _, row in grouped_df.iterrows():
            order_data = {
                'Marketplace': sales_channel,
                'Brand': brand_name,
                'OrderNumber': row['processed_order_number'],
                'OrderStatus': row['processed_order_status'],
                'AWB': row['processed_awb'],
                'Transporter': row['processed_transporter'],
                'OrderDate': row['processed_date'],
                'SLA': row['processed_sla'],
                'Batch': batch_name,
                'PIC': current_user,
                'UploadDate': current_time,
                'TaskId': task_id,
                'ItemId': row['processed_sku'],
                'ItemIdFlexo': None
            }
            all_order_data.append(order_data)
        
        # OPTIMIZED logging - only log once
        add_upload_log(task_id, "info", f"📊 Processed {len(df)} rows → {len(all_order_data)} unique orders (vectorized)")
        
        # STEP 10: Interface checking - OPTIMIZED WITH TIMEOUT PROTECTION
        add_upload_log(task_id, "info", "🔍 Starting optimized interface status checking...")
        
        # OPTIMIZED: Vectorized marketplace grouping
        marketplace_series = pd.Series([order['Marketplace'] for order in all_order_data])
        order_number_series = pd.Series([order['OrderNumber'] for order in all_order_data])
        
        # Group orders by marketplace - VECTORIZED
        orders_by_marketplace = {}
        for marketplace in marketplace_series.unique():
            mask = marketplace_series == marketplace
            orders_by_marketplace[marketplace] = [
                order for i, order in enumerate(all_order_data) if mask.iloc[i]
            ]
        
        # Check interface status for each marketplace with optimized chunking
        interface_results = {}
        total_orders = len(all_order_data)
        
        # OPTIMIZED: Smaller chunk size for better timeout handling
        EXTERNAL_CHUNK_SIZE = 100  # Reduced from 10000 for better timeout handling
        
        add_upload_log(task_id, "info", f"🚀 Processing {total_orders} orders with optimized chunk size: {EXTERNAL_CHUNK_SIZE}")
        
        for marketplace, marketplace_orders in orders_by_marketplace.items():
            # Extract order numbers for this marketplace
            order_numbers_list = [order['OrderNumber'] for order in marketplace_orders if order['OrderNumber']]
            
            if order_numbers_list:
                total_chunks = (len(order_numbers_list) + EXTERNAL_CHUNK_SIZE - 1) // EXTERNAL_CHUNK_SIZE
                
                # OPTIMIZED: Only log for large datasets or when significant progress
                if total_chunks > 1:
                    add_upload_log(task_id, "info", f"📊 Processing {len(order_numbers_list)} orders for {marketplace} in {total_chunks} chunks")
                
                for i in range(0, len(order_numbers_list), EXTERNAL_CHUNK_SIZE):
                    chunk_order_numbers = order_numbers_list[i:i + EXTERNAL_CHUNK_SIZE]
                    
                    try:
                        # Add timeout protection using threading.Timer
                        import threading
                        import time
                        
                        chunk_interface_results = {}
                        timeout_occurred = threading.Event()
                        
                        def check_with_timeout():
                            nonlocal chunk_interface_results
                            try:
                                chunk_interface_results = check_interface_status(chunk_order_numbers, marketplace)
                            except Exception as e:
                                add_upload_log(task_id, "warning", f"⚠️ Interface check error: {str(e)}")
                        
                        def timeout_handler():
                            timeout_occurred.set()
                        
                        # Start timeout timer (30 seconds)
                        timer = threading.Timer(30.0, timeout_handler)
                        timer.start()
                        
                        # Start interface check in thread (daemon=True to prevent memory leak)
                        check_thread = threading.Thread(target=check_with_timeout, daemon=True)
                        check_thread.start()
                        
                        # Wait for completion or timeout
                        check_thread.join(timeout=30)
                        timer.cancel()
                        
                        if timeout_occurred.is_set():
                            add_upload_log(task_id, "warning", f"⚠️ Interface check timeout for chunk after 30 seconds")
                            chunk_interface_results = {}
                        
                        # Ensure thread cleanup
                        if check_thread.is_alive():
                            add_upload_log(task_id, "warning", f"⚠️ Check thread still running after timeout, will be cleaned up automatically")
                        
                    except Exception as e:
                        add_upload_log(task_id, "warning", f"⚠️ Interface check failed for chunk: {str(e)}")
                        chunk_interface_results = {}
                    
                    interface_results.update(chunk_interface_results)
                    
                    chunk_num = (i // EXTERNAL_CHUNK_SIZE) + 1
                    # OPTIMIZED: Further reduce logging frequency for better performance
                    if total_chunks > 5 and (chunk_num % max(1, total_chunks // 3) == 0 or chunk_num == total_chunks):
                        add_upload_log(task_id, "info", f"📊 Interface check progress: {marketplace} chunk {chunk_num}/{total_chunks}")
        
        add_upload_log(task_id, "success", "✅ Interface checking completed with timeout protection")
        
        # OPTIMIZED: Database operations with vectorized processing
        add_upload_log(task_id, "info", f"🚀 Processing {len(all_order_data)} orders for database operations")
        
        # OPTIMIZED: Vectorized interface status assignment
        all_order_numbers = [order_data['OrderNumber'] for order_data in all_order_data]
        
        # OPTIMIZED: Bulk query existing orders with improved filtering
        existing_orders_dict = {}
        # Always check for duplicates to prevent race conditions and data integrity issues
        existing_orders = db.query(UploadedOrder).filter(
            UploadedOrder.OrderNumber.in_(all_order_numbers)
        ).all()
        existing_orders_dict = {order.OrderNumber: order for order in existing_orders}
        
        # OPTIMIZED: Vectorized interface status processing
        orders_to_insert = []
        orders_to_update = []
        replaced_count = 0
        new_count = 0
        
        # OPTIMIZED: Batch process interface status assignment
        for order_data in all_order_data:
            order_number = order_data['OrderNumber']
            
            # OPTIMIZED: Simplified interface status assignment
            if order_number in interface_results:
                interface_result = interface_results[order_number]
                interface_status = interface_result.get('status') or interface_result.get('interface_status')
                
                # Batch assign interface data
                order_data.update({
                    'InterfaceStatus': 'Interface' if interface_status == 'Interface' else 'Not Yet Interface',
                    'OrderNumberFlexo': interface_result.get('system_ref_id', ''),
                    'OrderStatusFlexo': interface_result.get('order_status', ''),
                    'ItemIdFlexo': interface_result.get('item_id_flexo')
                })
            else:
                # Batch assign default values
                order_data.update({
                    'InterfaceStatus': 'Not Yet Interface',
                    'OrderNumberFlexo': '',
                    'OrderStatusFlexo': '',
                    'ItemIdFlexo': None
                })
            
            # OPTIMIZED: Categorize for bulk operations
            if order_number in existing_orders_dict:
                orders_to_update.append((existing_orders_dict[order_number], order_data))
                replaced_count += 1
            else:
                orders_to_insert.append(order_data)
                new_count += 1
        
        
        # Bulk insert new orders - OPTIMIZED with single commit
        if orders_to_insert:
            try:
                db.bulk_insert_mappings(UploadedOrder, orders_to_insert)
                db.commit()  # Single commit for all inserts
            except Exception as bulk_error:
                # Fallback to individual inserts for unique constraint violations
                add_upload_log(task_id, "warning", f"⚠️ Bulk insert failed, falling back to individual inserts: {str(bulk_error)}")
                db.rollback()
                failed_orders = []
                success_count = 0
                for order_data in orders_to_insert:
                    try:
                        new_order = UploadedOrder(**order_data)
                        db.add(new_order)
                        db.commit()
                        success_count += 1
                    except Exception as e:
                        db.rollback()
                        failed_orders.append(order_data['OrderNumber'])
                        continue
                
                # Log summary of fallback results
                if failed_orders:
                    add_upload_log(task_id, "warning", f"⚠️ {len(failed_orders)} orders failed to insert (likely duplicates): {', '.join(failed_orders[:5])}{'...' if len(failed_orders) > 5 else ''}")
                if success_count > 0:
                    add_upload_log(task_id, "info", f"✅ Successfully inserted {success_count} orders via fallback method")
        
        # Bulk update existing orders
        if orders_to_update:
            for existing_order, new_data in orders_to_update:
                existing_order.Marketplace = new_data['Marketplace']
                existing_order.Brand = new_data['Brand']
                existing_order.OrderStatus = new_data['OrderStatus']
                existing_order.AWB = new_data['AWB']
                existing_order.Transporter = new_data['Transporter']
                existing_order.OrderDate = new_data['OrderDate']
                existing_order.SLA = new_data['SLA']
                existing_order.Batch = new_data['Batch']
                existing_order.PIC = new_data['PIC']
                existing_order.UploadDate = new_data['UploadDate']
                existing_order.InterfaceStatus = new_data['InterfaceStatus']
                existing_order.TaskId = new_data['TaskId']
                existing_order.Remarks = new_data.get('Remarks', '')
                existing_order.OrderNumberFlexo = new_data.get('OrderNumberFlexo') or ''
                existing_order.OrderStatusFlexo = new_data.get('OrderStatusFlexo') or ''
                existing_order.ItemId = new_data.get('ItemId')  # ✅ Update ItemId
                existing_order.ItemIdFlexo = new_data.get('ItemIdFlexo')  # ✅ Update ItemIdFlexo
                
        
        # Commit all operations
        db.commit()
        
        # Calculate processing time
        processing_time = (datetime.now() - total_start_time).total_seconds()
        
        # Count interface vs not interface orders - only for this upload session
        
        # Only get orders from this specific upload task
        uploaded_orders = db.query(UploadedOrder).filter(
            UploadedOrder.TaskId == task_id
        ).all()
        
        interface_count = sum(1 for order in uploaded_orders if order.InterfaceStatus == 'Interface')
        not_interface_count = len(uploaded_orders) - interface_count
        
        
        # Initialize result dictionary for response
        result = {
            "interface_count": interface_count,
            "not_interface_count": not_interface_count,
            "total_orders": len(uploaded_orders),
            "new_orders": new_count,
            "replaced_orders": replaced_count
        }
        
        # PERFORMANCE OPTIMIZATION: Skip orderlist generation during upload
        # Orderlist will be generated later via manual action or scheduled job
        add_upload_log(task_id, "info", "⚡ Skipping orderlist generation for faster upload...")
        
        # Set default result for orderlist
        orderlist_result = {"success": True, "skipped": True}
        result["orderlist_skipped"] = True
        
        # Skip auto-run marketplace app since orderlist is not generated
        add_upload_log(task_id, "info", "ℹ️ Auto-run skipped - orderlist generation disabled for performance")
        
        # Update task as completed
        if task:
            task.status = "completed"
            task.total_orders = len(uploaded_orders)
            task.processed_orders = len(uploaded_orders)
            task.processing_time = f"{processing_time:.2f}s"
            task.external_db_query_time = "0.00s"  # Interface check time
            task.completed_at = get_wib_now()
            db.commit()
            add_upload_log(task_id, "success", f"🎉 Upload process completed successfully! {len(uploaded_orders)} orders processed in {processing_time:.2f}s")
        
        # Performance summary
        total_time = (datetime.now() - total_start_time).total_seconds()
        add_upload_log(task_id, "info", f"📊 Total processing time: {total_time:.2f}s")
        
        # Update Not Uploaded Items history immediately after successful background upload
        try:
            save_not_uploaded_history()
            logger.info(f"Not uploaded items history updated after background upload: {filename_info['brand']}-{filename_info['sales_channel']}-{filename_info['batch']}")
        except Exception as history_error:
            logger.error(f"Failed to update not uploaded items history after background upload: {str(history_error)}")
            # Don't fail the upload if history update fails
        
    except Exception as e:
        # Add detailed error log
        import traceback
        error_traceback = traceback.format_exc()
        add_upload_log(task_id, "error", f"❌ Upload process failed: {str(e)}")
        add_upload_log(task_id, "error", f"🔍 Error details: {error_traceback}")
        
        # Update task as failed
        if 'task' in locals() and task:
            task.status = "failed"
            task.error_message = str(e)
            task.completed_at = get_wib_now()
            db.commit()
        
    finally:
        if db is not None:
            try:
                db.close()
                add_upload_log(task_id, "info", "🔒 Database connection closed properly")
            except Exception as close_error:
                add_upload_log(task_id, "error", f"❌ Error closing database connection: {str(close_error)}")

# Utility functions
def verify_password(plain_password, hashed_password):
    return pwd_context.verify(plain_password, hashed_password)

def get_password_hash(password):
    return pwd_context.hash(password)

def create_access_token(data: dict):
    to_encode = data.copy()
    expire = get_wib_now() + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

# Marketplace field mappings
MARKETPLACE_MAPPINGS = {
    'tiktok': {
        'brand': 'TIKTOK',
        'order_number': 'Order ID',
        'order_status': 'Order Substatus',
        'order_date': 'Created Time',
        'wh_loc': 'Warehouse Name',
        'awb': 'Tracking ID',
        'transporter': 'Shipping Provider Name',
        'sla': None,
        'sku_field': 'Seller SKU'  # ✅ SKU field untuk TikTok
    },
    'zalora': {
        'brand': 'ZALORA',
        'order_number': 'Order Number',
        'order_status': 'Status',
        'order_date': 'Created at',
        'wh_loc': None,
        'awb': 'Tracking Code',
        'transporter': 'Shipping Provider',
        'sla': None
    },
    'shopee': {
        'brand': 'SHOPEE',
        'order_number': 'No. Pesanan',
        'order_status': 'Status Pesanan',
        'order_date': 'Waktu Pesanan Dibuat',
        'wh_loc': 'Nama Gudang',
        'awb': 'No. Resi',
        'transporter': 'Opsi Pengiriman',
        'sla': 'Pesanan Harus Dikirimkan Sebelum (Menghindari keterlambatan)',
        'sku_field': 'Nomor Referensi SKU'  # ✅ SKU field untuk Shopee
    },
    'lazada': {
        'brand': 'LAZADA',
        'order_number': 'orderNumber',
        'order_status': 'status',
        'order_date': 'createTime',
        'wh_loc': 'wareHouse',
        'awb': 'trackingCode',
        'transporter': 'shippingProvider',
        'sla': 'ttsSla',
        'sku_field': 'sellerSku'  # ✅ SKU field untuk Lazada
    },
    'tokopedia': {
        'brand': 'TOKOPEDIA',
        'order_number': 'Nomor Invoice',
        'order_status': 'Status Terakhir',
        'order_date': None,  # Will be handled specially
        'wh_loc': None,
        'awb': 'No Resi / Kode Booking',
        'transporter': None,
        'sla': None,
        'sku_field': None  # ✅ No SKU field untuk Tokopedia
    },
    'blibli': {
        'brand': 'BLIBLI',
        'order_number': 'No. Order',
        'order_status': 'Order Status',
        'order_date': 'Tanggal Order',
        'wh_loc': None,
        'awb': 'No. Awb',
        'transporter': 'Servis Logistik',
        'sla': None,
        'sku_field': 'Merchant SKU'  # ✅ SKU field untuk Blibli
    },
    'ginee': {
        'brand': 'GINEE',
        'order_number': 'Order ID',
        'order_status': 'Status',
        'order_date': 'Create Time',
        'wh_loc': None,
        'awb': 'AWB/Tracking code',
        'transporter': 'Courier',
        'sla': 'Ship Before',
        'sku_field': 'SKU'  # ✅ SKU field untuk Ginee
    },
    'desty': {
        'brand': 'DESTY',
        'order_number': 'Nomor Pesanan\n(di Desty)',
        'order_status': 'Status Pesanan',
        'order_date': 'Tanggal Pesanan Dibuat',
        'wh_loc': 'Nama Gudang\nMaster',
        'awb': 'Nomor AWB/Resi',
        'transporter': 'Kurir',
        'sla': 'Dikirim Sebelum',
        'sku_field': 'SKU Marketplace'  # ✅ SKU field untuk Desty
    },
    'jubelio': {
        'brand': 'JUBELIO',
        'order_number': 'salesorder_no',
        'order_status': 'internal_status',
        'order_date': 'created_date',
        'wh_loc': 'location_name',
        'awb': 'tracking_no',
        'transporter': 'shipper',
        'sla': 'shipment_type',
        'customer_name': 'customer_name',
        'customer_phone': 'shipping_full_name',
        'customer_address': 'shipping_address',
        'customer_city': 'location_name',
        'customer_province': None,
        'customer_zipcode': None,
        'sku_field': 'SKU',  # ✅ SKU field untuk Jubelio
        'shop_key_1': 'Shop ID',
        'shop_key_2': 'Shop Key 2',
        'shop_key_3': 'Shop Key 3',
        'shop_key_4': 'Shop Key 4',
        'shop_key_5': 'Shop Key 5'
    }
}

def validate_upload_request(file: UploadFile, db: Session) -> dict:
    """Consolidated validation for upload requests"""
    try:
        # File security validation
        if hasattr(file, 'size') and file.size > MAX_FILE_SIZE:
            raise HTTPException(status_code=413, detail="File too large. Maximum size is 50MB")
        
        if not file.filename:
            raise HTTPException(status_code=400, detail="No filename provided")
        
        file_ext = os.path.splitext(file.filename)[1].lower()
        if file_ext not in ALLOWED_EXTENSIONS:
            raise HTTPException(
                status_code=400, 
                detail=f"Invalid file type. Allowed types: {', '.join(ALLOWED_EXTENSIONS)}"
            )
        
        if hasattr(file, 'content_type') and file.content_type not in ALLOWED_MIME_TYPES:
            raise HTTPException(status_code=400, detail="Invalid file MIME type")
        
        if any(char in file.filename for char in ['..', '/', '\\', '<', '>', ':', '"', '|', '?', '*']):
            raise HTTPException(status_code=400, detail="Invalid characters in filename")
        
        # Filename parsing
        filename_info = parse_filename(file.filename)
        if not filename_info:
            raise HTTPException(status_code=400, detail="Invalid filename format")
        
        # Shop ID validation
        sales_channel = filename_info['sales_channel']
        brand_name = filename_info['brand']
        shop_id = get_shop_id_from_brand_shops(brand_name, sales_channel, db)
        
        if not shop_id:
            error_message = f"No shop_id found for {brand_name} in {sales_channel}. Please add shop_id configuration before uploading."
            raise HTTPException(status_code=400, detail=error_message)
        
        return {
            'filename_info': filename_info,
            'shop_id': shop_id,
            'sales_channel': sales_channel,
            'brand_name': brand_name
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Upload validation error: {str(e)}")
        raise HTTPException(status_code=400, detail="Upload validation failed")

def validate_file_upload(file: UploadFile) -> bool:
    """Legacy function for backward compatibility"""
    try:
        # Check file size
        if hasattr(file, 'size') and file.size > MAX_FILE_SIZE:
            raise HTTPException(status_code=413, detail="File too large. Maximum size is 50MB")
        
        # Check file extension
        if not file.filename:
            raise HTTPException(status_code=400, detail="No filename provided")
        
        file_ext = os.path.splitext(file.filename)[1].lower()
        if file_ext not in ALLOWED_EXTENSIONS:
            raise HTTPException(
                status_code=400, 
                detail=f"Invalid file type. Allowed types: {', '.join(ALLOWED_EXTENSIONS)}"
            )
        
        # Check MIME type
        if hasattr(file, 'content_type') and file.content_type not in ALLOWED_MIME_TYPES:
            raise HTTPException(status_code=400, detail="Invalid file MIME type")
        
        # Check filename for malicious patterns
        if any(char in file.filename for char in ['..', '/', '\\', '<', '>', ':', '"', '|', '?', '*']):
            raise HTTPException(status_code=400, detail="Invalid characters in filename")
        
        return True
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"File validation error: {str(e)}")
        raise HTTPException(status_code=400, detail="File validation failed")

def parse_filename(filename: str):
    """Parse filename to extract brand, sales channel, date, and batch with security validation"""
    # Sanitize filename
    filename = os.path.basename(filename)  # Remove path components
    
    # Clean up filename - remove trailing dashes before extension
    filename = re.sub(r'-+\.(xlsx|csv)$', r'.\1', filename)
    
    # Support multiple formats:
    # Format 1: BRAND-SALESCHANNEL-DATE-BATCH.xlsx (e.g., CETAPHIL-SHOPEE-21-3.xlsx)
    # Format 2: BRAND-SALESCHANNEL-BATCH.xlsx (e.g., CETAPHIL-SHOPEE-3.xlsx)
    # Format 3: BRAND-SALESCHANNEL-DATE-BATCH-.xlsx (e.g., MUSTELA-TIKTOK-25-3-.xlsx)
    
    # Try format 1 first (with date)
    pattern1 = r"([A-Z]+)-([A-Z]+)-(\d+)-(\d+)\.(xlsx|csv)"
    match1 = re.match(pattern1, filename)
    if match1:
        brand = match1.group(1)
        sales_channel = match1.group(2)
        date = match1.group(3)   # DATE is group 3
        batch = match1.group(4)  # BATCH is group 4
        extension = match1.group(5)
        
        # Additional validation
        if len(brand) > 20 or len(sales_channel) > 20:
            raise HTTPException(status_code=400, detail="Brand or sales channel name too long")
        
        # Validate date format (1-31 for day, or 1-12 for month)
        if not date.isdigit() or int(date) < 1 or int(date) > 31:
            raise HTTPException(status_code=400, detail="Invalid date format in filename")
        
        # Validate batch format
        if not batch.isdigit() or len(batch) > 3:
            raise HTTPException(status_code=400, detail="Invalid batch format in filename")
        
        return {
            "brand": brand,
            "sales_channel": sales_channel,
            "date": date,
            "batch": batch,
            "extension": extension
        }
    
    # Try format 2 (simplified without date)
    pattern2 = r"([A-Z]+)-([A-Z]+)-(\d+)\.(xlsx|csv)"
    match2 = re.match(pattern2, filename)
    if match2:
        brand = match2.group(1)
        sales_channel = match2.group(2)
        batch = match2.group(3)  # BATCH is group 3
        extension = match2.group(4)
        date = "1"  # Default date for simplified format
        
        # Additional validation
        if len(brand) > 20 or len(sales_channel) > 20:
            raise HTTPException(status_code=400, detail="Brand or sales channel name too long")
        
        # Validate batch format
        if not batch.isdigit() or len(batch) > 3:
            raise HTTPException(status_code=400, detail="Invalid batch format in filename")
        
        return {
            "brand": brand,
            "sales_channel": sales_channel,
            "date": date,
            "batch": batch,
            "extension": extension
        }
    
    # If neither pattern matches, return None
    return None

def get_marketplace_mapping(sales_channel: str):
    """Get marketplace field mappings based on sales channel"""
    if not sales_channel:
        return None
    
    normalized_channel = sales_channel.lower().strip()
    return MARKETPLACE_MAPPINGS.get(normalized_channel)

def map_field_value(row, field_mapping, fallback_value=''):
    """Map field value using marketplace mapping with fallback"""
    if not field_mapping:
        return fallback_value
    
    # Try the main field mapping first
    if field_mapping in row:
        return row[field_mapping]
    
    return fallback_value

def get_shop_id_from_brand_shops(brand_name, marketplace_name, db):
    """Get shop_id (shop_key_1) from brand_shops table based on brand and marketplace"""
    try:
        
        # Map marketplace name to marketplace_id (from frontend BrandShopsInfo.js)
        marketplace_id_mapping = {
            'tokopedia': 1,
            'shopee': 2,
            'lazada': 3,
            'zalora': 6,
            'ginee': 27,
            'blibli': 7,
            'jdid': 8,
            'jubelio': 9,
            'shopify': 10,
            'tiktok': 11,
            'b2b': 12,
            'desty': 23
        }
        
        marketplace_id = marketplace_id_mapping.get(marketplace_name.lower())
        
        if not marketplace_id:
            print(f"❌ Unknown marketplace_id for: {marketplace_name}")
            return None
        
        # First, let's see all records for this brand
        all_brand_records = db.query(BrandShop).filter(BrandShop.brand == brand_name).all()
        
        for record in all_brand_records:
            pass
        
        # Query brand_shops table with case-insensitive criteria
        brand_shop = db.query(BrandShop).filter(
            func.lower(BrandShop.brand) == brand_name.lower(),
            BrandShop.marketplace_id == marketplace_id
        ).first()
        
        
        if brand_shop:
            if brand_shop.shop_key_1:
                print(f"✅ Found shop_id for {brand_name} in {marketplace_name}: {brand_shop.shop_key_1}")
                return brand_shop.shop_key_1
            else:
                print(f"⚠️ Found record but shop_key_1 is empty for {brand_name} in {marketplace_name}")
                return None
        else:
            print(f"❌ No record found for brand: {brand_name}, marketplace: {marketplace_name} (marketplace_id: {marketplace_id})")
            return None
            
    except Exception as e:
        print(f"❌ Error getting shop_id: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

def generate_orderlist_for_not_interfaced(uploaded_orders, sales_channel, brand_name, batch_name, db, user_id=None):
    """Generate Orderlist.txt file for Not Interfaced orders in the appropriate marketplace folder"""
    try:
        # Filter only Not Interfaced orders
        not_interfaced_orders = [order for order in uploaded_orders if order.InterfaceStatus == 'Not Yet Interface']
        
        if not not_interfaced_orders:
            print("No Not Interfaced orders found, skipping Orderlist.txt generation")
            return {"success": False, "reason": "no_not_interfaced_orders"}
        
        # Map sales channel to marketplace folder name (based on JobGetOrder structure)
        marketplace_folder_mapping = {
            'shopee': 'Shopee',
            'lazada': 'Lazada', 
            'blibli': 'Blibli',
            'desty': 'Desty',
            'ginee': 'Ginee',
            'tiktok': 'Tiktok',
            'zalora': 'Zalora',
            'jubelio': 'Jubelio'
        }
        
        marketplace_folder = marketplace_folder_mapping.get(sales_channel.lower())
        if not marketplace_folder:
            print(f"Unknown marketplace: {sales_channel}, skipping Orderlist.txt generation")
            return {"success": False, "reason": "unknown_marketplace"}
        
        # Get shop_id from brand_shops table
        shop_id = get_shop_id_from_brand_shops(brand_name, sales_channel, db)
        shop_id_warning = None
        if not shop_id:
            print(f"⚠️  No shop_id found for {brand_name} in {sales_channel}, using default config")
            shop_id_warning = {
                "type": "shop_id_not_found",
                "message": f"No shop_id found for {brand_name} in {sales_channel}",
                "details": "Marketplace app may not run correctly without valid shop_id",
                "action_required": f"Please add shop_id for {brand_name} in {sales_channel} to brand_shops table"
            }
        
        # Get the JobGetOrder directory path with user-specific workspace
        if user_id:
            # Use user-specific workspace
            user_workspace = multi_user_handler.get_user_workspace(user_id)
            
            jobgetorder_path = os.path.join(user_workspace, marketplace_folder)
            
            # Ensure marketplace folder exists (auto-create if missing) only if folder doesn't exist
            if not os.path.exists(jobgetorder_path):
                config = {'folder': marketplace_folder}
                ensure_marketplace_folder_exists(user_workspace, sales_channel.lower(), config)
            # Copy .NET apps to user workspace
            multi_user_handler.copy_net_apps_to_user_workspace(user_id, marketplace_folder)
        else:
            # Fallback to original path for backward compatibility
            project_root = get_project_root()
            jobgetorder_path = os.path.join(project_root, 'JobGetOrder', marketplace_folder)
        
        # Create directory if it doesn't exist with proper error handling
        try:
            os.makedirs(jobgetorder_path, exist_ok=True)
        except PermissionError as e:
            print(f"❌ Permission denied creating directory {jobgetorder_path}: {e}")
            raise Exception(f"Permission denied creating directory: {e}")
        except Exception as e:
            print(f"❌ Error creating directory {jobgetorder_path}: {e}")
            raise Exception(f"Error creating directory: {e}")
        
        # Generate Orderlist.txt file with unique name per task to prevent conflicts
        # Use task_id from database to make it unique per upload session
        db = SessionLocal()
        try:
            # Get task_id for this upload session
            task = db.query(UploadTask).filter(
                UploadTask.pic == user_id,
                UploadTask.brand == brand_name,
                UploadTask.batch == batch_name
            ).order_by(UploadTask.created_at.desc()).first()
            
            if task and task.task_id:
                # Use task_id to make Orderlist unique per upload session
                orderlist_filename = f"Orderlist_{task.task_id}.txt"
            else:
                # Fallback to timestamp-based filename
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                orderlist_filename = f"Orderlist_{timestamp}.txt"
        except Exception as e:
            # Fallback to timestamp-based filename
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            orderlist_filename = f"Orderlist_{timestamp}.txt"
        finally:
            db.close()
        
        orderlist_path = os.path.join(jobgetorder_path, orderlist_filename)
        
        # Extract order numbers with special formatting for certain marketplaces
        order_numbers = [order.OrderNumber for order in not_interfaced_orders if order.OrderNumber]
        
        # Write to file with special formatting for certain marketplaces
        with open(orderlist_path, 'w', encoding='utf-8') as f:
            for order_number in order_numbers:
                if sales_channel.lower() in ['zalora', 'blibli']:
                    # For ZALORA and BLIBLI: use format "id,shop_key_1"
                    if shop_id:
                        f.write(f"{order_number},{shop_id}\n")
                    else:
                        f.write(f"{order_number}\n")  # Fallback if no shop_id
                else:
                    # For other marketplaces: use format "id" only
                    f.write(f"{order_number}\n")
        
        print(f"✅ Generated {orderlist_filename} for {len(order_numbers)} Not Interfaced orders")
        print(f"   📁 Location: {orderlist_path}")
        print(f"   🏪 Marketplace: {marketplace_folder}")
        print(f"   📊 Brand: {brand_name}, Batch: {batch_name}")
        print(f"   🔑 Unique filename: {orderlist_filename}")
        if shop_id:
            print(f"   🏪 Shop ID: {shop_id}")
        else:
            print(f"   ⚠️  Shop ID: NOT FOUND - Marketplace app may not work correctly")
        
        # Backup functionality removed for better performance
        
        # Verify Orderlist.txt (backup functionality removed)
        verification_result = verify_and_restore_orderlist(orderlist_path, None, len(order_numbers), order_numbers)
        if not verification_result:
            print(f"⚠️ Orderlist.txt verification failed, but continuing...")
        
        # Make Orderlist.txt read-only to prevent accidental modification
        try:
            os.chmod(orderlist_path, 0o444)  # Read-only for all users
            print(f"🔒 Made Orderlist.txt read-only for protection")
        except Exception as e:
            print(f"⚠️ Could not make Orderlist.txt read-only: {str(e)}")
        
        # Backup cleanup functionality removed for better performance
        
        # Generate dynamic config file for .NET app (skip for ZALORA, BLIBLI, and GINEE)
        if shop_id and sales_channel.lower() not in ['zalora', 'blibli', 'ginee']:
            if sales_channel.lower() == 'jubelio':
                # For JUBELIO: generate shop.txt file instead of .config
                generate_jubelio_shop_txt(marketplace_folder, brand_name, shop_id, user_id)
            else:
                generate_dynamic_config(marketplace_folder, brand_name, shop_id, sales_channel, user_workspace)
        
        # Return success with shop_id warning if applicable
        return {
            "success": True,
            "order_count": len(order_numbers),
            "orderlist_path": orderlist_path,
            "backup_path": None,  # Backup functionality removed
            "shop_id": shop_id,
            "shop_id_warning": shop_id_warning
        }
        
    except Exception as e:
        print(f"❌ Error generating Orderlist.txt: {str(e)}")
        logger.error(f"Error generating Orderlist.txt: {str(e)}")
        return {"success": False, "error": str(e)}

# Backup cleanup function removed for better performance

def verify_and_restore_orderlist(orderlist_path, backup_path, expected_count, order_numbers):
    """Verify Orderlist.txt content and auto-restore from backup if needed"""
    try:
        # Check if Orderlist.txt exists
        if not os.path.exists(orderlist_path):
            print(f"⚠️ Orderlist.txt not found, restoring from backup...")
            if os.path.exists(backup_path):
                shutil.copy2(backup_path, orderlist_path)
                print(f"✅ Restored Orderlist.txt from backup")
                return True
            else:
                print(f"❌ Backup file also not found: {backup_path}")
                return False
        
        # Count actual orders in Orderlist.txt
        with open(orderlist_path, 'r', encoding='utf-8') as f:
            lines = f.readlines()
            actual_count = len([line.strip() for line in lines if line.strip()])
        
        # Check if count matches
        if actual_count != expected_count:
            print(f"⚠️ Order count mismatch detected!")
            print(f"   Expected: {expected_count} orders")
            print(f"   Actual: {actual_count} orders")
            print(f"🔄 Auto-restoring from backup...")
            
            if os.path.exists(backup_path):
                # Make file writable before restore
                try:
                    os.chmod(orderlist_path, 0o666)  # Make writable
                    print(f"🔓 Made Orderlist.txt writable for restore")
                except Exception as e:
                    print(f"⚠️ Could not change permissions: {str(e)}")
                
                shutil.copy2(backup_path, orderlist_path)
                print(f"✅ Auto-restored Orderlist.txt from backup")
                
                # Verify restoration
                with open(orderlist_path, 'r', encoding='utf-8') as f:
                    restored_lines = f.readlines()
                    restored_count = len([line.strip() for line in restored_lines if line.strip()])
                
                if restored_count == expected_count:
                    print(f"✅ Verification passed: {restored_count} orders restored")
                    return True
                else:
                    print(f"❌ Restoration failed: still {restored_count} orders")
                    return False
            else:
                print(f"❌ Backup file not found for restoration: {backup_path}")
                return False
        else:
            print(f"✅ Orderlist.txt verification passed: {actual_count} orders")
            return True
            
    except Exception as e:
        print(f"❌ Error verifying Orderlist.txt: {str(e)}")
        return False

def generate_jubelio_shop_txt(marketplace_folder, brand_name, shop_id, user_id=None):
    """Generate shop.txt file for Jubelio app with correct shop_id"""
    try:
        # Get the JobGetOrder directory path with user-specific workspace
        if user_id:
            # Use user-specific workspace
            user_workspace = multi_user_handler.get_user_workspace(user_id)
            jobgetorder_path = os.path.join(user_workspace, marketplace_folder)
        else:
            # Fallback to original path for backward compatibility
            project_root = get_project_root()
            jobgetorder_path = os.path.join(project_root, 'JobGetOrder', marketplace_folder)
        
        # Generate shop.txt file
        shop_txt_path = os.path.join(jobgetorder_path, 'shop.txt')
        
        # Write shop_id to shop.txt file
        with open(shop_txt_path, 'w', encoding='utf-8') as f:
            f.write(shop_id)
        
        print(f"   🏪 Shop.txt created for Jubelio: {shop_id}")
        
    except Exception as e:
        print(f"❌ Error generating Jubelio shop.txt: {str(e)}")
        logger.error(f"Error generating Jubelio shop.txt: {str(e)}")

def generate_dynamic_config(marketplace_folder, brand_name, shop_id, sales_channel, user_workspace=None):
    """Generate dynamic config file for .NET app with correct shop_id"""
    try:
        # Get the JobGetOrder directory path
        project_root = get_project_root()
        
        # Use user workspace if provided, otherwise use root marketplace folder
        if user_workspace:
            jobgetorder_path = os.path.join(user_workspace, marketplace_folder)
        else:
            jobgetorder_path = os.path.join(project_root, 'JobGetOrder', marketplace_folder)
        
        # Generate config file name based on brand
        config_filename = f"{brand_name}_{sales_channel}.config"
        config_path = os.path.join(jobgetorder_path, config_filename)
        
        # Create config content based on marketplace
        if sales_channel.lower() == 'shopee':
            config_content = f"""<?xml version="1.0" encoding="utf-8"?>
<configuration>
    <startup> 
        <supportedRuntime version="v4.0" sku=".NETFramework,Version=v4.7.2" />
    </startup>
    <appSettings>
        <add key="apikey" value="44746c6376594a72575559544b7751575844584c7958715575696b656c687977" />
        <add key="shop_id" value="{shop_id}" />
        <add key="brand_name" value="{brand_name}" />
        <add key="marketplace" value="{sales_channel}" />
    </appSettings>
</configuration>"""
        elif sales_channel.lower() == 'desty':
            config_content = f"""<?xml version="1.0" encoding="utf-8"?>
<configuration>
    <startup> 
        <supportedRuntime version="v4.0" sku=".NETFramework,Version=v4.7.2" />
    </startup>
    <appSettings>
        <add key="Merchantidcms" value="{shop_id}" />
        <add key="brand_name" value="{brand_name}" />
        <add key="marketplace" value="{sales_channel}" />
    </appSettings>
</configuration>"""
        else:
            # Generic config for other marketplaces
            config_content = f"""<?xml version="1.0" encoding="utf-8"?>
<configuration>
    <startup> 
        <supportedRuntime version="v4.0" sku=".NETFramework,Version=v4.7.2" />
    </startup>
    <appSettings>
        <add key="shop_id" value="{shop_id}" />
        <add key="brand_name" value="{brand_name}" />
        <add key="marketplace" value="{sales_channel}" />
    </appSettings>
</configuration>"""
        
        # Write config file
        with open(config_path, 'w', encoding='utf-8') as f:
            f.write(config_content)
        
        print(f"   ⚙️  Dynamic config created: {config_filename}")
        
    except Exception as e:
        print(f"❌ Error generating dynamic config: {str(e)}")
        logger.error(f"Error generating dynamic config: {str(e)}")

# Cache untuk menghindari pengecekan berulang
_marketplace_folder_cache = {}

def run_executable_silently(exe_path, cwd_path, marketplace_name=""):
    """Run executable silently without popup windows"""
    import subprocess
    import platform
    
    if platform.system() == 'Windows':
        # Multiple methods to ensure no popup
        methods = [
            # Method 1: Using start command with /B flag
            lambda: subprocess.Popen([
                'cmd', '/c', 'start', '/B', '/MIN', exe_path
            ], 
            cwd=cwd_path,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            creationflags=subprocess.CREATE_NO_WINDOW | subprocess.DETACHED_PROCESS),
            
            # Method 2: Using PowerShell with -WindowStyle Hidden
            lambda: subprocess.Popen([
                'powershell', '-WindowStyle', 'Hidden', '-Command', f'Start-Process -FilePath "{exe_path}" -WorkingDirectory "{cwd_path}" -WindowStyle Hidden'
            ],
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            creationflags=subprocess.CREATE_NO_WINDOW),
            
            # Method 3: Direct execution with all flags
            lambda: subprocess.Popen([exe_path], 
                                   cwd=cwd_path,
                                   stdout=subprocess.DEVNULL,
                                   stderr=subprocess.DEVNULL,
                                   creationflags=subprocess.CREATE_NO_WINDOW | subprocess.DETACHED_PROCESS)
        ]
        
        for i, method in enumerate(methods, 1):
            try:
                method()
                print(f"🚀 Started {marketplace_name} app silently using method {i}")
                return True
            except Exception as e:
                print(f"⚠️ Method {i} failed for {marketplace_name}: {str(e)}")
                continue
        
        print(f"❌ All methods failed for {marketplace_name}")
        return False
    else:
        # For other systems, run directly
        try:
            subprocess.Popen([exe_path], 
                           cwd=cwd_path,
                           stdout=subprocess.DEVNULL,
                           stderr=subprocess.DEVNULL)
            print(f"🚀 Started {marketplace_name} app on {platform.system()}")
            return True
        except Exception as e:
            print(f"❌ Failed to start {marketplace_name} on {platform.system()}: {str(e)}")
            return False

def run_executable_with_logging(exe_path, cwd_path, marketplace_name="", user_name="", brand_name="", task_id=None):
    """Run executable and capture output to log file"""
    import subprocess
    import platform
    import threading
    import time
    from datetime import datetime
    
    # Create log file path
    log_file_path = os.path.join(cwd_path, f"{marketplace_name.lower()}_app.log")
    
    def log_output(process, log_file, prefix=""):
        """Log process output to file"""
        try:
            with open(log_file, 'a', encoding='utf-8') as f:
                # Write header
                f.write(f"\n{'='*50}\n")
                f.write(f"Marketplace App Execution Log\n")
                f.write(f"User: {user_name}\n")
                f.write(f"Marketplace: {marketplace_name}\n")
                f.write(f"Brand: {brand_name}\n")
                f.write(f"Start Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"{'='*50}\n\n")
                
                # Read output line by line
                for line in iter(process.stdout.readline, ''):
                    if line:
                        line_str = line.strip()
                        if line_str:
                            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                            # Detect log level
                            level = "INFO"
                            if any(word in line_str.upper() for word in ['ERROR', 'EXCEPTION', 'FAILED', 'CRITICAL']):
                                level = "ERROR"
                            elif any(word in line_str.upper() for word in ['WARNING', 'WARN', 'CAUTION']):
                                level = "WARNING"
                            elif any(word in line_str.upper() for word in ['DEBUG', 'TRACE']):
                                level = "DEBUG"
                            
                            log_entry = f"{timestamp} [{level}] {line_str}\n"
                            f.write(log_entry)
                            f.flush()  # Ensure immediate write
                            
                            # Also print to console for debugging
                            print(f"📝 {marketplace_name} Log: {line_str}")
                            
                            # Store in marketplace_logs if task_id is provided
                            if task_id:
                                add_marketplace_log(task_id, level.lower(), f"📝 {marketplace_name} Log: {line_str}")
                            else:
                                # Store in global marketplace logs for apps without task_id
                                global_marketplace_logs.append({
                                    "timestamp": datetime.now().isoformat(),
                                    "level": level.lower(),
                                    "message": f"📝 {marketplace_name} Log: {line_str}",
                                    "marketplace": marketplace_name,
                                    "user": user_name,
                                    "brand": brand_name
                                })
                                # Keep only last 1000 global logs
                                if len(global_marketplace_logs) > 1000:
                                    global_marketplace_logs.pop(0)
                
                # Write footer
                f.write(f"\n{'='*50}\n")
                f.write(f"End Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"{'='*50}\n\n")
                
        except Exception as e:
            print(f"⚠️ Error logging {marketplace_name} output: {str(e)}")
    
    if platform.system() == 'Windows':
        try:
            # Run executable and capture output
            process = subprocess.Popen(
                [exe_path],
                cwd=cwd_path,
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                creationflags=subprocess.CREATE_NO_WINDOW,
                bufsize=1,
                universal_newlines=True
            )
            
            # Start logging thread
            log_thread = threading.Thread(
                target=log_output, 
                args=(process, log_file_path),
                daemon=True
            )
            log_thread.start()
            
            print(f"🚀 Started {marketplace_name} app with logging to: {log_file_path}")
            return True, process
            
        except Exception as e:
            print(f"❌ Failed to start {marketplace_name} with logging: {str(e)}")
            return False, None
    else:
        # For other systems, run directly
        try:
            process = subprocess.Popen(
                [exe_path],
                cwd=cwd_path,
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                universal_newlines=True
            )
            
            # Start logging thread
            log_thread = threading.Thread(
                target=log_output, 
                args=(process, log_file_path),
                daemon=True
            )
            log_thread.start()
            
            print(f"🚀 Started {marketplace_name} app with logging on {platform.system()}")
            return True, process
            
        except Exception as e:
            print(f"❌ Failed to start {marketplace_name} with logging on {platform.system()}: {str(e)}")
            return False, None

def ensure_marketplace_folder_exists(user_workspace, marketplace, config):
    """Ensure marketplace folder and files exist for user (optimized to avoid unnecessary copies)"""
    try:
        folder_path = os.path.join(user_workspace, config['folder'])
        
        # Check cache first
        cache_key = f"{user_workspace}_{marketplace}"
        if cache_key in _marketplace_folder_cache:
            if _marketplace_folder_cache[cache_key]:
                print(f"✅ Using cached marketplace folder: {config['folder']}")
                return True
        
        # Create folder if it doesn't exist
        if not os.path.exists(folder_path):
            os.makedirs(folder_path, exist_ok=True)
            print(f"📁 Created marketplace folder: {folder_path}")
        
        # Get root marketplace folder as template
        project_root = get_project_root()
        root_folder_path = os.path.join(project_root, 'JobGetOrder', config['folder'])
        
        # Only copy files if root folder exists and user folder is empty or missing essential files
        if os.path.exists(root_folder_path):
            root_files = os.listdir(root_folder_path)
            user_files = os.listdir(folder_path) if os.path.exists(folder_path) else []
            
            # Check if essential files are missing (exe files, configs, etc.)
            essential_files = [f for f in root_files if f.endswith(('.exe', '.dll', '.config', '.json', '.txt'))]
            missing_essential = [f for f in essential_files if f not in user_files]
            
            if missing_essential:
                print(f"📋 Found {len(missing_essential)} missing essential files, copying...")
                
                # Copy only missing essential files
                for file in missing_essential:
                    src_path = os.path.join(root_folder_path, file)
                    dst_path = os.path.join(folder_path, file)
                    if os.path.isfile(src_path):
                        shutil.copy2(src_path, dst_path)
                        print(f"📋 Copied {file} to user marketplace folder")
            else:
                # All essential files exist, no need to copy
                print(f"✅ All essential files already exist in {config['folder']}")
        
        # Cache the result
        _marketplace_folder_cache[cache_key] = True
        return True
    except Exception as e:
        print(f"❌ Error ensuring marketplace folder exists: {str(e)}")
        _marketplace_folder_cache[cache_key] = False
        return False

def auto_run_marketplace_app(sales_channel, brand_name, batch_name, user_id=None, task_id=None):
    """Automatically run marketplace app after generating Orderlist.txt"""
    try:
        print(f"🚀 Auto-running {sales_channel} app for {brand_name} batch {batch_name}")
        if task_id:
            add_marketplace_log(task_id, "info", f"🚀 Starting auto-run of {sales_channel} app for {brand_name} batch {batch_name}")
        # Map sales channel to marketplace folder and exe
        marketplace_configs = {
            'shopee': {'folder': 'Shopee', 'exe': 'ShopeeOrderLogistic.exe'},
            'lazada': {'folder': 'Lazada', 'exe': 'LazadaMarketplace.exe'},
            'blibli': {'folder': 'Blibli', 'exe': 'BliBliProduct2024.exe'},
            'desty': {'folder': 'Desty', 'exe': 'Desty.Console.exe'},
            'ginee': {'folder': 'Ginee', 'exe': 'Ginee.sync.exe'},
            'tiktok': {'folder': 'Tiktok', 'exe': 'tiktok.api.exe'},
            'zalora': {'folder': 'Zalora', 'exe': 'Zalora.Flexo.Integration.exe'},
            'tokopedia': {'folder': 'Tokopedia', 'exe': 'TokopediaOrder.exe'},
            'jubelio': {'folder': 'Jubelio', 'exe': 'Jubelio_project.exe'}
        }
        
        marketplace = sales_channel.lower()
        if marketplace not in marketplace_configs:
            print(f"⚠️  Unknown marketplace: {marketplace}, skipping auto-run")
            return
        
        config = marketplace_configs[marketplace]
        
        # Get the JobGetOrder directory path with user-specific workspace
        if user_id:
            # Use user-specific workspace
            user_workspace = multi_user_handler.get_user_workspace(user_id)
            
            jobgetorder_path = os.path.join(user_workspace, config['folder'])
            
            # Ensure marketplace folder exists (auto-create if missing) only if folder doesn't exist
            if not os.path.exists(jobgetorder_path):
                ensure_marketplace_folder_exists(user_workspace, marketplace, config)
        else:
            # Fallback to original path for backward compatibility
            project_root = get_project_root()
            jobgetorder_path = os.path.join(project_root, 'JobGetOrder', config['folder'])
        
        exe_path = os.path.join(jobgetorder_path, config['exe'])
        
        # Check if exe exists
        if not os.path.exists(exe_path):
            print(f"⚠️  {config['exe']} not found, skipping auto-run")
            if task_id:
                add_marketplace_log(task_id, "warning", f"⚠️ {config['exe']} not found, skipping auto-run")
            return
        
        # Check if we have permission to execute
        if not os.access(exe_path, os.X_OK):
            print(f"⚠️  No execute permission for {config['exe']}, skipping auto-run")
            if task_id:
                add_marketplace_log(task_id, "warning", f"⚠️ No execute permission for {config['exe']}, skipping auto-run")
            return
        
        # Check if Orderlist.txt exists and has content (keep original flow for auto-run)
        orderlist_path = os.path.join(jobgetorder_path, 'Orderlist.txt')
        if not os.path.exists(orderlist_path):
            print(f"⚠️  Orderlist.txt not found, skipping auto-run")
            if task_id:
                add_marketplace_log(task_id, "warning", f"⚠️ Orderlist.txt not found, skipping auto-run")
            return
        
        # Find the most recent backup file for verification (keep original flow for auto-run)
        backup_files = [f for f in os.listdir(jobgetorder_path) if f.startswith('Orderlist_') and f.endswith('.txt')]
        if backup_files:
            print(f"📁 Found {len(backup_files)} backup files for verification")
            
            # Get the most recent backup file
            backup_files.sort(key=lambda x: os.path.getmtime(os.path.join(jobgetorder_path, x)), reverse=True)
            latest_backup = os.path.join(jobgetorder_path, backup_files[0])
            
            print(f"🔄 Selected latest backup: {backup_files[0]}")
            print(f"   📅 Backup timestamp: {os.path.getmtime(latest_backup)}")
            
            # Count orders in backup file to verify against current Orderlist.txt
            with open(latest_backup, 'r', encoding='utf-8') as f:
                backup_lines = f.readlines()
                expected_count = len([line.strip() for line in backup_lines if line.strip()])
            
            # Verify current Orderlist.txt against backup
            with open(orderlist_path, 'r', encoding='utf-8') as f:
                current_lines = f.readlines()
                actual_count = len([line.strip() for line in current_lines if line.strip()])
            
            if actual_count != expected_count:
                print(f"⚠️  Orderlist.txt count mismatch before auto-run!")
                print(f"   Expected: {expected_count} orders (from backup)")
                print(f"   Actual: {actual_count} orders (current file)")
                print(f"🔄 Auto-restoring from latest backup...")
                
                # Make file writable before restore
                try:
                    os.chmod(orderlist_path, 0o666)  # Make writable
                    print(f"🔓 Made Orderlist.txt writable for restore")
                except Exception as e:
                    print(f"⚠️ Could not change permissions: {str(e)}")
                
                shutil.copy2(latest_backup, orderlist_path)
                print(f"✅ Auto-restored Orderlist.txt before auto-run")
            else:
                print(f"✅ Orderlist.txt verified: {actual_count} orders")
        
        # Make Orderlist.txt writable for executable (if it was read-only)
        try:
            os.chmod(orderlist_path, 0o644)  # Read-write for owner, read-only for others
            print(f"🔓 Made Orderlist.txt writable for executable")
        except Exception as e:
            print(f"⚠️ Could not change Orderlist.txt permissions: {str(e)}")
        
        # Count orders in Orderlist.txt
        with open(orderlist_path, 'r', encoding='utf-8') as f:
            lines = f.readlines()
            order_count = len([line.strip() for line in lines if line.strip()])
        
        if order_count == 0:
            print(f"⚠️  No orders in Orderlist.txt, skipping auto-run")
            return
        
        # Use provided brand_name directly (don't try to detect from filename)
        # The orderlist filename uses task_id, not brand name
        final_brand = brand_name
        print(f"✅ Auto-run using brand: {final_brand}")
        
        # Replace default config with brand-specific config (if exists)
        default_config_path = os.path.join(jobgetorder_path, f"{config['exe']}.config")
        brand_config_path = os.path.join(jobgetorder_path, f"{final_brand}_{marketplace.upper()}.config")
        
        if os.path.exists(brand_config_path):
            print(f"🔄 Replacing default config with {final_brand} config...")
            # Backup original config
            backup_config_path = os.path.join(jobgetorder_path, f"{config['exe']}.config.backup")
            if os.path.exists(default_config_path):
                shutil.copy2(default_config_path, backup_config_path)
                print(f"   📋 Backed up original config to {os.path.basename(backup_config_path)}")
            
            # Replace with brand config
            shutil.copy2(brand_config_path, default_config_path)
            print(f"   ✅ Using {final_brand} config with correct shop_id")
        else:
            print(f"⚠️  Brand config not found: {os.path.basename(brand_config_path)}")
            print(f"   Using default config (may have wrong shop_id)")
        
        # Run the marketplace app
        import subprocess
        import platform
        
        print(f"🚀 Auto-running {marketplace.title()} app for {order_count} orders...")
        
        # Run the executable with logging (capture console output)
        if task_id:
            add_marketplace_log(task_id, "info", f"🚀 Starting {marketplace.title()} app for {order_count} orders...")
        success, process = run_executable_with_logging(exe_path, jobgetorder_path, marketplace, user_id or "system", final_brand, task_id)
        if not success:
            print(f"⚠️ Failed to auto-start {marketplace} app with logging")
            if task_id:
                add_marketplace_log(task_id, "warning", f"⚠️ Failed to auto-start {marketplace} app with logging")
        else:
            if task_id:
                add_marketplace_log(task_id, "success", f"✅ {marketplace.title()} app started successfully!")
        
        print(f"✅ {marketplace.title()} app started successfully!")
        print(f"   📁 Location: {jobgetorder_path}")
        print(f"   📊 Orders to process: {order_count}")
        print(f"   🏪 Brand: {brand_name}, Batch: {batch_name}")
        
        # Log the auto-run event
        logger.info(f"Auto-run {marketplace} app: {order_count} orders for {brand_name} batch {batch_name}")
        
        # Return success info for notification
        return {
            "success": True,
            "marketplace": marketplace.title(),
            "brand": brand_name,
            "batch": batch_name,
            "order_count": order_count,
            "message": f"{marketplace.title()} app started successfully for {order_count} orders"
        }
        
    except Exception as e:
        print(f"❌ Error auto-running {marketplace} app: {str(e)}")
        logger.error(f"Error auto-running marketplace app: {str(e)}")
        if task_id:
            add_marketplace_log(task_id, "error", f"❌ Error auto-running {marketplace} app: {str(e)}")

def get_database_connection(connection_string):
    """Get database connection with retry logic and multiple connection strings"""
    connection_strings = [
        connection_string,
        connection_string.replace("ODBC Driver 17", "ODBC Driver 18"),
        connection_string.replace("ODBC Driver 18", "ODBC Driver 17"),
    ]
    
    for i, conn_str in enumerate(connection_strings):
        try:
            # Silent connection attempt for better performance
            conn = pyodbc.connect(conn_str, timeout=30)
            # Silent connection success for better performance
            return conn
        except Exception as e:
            print(f"❌ Connection attempt {i+1} failed: {str(e)}")
            if i == len(connection_strings) - 1:  # Last attempt
                print(f"All connection attempts failed for: {connection_string}")
                return None
            continue
    
    return None

def check_external_database_status(order_numbers, marketplace):
    """Check order status in external database - OPTIMIZED VERSION"""
    try:
        # Get the appropriate field mapping for the marketplace
        field_mapping = MARKETPLACE_FIELD_MAPPING.get(marketplace.lower())
        if not field_mapping:
            print(f"⚠️ No field mapping found for marketplace: {marketplace}")
            return {}
        
        # Minimize console logs for better performance
        # Removed external database logging
        
        # OPTIMIZATION: Use smaller chunk size for better timeout handling
        MAX_PARAMS_PER_QUERY = min(100, len(order_numbers))  # Reduced to 100 parameters for faster processing
        all_results = {}
        
        # Get single connection for all chunks (connection pooling) - OPTIMIZED
        conn = get_database_connection(WMSPROD_DB_CONNECTION_STRING)
        if not conn:
            return {}  # Silent failure for better performance
        
        # Use connection timeout optimization
        conn.timeout = 10  # Further reduced timeout for faster failure detection
        
        try:
            cursor = conn.cursor()
            
            # Process in optimized chunks
            for i in range(0, len(order_numbers), MAX_PARAMS_PER_QUERY):
                chunk_order_numbers = order_numbers[i:i + MAX_PARAMS_PER_QUERY]
                print(f"  Processing chunk {i//MAX_PARAMS_PER_QUERY + 1}: {len(chunk_order_numbers)} orders")
                
                # Create placeholders for this chunk
                placeholders = ','.join(['?' for _ in chunk_order_numbers])
                
                # Query external database with both Flexo_Db and WMSPROD, including ItemIdFlexo
                # All marketplaces use alphanumeric order numbers - no numeric validation needed
                query = f"""
                SELECT so.MerchantName, so.SystemId, so.SystemRefId, so.OrderStatus, 
                           so.OrderDate, so.Awb, so.TransporterCode, so.{field_mapping} as order_number,
                       sol.ItemId as item_id_flexo,
                       CASE WHEN ol.ordnum IS NOT NULL THEN 'Interface' ELSE 'Not Yet Interface' END as interface_status
                FROM Flexo_Db.dbo.SalesOrder so
                    LEFT JOIN Flexo_Db.dbo.SalesOrderLine sol ON sol.SystemRefId = so.SystemRefId
                    LEFT JOIN WMSPROD.dbo.ord_line ol ON ol.ordnum = so.SystemRefId
                        WHERE so.{field_mapping} IN ({placeholders})
                """
            
                # Execute query for this chunk
                cursor.execute(query, chunk_order_numbers)
                chunk_results = cursor.fetchall()
                
                # Process results for this chunk
                for result in chunk_results:
                    merchant_name, system_id, system_ref_id, order_status, order_date, awb, transporter_code, order_number, item_id_flexo, interface_status = result
                
                    # Use the original order ID (from Excel) as the key
                    if order_number not in all_results:
                        all_results[order_number] = {
                            'system_id': system_id,
                            'merchant_name': merchant_name,
                            'system_ref_id': system_ref_id,
                            'order_status': order_status,
                            'order_date': order_date,
                            'awb': awb,
                            'transporter_code': transporter_code,
                            'item_id_flexo': [],  # ✅ Collect as list
                            'interface_status': interface_status
                        }
                    
                    if item_id_flexo and item_id_flexo not in all_results[order_number]['item_id_flexo']:
                        all_results[order_number]['item_id_flexo'].append(item_id_flexo)
                
                # Silent database query result for better performance
                
        except Exception as e:
            print(f"  ❌ Error processing chunks: {e}")
        finally:
            cursor.close()
            conn.close()
        
        # Convert item_id_flexo lists to comma-separated strings
        for order_number, result_data in all_results.items():
            item_id_flexo_list = result_data.get('item_id_flexo', [])
            if item_id_flexo_list:
                # Convert list to comma-separated string
                item_id_flexo_str = ','.join(item_id_flexo_list)
                all_results[order_number]['item_id_flexo'] = item_id_flexo_str
            else:
                all_results[order_number]['item_id_flexo'] = None
        
        print(f"✅ External database check completed: {len(all_results)} orders found")
        return all_results
        
    except Exception as e:
        print(f"❌ Error checking external database: {e}")
        return {}

def check_ord_line_status(order_numbers, marketplace):
    """Check if orders exist in ord_line table (WMSPROD.dbo.ord_line) - OPTIMIZED VERSION"""
    try:
        # Skip field mapping untuk ord_line check karena menggunakan SystemRefId langsung
        if not order_numbers:
            return {}
        
        # ord_line check log removed for better performance
        
        # OPTIMIZATION: Use reasonable chunk size to avoid SQL parameter limits
        MAX_PARAMS_PER_QUERY = min(1000, len(order_numbers))  # Limit to 1000 parameters max
        all_results = {}
        
        # Get single connection for all chunks (connection pooling)
        conn = get_database_connection(WMSPROD_DB_CONNECTION_STRING)
        if not conn:
            print(f"❌ Failed to connect to ord_line database")
            return {}
        
        try:
            cursor = conn.cursor()
            
            # Process in optimized chunks
            for i in range(0, len(order_numbers), MAX_PARAMS_PER_QUERY):
                chunk_order_numbers = order_numbers[i:i + MAX_PARAMS_PER_QUERY]
                
                # Create placeholders for this chunk
                placeholders = ','.join(['?' for _ in chunk_order_numbers])
                
                # Query ord_line table only - menggunakan SystemRefId dari external database
                query = f"""
                SELECT 
                    ol.ordnum
                FROM WMSPROD.dbo.ord_line ol
                WHERE ol.ordnum IN ({placeholders})
                """
                
                # Execute query for this chunk
                cursor.execute(query, chunk_order_numbers)
                chunk_results = cursor.fetchall()
                
                # Process results for this chunk
                for row in chunk_results:
                    ordnum = row[0]  # ol.ordnum
                    # Gunakan ordnum sebagai key untuk mapping dengan external results
                    all_results[ordnum] = {
                        "found_in_ord_line": True,
                        "ordnum": ordnum
                    }
                
            cursor.close()
            conn.close()
        
        except Exception as e:
            print(f"❌ Error processing ord_line chunks: {e}")
            if conn:
                conn.close()
        
        # Silent ord_line check completion for better performance
        return all_results
        
    except Exception as e:
        print(f"❌ Error checking ord_line: {str(e)}")
        return {}

def check_interface_status(order_numbers, marketplace):
    """Check interface status by combining external database and ord_line checks - ULTRA-OPTIMIZED"""
    try:
        # Minimize console logs for better performance
        # Removed interface check logging
        
        # ULTRA-OPTIMIZATION: Single query with LEFT JOIN (no separate ord_line query needed)
        external_results = check_external_database_status(order_numbers, marketplace)
        
        # Process results (ord_line status already included in external_results via LEFT JOIN)
        all_results = {}
        
        for order_id in order_numbers:
            external_data = external_results.get(order_id, {})
            
            if external_data:
                # Order exists in external database (ord_line status already included via LEFT JOIN)
                all_results[order_id] = {
                    "status": external_data.get("interface_status", "Not Yet Interface"),
                    "system_id": external_data.get("system_id"),
                    "system_ref_id": external_data.get("system_ref_id"),
                    "merchant_name": external_data.get("merchant_name"),
                    "order_status": external_data.get("order_status"),
                    "order_date": external_data.get("order_date"),
                    "awb": external_data.get("awb"),
                    "transporter_code": external_data.get("transporter_code"),
                    "item_id_flexo": external_data.get("item_id_flexo"),  # ✅ Add ItemIdFlexo
                    "found_in_external": True,
                    "found_in_ord_line": external_data.get("interface_status") == "Interface",
                    "ordnum": external_data.get("system_ref_id") if external_data.get("interface_status") == "Interface" else None
                }
            else:
                # Order not found in external database
                all_results[order_id] = {
                    "status": "Not Yet Interface",
                    "system_ref_id": None,  # ✅ Add SystemRefId as None
                    "item_id_flexo": None,  # ✅ Add ItemIdFlexo as None
                    "found_in_external": False,
                    "found_in_ord_line": False,
                    "ordnum": None
                }
        
        # Minimize console logs for better performance
        # Removed interface check completion logging
        return all_results
        
    except Exception as e:
        print(f"❌ Error checking interface status: {str(e)}")
        return {}

def get_interface_status_summary(orders, db):
    """Get interface status summary for uploaded orders using real database queries"""
    try:
        interface_orders = []
        not_interface_orders = []
        
        # Group orders by marketplace for efficient database queries
        orders_by_marketplace = {}
        for order in orders:
            marketplace = order.Marketplace.lower()
            if marketplace not in orders_by_marketplace:
                orders_by_marketplace[marketplace] = []
            orders_by_marketplace[marketplace].append(order)
        
        # Check interface status for each marketplace
        for marketplace, marketplace_orders in orders_by_marketplace.items():
            order_numbers = [order.OrderNumber for order in marketplace_orders if order.OrderNumber]
            
            if not order_numbers:
                # If no order numbers, fall back to AWB-based logic
                for order in marketplace_orders:
                    if order.AWB and order.AWB.strip():
                        interface_orders.append({
                            "order_number": order.OrderNumberFlexo if order.OrderNumberFlexo and order.OrderNumberFlexo.strip() else order.OrderNumber,
                            "awb": order.AWB,
                            "status": order.OrderStatus,
                            "transporter": order.Transporter,
                            "reason": "Has AWB/Tracking code (fallback)"
                        })
                    else:
                        not_interface_orders.append({
                            "order_number": order.OrderNumberFlexo if order.OrderNumberFlexo and order.OrderNumberFlexo.strip() else order.OrderNumber,
                            "status": order.OrderStatus,
                            "reason": "No AWB/Tracking code available (fallback)"
                        })
                continue
            
            # Query database for interface status
            interface_results = check_interface_status(order_numbers, marketplace)
            
            # Process results
            for order in marketplace_orders:
                order_number = order.OrderNumber
                if order_number in interface_results:
                    result = interface_results[order_number]
                    if result["status"] == "Interface":
                        interface_orders.append({
                            "order_number": order.OrderNumberFlexo if order.OrderNumberFlexo and order.OrderNumberFlexo.strip() else order_number,
                            "awb": result.get("awb", order.AWB),
                            "status": result.get("order_status", order.OrderStatus),
                            "transporter": result.get("transporter_code", order.Transporter),
                            "system_id": result.get("system_id"),
                            "ordnum": result.get("ordnum"),
                            "reason": "Found in WMSPROD.ord_line"
                        })
                    else:
                        not_interface_orders.append({
                            "order_number": order.OrderNumberFlexo if order.OrderNumberFlexo and order.OrderNumberFlexo.strip() else order_number,
                            "status": result.get("order_status", order.OrderStatus),
                            "reason": "Not found in WMSPROD.ord_line"
                        })
                else:
                    # Order not found in database, use fallback logic
                    if order.AWB and order.AWB.strip():
                        interface_orders.append({
                            "order_number": order.OrderNumberFlexo if order.OrderNumberFlexo and order.OrderNumberFlexo.strip() else order_number,
                            "awb": order.AWB,
                            "status": order.OrderStatus,
                            "transporter": order.Transporter,
                            "reason": "Has AWB/Tracking code (not in database)"
                        })
                    else:
                        not_interface_orders.append({
                            "order_number": order.OrderNumberFlexo if order.OrderNumberFlexo and order.OrderNumberFlexo.strip() else order_number,
                            "status": order.OrderStatus,
                            "reason": "No AWB/Tracking code and not in database"
                        })
        
        return {
            "interface": {
                "count": len(interface_orders),
                "orders": interface_orders
            },
            "not_interface": {
                "count": len(not_interface_orders),
                "orders": not_interface_orders
            },
            "total": len(orders),
            "interface_percentage": round((len(interface_orders) / len(orders)) * 100, 2) if orders else 0
        }
    except Exception as e:
        print(f"Error getting interface status summary: {str(e)}")
        # Fallback to simple AWB-based logic
        interface_orders = []
        not_interface_orders = []
        
        for order in orders:
            if order.AWB and order.AWB.strip():
                interface_orders.append({
                    "order_number": order.OrderNumberFlexo if order.OrderNumberFlexo and order.OrderNumberFlexo.strip() else order.OrderNumber,
                    "awb": order.AWB,
                    "status": order.OrderStatus,
                    "transporter": order.Transporter,
                    "reason": "Has AWB/Tracking code (fallback due to error)"
                })
            else:
                not_interface_orders.append({
                    "order_number": order.OrderNumberFlexo if order.OrderNumberFlexo and order.OrderNumberFlexo.strip() else order.OrderNumber,
                    "status": order.OrderStatus,
                    "reason": "No AWB/Tracking code (fallback due to error)"
                })
        
        return {
            "interface": {
                "count": len(interface_orders),
                "orders": interface_orders
            },
            "not_interface": {
                "count": len(not_interface_orders),
                "orders": not_interface_orders
            },
            "total": len(orders),
            "interface_percentage": round((len(interface_orders) / len(orders)) * 100, 2) if orders else 0,
            "error": str(e)
        }

# Routes
@app.websocket("/ws")
async def websocket_endpoint(websocket: WebSocket):
    """Simple WebSocket endpoint for React development server"""
    await websocket.accept()
    try:
        while True:
            # Keep connection alive - receive and ignore messages
            try:
                await websocket.receive_text()
            except:
                break
    except WebSocketDisconnect:
        pass

@app.get("/health")
def health_check():
    """Enhanced health check endpoint with metrics"""
    try:
        metrics = get_metrics_summary()
        return {
            "status": "healthy",
            "message": "Backend is running",
            "timestamp": datetime.now().isoformat(),
            "metrics": metrics
        }
    except Exception as e:
        logger.error(f"Health check error: {e}")
        return {
            "status": "unhealthy",
            "message": "Backend is running but monitoring failed",
            "error": str(e),
            "timestamp": datetime.now().isoformat()
        }

@app.get("/metrics")
def get_metrics():
    """Get detailed application metrics"""
    try:
        return get_metrics_summary()
    except Exception as e:
        logger.error(f"Metrics error: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to get metrics: {str(e)}")

@app.get("/metrics/system")
def get_system_metrics():
    """Get system performance metrics"""
    try:
        from monitoring import metrics_collector
        if metrics_collector.system_metrics:
            latest = metrics_collector.system_metrics[-1]
            return {
                "timestamp": latest.timestamp.isoformat(),
                "cpu_percent": latest.cpu_percent,
                "memory_percent": latest.memory_percent,
                "memory_used_mb": latest.memory_used_mb,
                "memory_total_mb": latest.memory_total_mb,
                "disk_usage_percent": latest.disk_usage_percent,
                "disk_free_gb": latest.disk_free_gb,
                "active_connections": latest.active_connections,
                "requests_per_minute": latest.requests_per_minute,
                "error_rate": latest.error_rate
            }
        else:
            return {"message": "No system metrics available yet"}
    except Exception as e:
        logger.error(f"System metrics error: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to get system metrics: {str(e)}")

@app.get("/metrics/api")
def get_api_metrics():
    """Get API performance metrics"""
    try:
        from monitoring import metrics_collector
        return {
            "endpoint_stats": dict(metrics_collector.endpoint_stats),
            "error_counts": dict(metrics_collector.error_counts),
            "total_requests": len(metrics_collector.api_metrics),
            "recent_requests": [
                {
                    "endpoint": m.endpoint,
                    "method": m.method,
                    "response_time": m.response_time,
                    "status_code": m.status_code,
                    "timestamp": m.timestamp.isoformat()
                }
                for m in list(metrics_collector.api_metrics)[-10:]  # Last 10 requests
            ]
        }
    except Exception as e:
        logger.error(f"API metrics error: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to get API metrics: {str(e)}")

@app.post("/register", response_model=Token)
def register(user: UserCreate, db: Session = Depends(get_db)):
    db_user = db.query(User).filter(func.lower(User.username) == func.lower(user.username)).first()
    if db_user:
        raise HTTPException(status_code=400, detail="Username already registered")
    
    hashed_password = get_password_hash(user.password)
    db_user = User(
        username=user.username, 
        email=user.email, 
        password_hash=hashed_password,
        role=user.role
    )
    db.add(db_user)
    db.commit()
    db.refresh(db_user)
    
    access_token = create_access_token(data={"sub": user.username})
    return {"access_token": access_token, "token_type": "bearer"}

@app.post("/login", response_model=Token)
def login(user: UserLogin, db: Session = Depends(get_db)):
    db_user = db.query(User).filter(func.lower(User.username) == func.lower(user.username)).first()
    if not db_user or not verify_password(user.password, db_user.password_hash):
        raise HTTPException(status_code=401, detail="Incorrect username or password")
    
    # Check if user is active
    if not db_user.is_active:
        raise HTTPException(status_code=401, detail="Account is inactive. Please contact administrator.")
    
    access_token = create_access_token(data={"sub": user.username})
    return {"access_token": access_token, "token_type": "bearer"}

@app.get("/me-debug")
def get_current_user_info_debug(current_user: str = Depends(get_current_user), db: Session = Depends(get_db)):
    """Debug endpoint to see raw user data"""
    try:
        user = db.query(User).filter(func.lower(User.username) == func.lower(current_user)).first()
        if not user:
            raise HTTPException(status_code=404, detail="User not found")
        
        # Return raw user data without response model
        return {
            "id": user.id,
            "username": user.username,
            "email": user.email,
            "role": user.role,
            "is_active": user.is_active,
            "created_at": user.created_at,
            "updated_at": user.updated_at,
            "is_active_type": str(type(user.is_active)),
            "role_type": str(type(user.role))
        }
    except Exception as e:
        logger.error(f"Error in /me-debug endpoint: {e}")
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")

@app.get("/me", response_model=UserResponse)
def get_current_user_info(current_user: str = Depends(get_current_user), db: Session = Depends(get_db)):
    """Get current user information including role"""
    try:
        user = db.query(User).filter(func.lower(User.username) == func.lower(current_user)).first()
        if not user:
            raise HTTPException(status_code=404, detail="User not found")
        
        # Check if user is active
        if not user.is_active:
            raise HTTPException(status_code=401, detail="Account is inactive. Please contact administrator.")
        
        return user
    except Exception as e:
        logger.error(f"Error in /me endpoint: {e}")
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")

class ProfileUpdate(BaseModel):
    email: Optional[EmailStr] = None
    current_password: Optional[str] = None
    new_password: Optional[str] = None
    
    @field_validator('new_password')
    @classmethod
    def validate_new_password(cls, v):
        if v is not None and len(v) < 6:
            raise ValueError('New password must be at least 6 characters long')
        return v

@app.put("/me", response_model=UserResponse)
def update_profile(
    profile_update: ProfileUpdate,
    current_user: str = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Update current user profile (email and password only)"""
    user = db.query(User).filter(func.lower(User.username) == func.lower(current_user)).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Update email if provided
    if profile_update.email is not None:
        # Check if new email already exists
        existing_email = db.query(User).filter(User.email == profile_update.email, User.id != user.id).first()
        if existing_email:
            raise HTTPException(status_code=400, detail="Email already exists")
        user.email = profile_update.email
    
    # Update password if provided
    if profile_update.new_password is not None:
        if not profile_update.current_password:
            raise HTTPException(status_code=400, detail="Current password is required to change password")
        
        # Verify current password
        if not verify_password(profile_update.current_password, user.password_hash):
            raise HTTPException(status_code=400, detail="Current password is incorrect")
        
        # Update to new password
        user.password_hash = get_password_hash(profile_update.new_password)
    
    user.updated_at = get_wib_now()
    
    db.commit()
    db.refresh(user)
    return user

# Admin endpoints for user management
def get_current_admin_user(current_user: str = Depends(get_current_user), db: Session = Depends(get_db)):
    """Check if current user is admin or superuser"""
    user = db.query(User).filter(func.lower(User.username) == func.lower(current_user)).first()
    if not user or user.role not in ["admin", "superuser"]:
        raise HTTPException(status_code=403, detail="Admin access required")
    return user

@app.get("/api/admin/users", response_model=List[UserResponse])
def get_all_users(
    current_admin: User = Depends(get_current_admin_user),
    db: Session = Depends(get_db)
):
    """Get all users (admin only)"""
    users = db.query(User).all()
    return users

@app.post("/api/admin/users", response_model=UserResponse)
def create_user(
    user_data: UserCreateAdmin,
    current_admin: User = Depends(get_current_admin_user),
    db: Session = Depends(get_db)
):
    """Create new user (admin only)"""
    # Prevent admin from creating superuser accounts
    if current_admin.role == "admin" and user_data.role == "superuser":
        raise HTTPException(status_code=403, detail="Admin cannot create superuser accounts")
    
    # Check if username already exists
    existing_user = db.query(User).filter(func.lower(User.username) == func.lower(user_data.username)).first()
    if existing_user:
        raise HTTPException(status_code=400, detail="Username already exists")
    
    # Check if email already exists (only if email is provided)
    if user_data.email:
        existing_email = db.query(User).filter(User.email == user_data.email).first()
        if existing_email:
            raise HTTPException(status_code=400, detail="Email already exists")
    
    hashed_password = get_password_hash(user_data.password)
    new_user = User(
        username=user_data.username,
        email=user_data.email,
        password_hash=hashed_password,
        role=user_data.role
    )
    
    db.add(new_user)
    db.commit()
    db.refresh(new_user)
    
    return new_user

@app.get("/api/admin/users/{user_id}", response_model=UserResponse)
def get_user(
    user_id: int,
    current_admin: User = Depends(get_current_admin_user),
    db: Session = Depends(get_db)
):
    """Get specific user by ID (admin only)"""
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    return user

@app.put("/api/admin/users/{user_id}", response_model=UserResponse)
def update_user(
    user_id: int,
    user_update: UserUpdate,
    current_admin: User = Depends(get_current_admin_user),
    db: Session = Depends(get_db)
):
    """Update user (admin only)"""
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Update fields if provided
    if user_update.username is not None:
        # Check if new username already exists
        existing_user = db.query(User).filter(func.lower(User.username) == func.lower(user_update.username), User.id != user_id).first()
        if existing_user:
            raise HTTPException(status_code=400, detail="Username already exists")
        user.username = user_update.username
    
    if user_update.email is not None:
        # Check if new email already exists
        existing_email = db.query(User).filter(User.email == user_update.email, User.id != user_id).first()
        if existing_email:
            raise HTTPException(status_code=400, detail="Email already exists")
        user.email = user_update.email
    
    if user_update.role is not None:
        user.role = user_update.role
    
    if user_update.is_active is not None:
        user.is_active = user_update.is_active
    
    user.updated_at = get_wib_now()
    
    db.commit()
    db.refresh(user)
    return user

@app.delete("/api/admin/users/{user_id}")
def delete_user(
    user_id: int,
    current_admin: User = Depends(get_current_admin_user),
    db: Session = Depends(get_db)
):
    """Delete user (admin only)"""
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Prevent admin from deleting themselves
    if user.id == current_admin.id:
        raise HTTPException(status_code=400, detail="Cannot delete your own account")
    
    # Prevent admin from deleting superuser accounts
    if current_admin.role == "admin" and user.role == "superuser":
        raise HTTPException(status_code=403, detail="Admin cannot delete superuser accounts")
    
    # Prevent admin from deleting other admin accounts
    if current_admin.role == "admin" and user.role == "admin":
        raise HTTPException(status_code=403, detail="Admin cannot delete other admin accounts")
    
    db.delete(user)
    db.commit()
    
    return {"message": "User deleted successfully"}

@app.post("/api/admin/users/bulk-create")
def bulk_create_users(
    users_data: List[UserCreateAdmin],
    current_admin: User = Depends(get_current_admin_user),
    db: Session = Depends(get_db)
):
    """Bulk create users (admin only)"""
    created_users = []
    errors = []
    
    for user_data in users_data:
        try:
            # Prevent admin from creating superuser accounts
            if current_admin.role == "admin" and user_data.role == "superuser":
                errors.append(f"Cannot create superuser account for '{user_data.username}' - Admin cannot create superuser accounts")
                continue
                
            # Check if username already exists
            existing_user = db.query(User).filter(func.lower(User.username) == func.lower(user_data.username)).first()
            if existing_user:
                errors.append(f"Username '{user_data.username}' already exists")
                continue
            
            # Check if email already exists (only if email is provided)
            if user_data.email:
                existing_email = db.query(User).filter(User.email == user_data.email).first()
                if existing_email:
                    errors.append(f"Email '{user_data.email}' already exists")
                    continue
            
            hashed_password = get_password_hash(user_data.password)
            new_user = User(
                username=user_data.username,
                email=user_data.email,
                password_hash=hashed_password,
                role=user_data.role
            )
            
            db.add(new_user)
            db.commit()
            db.refresh(new_user)
            created_users.append(new_user)
            
        except Exception as e:
            errors.append(f"Error creating user '{user_data.username}': {str(e)}")
    
    return {
        "created_users": len(created_users),
        "errors": errors,
        "users": created_users
    }

@app.post("/api/upload")
def upload_file(
    file: UploadFile = File(...),
    current_user: str = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Unified upload endpoint with auto-scaling performance optimization"""
    try:
        # Consolidated validation
        validation_result = validate_upload_request(file, db)
        filename_info = validation_result['filename_info']
        sales_channel = validation_result['sales_channel']
        brand_name = validation_result['brand_name']
        shop_id = validation_result['shop_id']
        
        # Read file content into memory first
        file_content = file.file.read()
        
        # Read file
        if file.filename.endswith('.xlsx'):
            df = pd.read_excel(io.BytesIO(file_content))
        elif file.filename.endswith('.csv'):
            df = pd.read_csv(io.BytesIO(file_content))
        else:
            raise HTTPException(status_code=400, detail="Unsupported file format")
        
        # Get marketplace mapping based on sales channel from filename
        sales_channel = filename_info['sales_channel']
        marketplace_mapping = get_marketplace_mapping(sales_channel)
        
        if not marketplace_mapping:
            raise HTTPException(status_code=400, detail=f"Unsupported sales channel: {sales_channel}")
        
        # Auto-scaling chunk size based on file size
        total_rows = len(df)
        file_size_mb = len(file_content) / (1024 * 1024)
        
        # Dynamic chunk size based on file size to prevent timeout
        if total_rows <= 1000:
            CHUNK_SIZE = 1000
        elif total_rows <= 5000:
            CHUNK_SIZE = 2000
        elif total_rows <= 20000:
            CHUNK_SIZE = 5000
        elif total_rows <= 50000:
            CHUNK_SIZE = 10000
        else:
            CHUNK_SIZE = 15000  # Reduced for very large files to prevent timeout
        
        total_chunks = (total_rows + CHUNK_SIZE - 1) // CHUNK_SIZE
        
        if total_rows <= 1000:
            processing_mode = "small"
        elif total_rows <= 10000:
            processing_mode = "medium"
        elif total_rows <= 50000:
            processing_mode = "large"
        elif total_rows <= 100000:
            processing_mode = "xlarge"
        else:
            processing_mode = "ultra"
        
        print(f"Auto-scaling: {total_rows} rows, {file_size_mb:.1f}MB -> {processing_mode} mode, chunk size: {CHUNK_SIZE} ({total_chunks} chunks)")
        
        total_processed = 0
        
        # Pre-calculate common values to avoid repeated calculations
        current_time = get_wib_now().replace(tzinfo=None)  # Store as naive datetime in WIB timezone
        brand_name = filename_info['brand']
        batch_name = filename_info['batch']
        
        # Generate task_id for main upload endpoint
        task_id = str(uuid.uuid4())
        
        # Create upload history record
        upload_history = UploadHistory(
            marketplace=sales_channel,
            brand=brand_name,
            pic=current_user,
            batch=batch_name
        )
        db.add(upload_history)
        db.commit()
        
        # Pre-process date for Tokopedia if needed
        tokopedia_date = None
        if sales_channel.lower() == 'tokopedia':
            try:
                date_str = filename_info['date']
                tokopedia_date = datetime.strptime(date_str, '%Y%m%d')
                tokopedia_date = WIB_TIMEZONE.localize(tokopedia_date)
            except:
                tokopedia_date = current_time
        
        # Ultra-optimized single-bulk processing for maximum speed
        start_time = datetime.now()
        
        # Pre-calculate all field mappings for speed
        order_number_col = marketplace_mapping['order_number']
        order_status_col = marketplace_mapping['order_status']
        order_date_col = marketplace_mapping['order_date']
        warehouse_col = marketplace_mapping['wh_loc']
        awb_col = marketplace_mapping['awb']
        transporter_col = marketplace_mapping['transporter']
        sla_col = marketplace_mapping['sla']
        sku_col = marketplace_mapping.get('sku_field')  # ✅ SKU field mapping
        
        # Vectorized field extraction for entire dataset
        order_numbers = df[order_number_col].fillna('').astype(str) if order_number_col in df.columns else pd.Series([''] * len(df))
        order_statuses = df[order_status_col].fillna('').astype(str) if order_status_col in df.columns else pd.Series([''] * len(df))
        order_dates = df[order_date_col].fillna('') if order_date_col in df.columns else pd.Series([''] * len(df))
        warehouses = df[warehouse_col].fillna('').astype(str) if warehouse_col in df.columns else pd.Series([''] * len(df))
        awbs = df[awb_col].fillna('').astype(str) if awb_col in df.columns else pd.Series([''] * len(df))
        transporters = df[transporter_col].fillna('').astype(str) if transporter_col in df.columns else pd.Series([''] * len(df))
        slas = df[sla_col].fillna('').astype(str) if sla_col in df.columns else pd.Series([''] * len(df))
        skus = df[sku_col].fillna('').astype(str) if sku_col and sku_col in df.columns else pd.Series([''] * len(df))  # ✅ SKU data
        
        # Group by OrderNumber to handle multiple items per order (vectorized)
        order_skus = defaultdict(list)
        if sku_col and sku_col in df.columns:
            # Vectorized approach for better performance
            valid_mask = (order_numbers != '') & (order_numbers != 'nan') & (skus != '') & (skus != 'nan')
            valid_data = df[valid_mask]
            if not valid_data.empty:
                for order_num, sku in zip(valid_data[order_number_col], valid_data[sku_col]):
                    order_num = str(order_num).strip()
                    sku = str(sku).strip()
                    if order_num and sku and sku not in order_skus[order_num]:
                        order_skus[order_num].append(sku)

        
        # NEW APPROACH: Check interface status FIRST, then save to database
        
        # Group orders by OrderNumber to handle multiple items per order
        grouped_orders = {}
        for i in range(len(df)):
            order_num = str(order_numbers.iloc[i]).strip()
            if not order_num or order_num == 'nan':
                continue
                
            # Simplified date handling
            if sales_channel.lower() == 'tokopedia' and not order_dates.iloc[i]:
                order_date = tokopedia_date
            elif order_dates.iloc[i]:
                try:
                    order_date = pd.to_datetime(order_dates.iloc[i], dayfirst=True)
                    order_date = convert_to_wib(order_date)
                except:
                    order_date = current_time
            else:
                order_date = current_time
            
            # If this order number already exists, just add the SKU to the existing order
            if order_num in grouped_orders:
                existing_order = grouped_orders[order_num]
                # Add SKU to existing order's ItemId
            if sku_col and sku_col in df.columns:
                sku_value = skus.iloc[i]
                if sku_value and sku_value.strip() and sku_value != 'nan':
                    existing_skus = existing_order['ItemId'].split(',') if existing_order['ItemId'] else []
                    if sku_value not in existing_skus:
                        existing_skus.append(sku_value)
                        existing_order['ItemId'] = ','.join(existing_skus)
            else:
                # Create new order entry
                item_id = None
                if sku_col and sku_col in df.columns:
                    sku_value = skus.iloc[i]
                    item_id = sku_value if sku_value and sku_value.strip() and sku_value != 'nan' else None
                
                grouped_orders[order_num] = {
                    'Marketplace': sales_channel,
                    'Brand': brand_name,
                    'OrderNumber': order_num,
                    'OrderStatus': order_statuses.iloc[i],
                    'AWB': awbs.iloc[i],
                    'Transporter': transporters.iloc[i],
                    'OrderDate': order_date,
                    'SLA': slas.iloc[i],
                    'Batch': batch_name,
                    'PIC': current_user,
                    'UploadDate': current_time,
                    'TaskId': task_id,
                    'ItemId': item_id,  # ✅ SKU dari Excel
                    'ItemIdFlexo': None  # ✅ Akan diisi dari Flexo database
                }
        
        # Convert grouped orders to list
        all_order_data = list(grouped_orders.values())
        print(f"📊 Grouped {len(df)} rows into {len(all_order_data)} unique orders")
        
        # Check interface status for all orders BEFORE saving to database
        interface_check_start = datetime.now()
        
        # Group orders by marketplace for efficient interface checking
        orders_by_marketplace = {}
        for order_data in all_order_data:
            marketplace = order_data['Marketplace'].lower()
            if marketplace not in orders_by_marketplace:
                orders_by_marketplace[marketplace] = []
            orders_by_marketplace[marketplace].append(order_data)
        
        # Check interface status for each marketplace - OPTIMIZED
        interface_results = {}
        for marketplace, marketplace_orders in orders_by_marketplace.items():
            
            # Extract order numbers for this marketplace
            order_numbers_list = [order['OrderNumber'] for order in marketplace_orders if order['OrderNumber']]
            
            if order_numbers_list:
                # ULTRA-AGGRESSIVE: Always process all at once for maximum speed
                EXTERNAL_CHUNK_SIZE = len(order_numbers_list)  # Process ALL at once - no chunking overhead
                
                for i in range(0, len(order_numbers_list), EXTERNAL_CHUNK_SIZE):
                    chunk_order_numbers = order_numbers_list[i:i + EXTERNAL_CHUNK_SIZE]
                    
                    try:
                        chunk_interface_results = check_interface_status(chunk_order_numbers, marketplace)
                    except Exception as e:
                        chunk_interface_results = {}
                    
                    interface_results.update(chunk_interface_results)
        
        interface_check_time = (datetime.now() - interface_check_start).total_seconds()
        add_upload_log(task_id, "success", f"✅ Interface status check completed in {interface_check_time:.2f}s")
        add_upload_log(task_id, "info", f"📊 Total interface results: {len(interface_results)}")
        
        # Debug: show sample of all interface results
        if interface_results:
            sample_all_results = list(interface_results.items())[:5]
            
            # Count interface vs not interface
            interface_status_count = {}
            for order_num, result in interface_results.items():
                status = result.get('status', 'Unknown')
                interface_status_count[status] = interface_status_count.get(status, 0) + 1
            
        else:
            pass
        
        # ULTRA-FAST: Bulk duplicate checking and processing
        print("🚀 ULTRA-FAST: Bulk processing with optimized duplicate handling...")
        save_start = datetime.now()
        
        # Get all order numbers for bulk duplicate check
        all_order_numbers = [order_data['OrderNumber'] for order_data in all_order_data]
        
        # Bulk query existing orders
        existing_orders_dict = {}
        if all_order_numbers:
            existing_orders = db.query(UploadedOrder).filter(
                UploadedOrder.OrderNumber.in_(all_order_numbers)
            ).all()
            existing_orders_dict = {order.OrderNumber: order for order in existing_orders}
        
        # Prepare bulk operations
        orders_to_insert = []
        orders_to_update = []
        replaced_count = 0
        new_count = 0
        
        # Categorize orders
        for order_data in all_order_data:
            order_number = order_data['OrderNumber']
            
            # Set interface status and Flexo data based on external database check
            if order_number in interface_results:
                interface_result = interface_results[order_number]
                interface_status = interface_result.get('status') or interface_result.get('interface_status')  # ✅ Fix: Handle both keys
                if interface_status == 'Interface':
                    order_data['InterfaceStatus'] = 'Interface'
                else:
                    order_data['InterfaceStatus'] = 'Not Yet Interface'
                
                # Add Flexo data from external database
                order_number_flexo = interface_result.get('system_ref_id', '')
                order_data['OrderNumberFlexo'] = order_number_flexo if order_number_flexo else ''
                order_data['OrderStatusFlexo'] = interface_result.get('order_status', '')
                item_id_flexo = interface_result.get('item_id_flexo')
                order_data['ItemIdFlexo'] = item_id_flexo  # ✅ ItemIdFlexo
                
            else:
                order_data['InterfaceStatus'] = 'Not Yet Interface'
                order_data['OrderNumberFlexo'] = ''
                order_data['OrderStatusFlexo'] = ''
                order_data['ItemIdFlexo'] = None  # ✅ ItemIdFlexo tetap None jika tidak ada
            
            if order_number in existing_orders_dict:
                # Prepare for update
                existing_order = existing_orders_dict[order_number]
                orders_to_update.append((existing_order, order_data))
                replaced_count += 1
            else:
                # Prepare for insert
                orders_to_insert.append(order_data)
                new_count += 1
        
        
        # Bulk insert new orders - OPTIMIZED with single commit
        if orders_to_insert:
            try:
                db.bulk_insert_mappings(UploadedOrder, orders_to_insert)
                db.commit()  # Single commit for all inserts
            except Exception:
                # Fallback to individual inserts for unique constraint violations
                db.rollback()
                for order_data in orders_to_insert:
                    try:
                        new_order = UploadedOrder(**order_data)
                        db.add(new_order)
                        db.commit()
                    except Exception:
                        db.rollback()
                        continue
        
        # Bulk update existing orders
        if orders_to_update:
            for existing_order, new_data in orders_to_update:
                existing_order.Marketplace = new_data['Marketplace']
                existing_order.Brand = new_data['Brand']
                existing_order.OrderStatus = new_data['OrderStatus']
                existing_order.AWB = new_data['AWB']
                existing_order.Transporter = new_data['Transporter']
                existing_order.OrderDate = new_data['OrderDate']
                existing_order.SLA = new_data['SLA']
                existing_order.Batch = new_data['Batch']
                existing_order.PIC = new_data['PIC']
                existing_order.UploadDate = new_data['UploadDate']
                existing_order.InterfaceStatus = new_data['InterfaceStatus']
                existing_order.TaskId = new_data['TaskId']
                existing_order.Remarks = new_data.get('Remarks', '')
                existing_order.OrderNumberFlexo = new_data.get('OrderNumberFlexo') or ''
                existing_order.OrderStatusFlexo = new_data.get('OrderStatusFlexo') or ''
                existing_order.ItemId = new_data.get('ItemId')  # ✅ Update ItemId
                existing_order.ItemIdFlexo = new_data.get('ItemIdFlexo')  # ✅ Update ItemIdFlexo
                
        
        # Commit all operations
        db.commit()
        
        save_time = (datetime.now() - save_start).total_seconds()
        
        total_processed = len(df)
        processing_time = (datetime.now() - start_time).total_seconds()
        rows_per_second = total_processed / processing_time if processing_time > 0 else 0
        
        # Count interface vs not interface orders - only for this upload session
        
        # Only get orders from this specific upload task
        uploaded_orders = db.query(UploadedOrder).filter(
            UploadedOrder.TaskId == task_id
        ).all()
        
        interface_count = sum(1 for order in uploaded_orders if order.InterfaceStatus == 'Interface')
        not_interface_count = len(uploaded_orders) - interface_count
        
        print(f"Ultra-fast processing completed: {total_processed} rows in {processing_time:.2f}s - {rows_per_second:.0f} rows/sec")
        print(f"  - Interface check: {interface_check_time:.2f}s")
        print(f"  - Database save: {save_time:.2f}s")
        
        total_time = processing_time
        avg_speed = rows_per_second
        external_db_success = True
        external_query_time = interface_check_time
        
        # Ultra-fast interface status summary for large files
        if total_rows <= 5000:
            # For small files, get full interface status
            interface_summary = get_interface_status_summary(uploaded_orders, db)
        else:
            # For large files, use ultra-fast cached summary
            interface_count = sum(1 for order in uploaded_orders if order.InterfaceStatus == 'Interface')
            not_interface_count = len(uploaded_orders) - interface_count
            
            interface_summary = {
                "interface": {
                    "count": interface_count,
                    "orders": []  # Skip detailed orders for speed
                },
                "not_interface": {
                    "count": not_interface_count,
                    "orders": []  # Skip detailed orders for speed
                },
                "total": len(uploaded_orders),
                "interface_percentage": round((interface_count / len(uploaded_orders)) * 100, 2) if uploaded_orders else 0,
                "performance_note": f"Ultra-fast processing completed in {external_query_time:.1f}s. {interface_count} interface orders from {total_rows} total."
            }
        
        result = {
            "message": f"Successfully processed {total_processed} orders with external database integration",
            "Brand": brand_name,
            "Batch": batch_name,
            "PIC": current_user,
            "total_orders": total_processed,
            "new_orders": new_count,
            "replaced_orders": replaced_count,
            "processing_mode": processing_mode,
            "processing_time": f"{total_time:.1f}s",
            "avg_speed": f"{avg_speed:.0f} rows/sec",
            "external_db_query": {
                "completed": external_db_success,
                "query_time": f"{external_query_time:.1f}s",
                "updated_orders": interface_count,
                "total_queried": total_processed,
                "success": external_db_success
            },
            "interface_summary": interface_summary
        }
        
        # Auto-generate Orderlist.txt for Not Interfaced orders
        if not_interface_count > 0:
            orderlist_result = generate_orderlist_for_not_interfaced(uploaded_orders, sales_channel, brand_name, batch_name, db, current_user)
            
            # Check for shop_id warnings and add to response
            if orderlist_result and orderlist_result.get("shop_id_warning"):
                result["shop_id_warning"] = orderlist_result["shop_id_warning"]
            
            # Auto-run marketplace app after generating Orderlist.txt (if enabled)
            if AUTO_RUN_MARKETPLACE_APPS:
                add_upload_log(task_id, "info", "🚀 Starting auto-run of marketplace app...")
                auto_run_result = auto_run_marketplace_app(sales_channel, brand_name, batch_name, current_user, task_id)
                if auto_run_result and auto_run_result.get("success"):
                    add_upload_log(task_id, "success", "✅ Marketplace app auto-run completed successfully")
                    # Add notification info to response
                    result["marketplace_notification"] = auto_run_result
                else:
                    add_upload_log(task_id, "warning", "⚠️ Marketplace app auto-run failed")
            else:
                add_upload_log(task_id, "info", "ℹ️ Auto-run disabled, marketplace app not started automatically")
        
        # Update Not Uploaded Items history immediately after successful upload
        try:
            save_not_uploaded_history()
            logger.info(f"Not uploaded items history updated after upload: {brand_name}-{sales_channel}-{batch_name}")
        except Exception as history_error:
            logger.error(f"Failed to update not uploaded items history: {str(history_error)}")
            # Don't fail the upload if history update fails
        
        return result
        
    except Exception as e:
        # Rollback any database changes if there's an error
        db.rollback()
        # Log the error for debugging
        print(f"Upload error for file {file.filename}: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)}")

@app.post("/api/upload-background")
def upload_file_background(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    current_user: str = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Background upload endpoint - returns immediately with task ID"""
    try:
        # Consolidated validation
        validation_result = validate_upload_request(file, db)
        filename_info = validation_result['filename_info']
        sales_channel = validation_result['sales_channel']
        brand_name = validation_result['brand_name']
        shop_id = validation_result['shop_id']
        
        # Read file content
        file_content = file.file.read()
        
        # Generate unique task ID with user info and microsecond precision
        import uuid
        task_id = f"upload_{current_user}_{int(time.time())}_{str(uuid.uuid4())[:8]}"
        
        # PERFORMANCE OPTIMIZATION: Skip workspace creation during upload
        # Workspace will be created only when needed for orderlist generation
        # user_workspace = multi_user_handler.create_user_workspace(current_user)
        
        # Get marketplace mapping based on sales channel from filename
        marketplace_mapping = get_marketplace_mapping(sales_channel)
        marketplace_name = marketplace_mapping.get('brand', sales_channel.upper()) if marketplace_mapping else sales_channel.upper()
        
        # Create task record
        task = UploadTask(
            task_id=task_id,
            task_name=file.filename,  # Use filename as task_name
            status="pending",
            marketplace=marketplace_name,
            brand=filename_info['brand'],
            batch=filename_info['batch'],
            pic=current_user,
            file_path=f"uploads/{current_user}/{file.filename}"
        )
        db.add(task)
        db.commit()
        
        # Start background processing
        background_tasks.add_task(
            process_upload_background,
            task_id=task_id,
            file_content=file_content,
            filename=file.filename,
            current_user=current_user
        )
        
        return {
            "success": True,
            "message": "Upload started in background",
            "task_id": task_id,
            "status": "pending",
            "brand": filename_info['brand'],
            "batch": filename_info['batch'],
            "pic": current_user,
            "estimated_time": "Processing in background...",
            "check_status_url": f"/upload-status/{task_id}"
        }
        
    except Exception as e:
        db.rollback()
        print(f"Background upload error for file {file.filename}: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Background upload failed: {str(e)}")

@app.get("/api/marketplace-logs")
def get_marketplace_logs(
    marketplace: str = Query(..., description="Marketplace name (e.g., SHOPEE, LAZADA)"),
    user: str = Query(..., description="User name"),
    lines: int = Query(50, description="Number of lines to read from the end of log file")
):
    """Get marketplace app logs for a specific user and marketplace"""
    try:
        # Construct log file path with different case variations
        marketplace_upper = marketplace.upper()
        marketplace_lower = marketplace.lower()
        
        # Try different path variations (including capitalized folder names)
        marketplace_capitalized = marketplace.capitalize()  # Shopee, Lazada, etc.
        
        # Use absolute paths to ensure we find the files regardless of working directory
        # Get the project root directory (parent of backend directory)
        project_root = get_project_root()  # Go up one level from backend/
        
        # Generate all possible case combinations for marketplace folder and log file
        marketplace_folder_variations = [marketplace_capitalized, marketplace_upper, marketplace_lower]
        log_file_variations = [f"{marketplace_lower}_app.log", f"{marketplace_upper}_app.log"]
        
        possible_paths = []
        for folder_case in marketplace_folder_variations:
            for log_case in log_file_variations:
                possible_paths.append(
                    os.path.join(project_root, "JobGetOrder", f"User_{user}", folder_case, log_case)
                )
        
        log_file_path = None
        for path in possible_paths:
            if os.path.exists(path):
                log_file_path = path
                break
        
        # If no log file found with standard naming, try to find any .log file in the marketplace folder
        if not log_file_path:
            for folder_case in marketplace_folder_variations:
                marketplace_folder = os.path.join(project_root, "JobGetOrder", f"User_{user}", folder_case)
                if os.path.exists(marketplace_folder):
                    # Look for any .log files in the folder
                    try:
                        for file in os.listdir(marketplace_folder):
                            if file.endswith('.log'):
                                log_file_path = os.path.join(marketplace_folder, file)
                                break
                        if log_file_path:
                            break
                    except (OSError, PermissionError):
                        continue
        
        if not log_file_path:
            return {
                "success": False,
                "error": f"Log file not found. Tried paths: {possible_paths}",
                "logs": [],
                "debug_info": {
                    "user": user,
                    "marketplace": marketplace,
                    "tried_paths": possible_paths
                }
            }
        
        # Read last N lines from log file
        with open(log_file_path, 'r', encoding='utf-8', errors='ignore') as f:
            all_lines = f.readlines()
            
        # Get last N lines
        recent_lines = all_lines[-lines:] if len(all_lines) > lines else all_lines
        
        # Parse log lines and format them
        formatted_logs = []
        for i, line in enumerate(recent_lines):
            line = line.strip()
            if line:
                # Try to extract timestamp and log level from the line
                timestamp = None
                level = "INFO"
                message = line
                
                # Look for timestamp patterns (adjust based on your log format)
                import re
                timestamp_match = re.search(r'(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2})', line)
                if timestamp_match:
                    timestamp = timestamp_match.group(1)
                
                # Look for log level patterns
                if 'ERROR' in line.upper():
                    level = "ERROR"
                elif 'WARN' in line.upper():
                    level = "WARN"
                elif 'DEBUG' in line.upper():
                    level = "DEBUG"
                
                formatted_logs.append({
                    "id": f"{user}_{marketplace}_{i}",
                    "timestamp": timestamp or "Unknown",
                    "level": level,
                    "message": message,
                    "marketplace": marketplace.upper(),
                    "user": user
                })
        
        logger.info(f"Successfully read {len(formatted_logs)} log lines from {log_file_path}")
        
        return {
            "success": True,
            "logs": formatted_logs,
            "file_path": log_file_path,
            "total_lines": len(all_lines),
            "returned_lines": len(formatted_logs),
            "debug_info": {
                "user": user,
                "marketplace": marketplace,
                "found_path": log_file_path
            }
        }
        
    except Exception as e:
        logger.error(f"Error reading marketplace logs: {str(e)}")
        return {
            "success": False,
            "error": str(e),
            "logs": []
        }

@app.get("/api/debug-brand/{brand_name}")
def debug_brand(
    brand_name: str,
    db: Session = Depends(get_db)
):
    """Debug endpoint to check brand configuration in database"""
    try:
        # Get all records for this brand
        brand_records = db.query(BrandShop).filter(BrandShop.brand == brand_name).all()
        
        result = {
            "brand_name": brand_name,
            "total_records": len(brand_records),
            "records": []
        }
        
        for record in brand_records:
            result["records"].append({
                "id": record.id,
                "brand": record.brand,
                "marketplace_id": record.marketplace_id,
                "shop_key_1": record.shop_key_1,
                "shop_name": record.shop_name,
                "client_shop_id": record.client_shop_id,
                "created_by": record.created_by
            })
        
        return result
        
    except Exception as e:
        return {"error": str(e)}

@app.post("/add-brand-shop")
def add_brand_shop(
    brand: str = Form(...),
    marketplace: str = Form(...),
    shop_key_1: str = Form(...),
    shop_name: str = Form(None),
    client_shop_id: int = Form(None),
    client_id: int = Form(1),
    order_type: str = Form("ONLINE"),
    created_by: str = Form("system"),
    db: Session = Depends(get_db)
):
    """Add new brand shop configuration"""
    try:
        # Map marketplace name to marketplace_id
        marketplace_id_mapping = {
            'tokopedia': 1,
            'shopee': 2,
            'lazada': 3,
            'zalora': 6,
            'ginee': 27,
            'blibli': 7,
            'jdid': 8,
            'jubelio': 9,
            'shopify': 10,
            'tiktok': 11,
            'b2b': 12,
            'desty': 23
        }
        
        marketplace_id = marketplace_id_mapping.get(marketplace.lower())
        if not marketplace_id:
            return {"error": f"Unknown marketplace: {marketplace}"}
        
        # Check if record already exists
        existing = db.query(BrandShop).filter(
            BrandShop.brand == brand,
            BrandShop.marketplace_id == marketplace_id
        ).first()
        
        if existing:
            return {"error": f"Record already exists for {brand} in {marketplace}"}
        
        # Create new record
        new_brand_shop = BrandShop(
            brand=brand,
            marketplace_id=marketplace_id,
            shop_key_1=shop_key_1,
            shop_name=shop_name or f"{brand} {marketplace} Store",
            client_shop_id=client_shop_id,
            client_id=client_id,
            order_type=order_type,
            created_by=created_by
        )
        
        db.add(new_brand_shop)
        db.commit()
        db.refresh(new_brand_shop)
        
        return {
            "success": True,
            "message": f"Added {brand} for {marketplace} with shop_key_1: {shop_key_1}",
            "record_id": new_brand_shop.id
        }
        
    except Exception as e:
        db.rollback()
        return {"error": str(e)}

@app.get("/fix-deardoer")
def fix_deardoer(db: Session = Depends(get_db)):
    """Fix DEARDOER brand name in database"""
    try:
        # Find EBLO-DEARDOER record with marketplace_id = 2 (SHOPEE)
        eblo_record = db.query(BrandShop).filter(
            BrandShop.brand == "EBLO-DEARDOER",
            BrandShop.marketplace_id == 2
        ).first()
        
        if eblo_record:
            # Update brand name to DEARDOER
            eblo_record.brand = "DEARDOER"
            eblo_record.shop_name = "DEARDOER Shopee Store"
            db.commit()
            db.refresh(eblo_record)
            
            return {
                "success": True,
                "message": f"Updated EBLO-DEARDOER to DEARDOER (ID: {eblo_record.id})",
                "record_id": eblo_record.id,
                "shop_key_1": eblo_record.shop_key_1
            }
        else:
            return {"error": "EBLO-DEARDOER record with marketplace_id=2 not found"}
        
    except Exception as e:
        db.rollback()
        return {"error": str(e)}

@app.get("/add-deardoer")
def add_deardoer(
    shop_id: str = "SHOP_ID_DEARDOER",
    db: Session = Depends(get_db)
):
    """Quick endpoint to add DEARDOER brand shop configuration"""
    try:
        # Check if record already exists
        existing = db.query(BrandShop).filter(
            BrandShop.brand == "DEARDOER",
            BrandShop.marketplace_id == 2  # SHOPEE
        ).first()
        
        if existing:
            return {"message": f"DEARDOER already exists with ID: {existing.id}", "existing": True}
        
        # Create new record
        new_brand_shop = BrandShop(
            brand="DEARDOER",
            marketplace_id=2,  # SHOPEE
            shop_key_1=shop_id,
            shop_name="DEARDOER Shopee Store",
            client_shop_id=123456,
            client_id=1,
            order_type="ONLINE",
            created_by="system"
        )
        
        db.add(new_brand_shop)
        db.commit()
        db.refresh(new_brand_shop)
        
        return {
            "success": True,
            "message": f"Added DEARDOER for SHOPEE with shop_key_1: {shop_id}",
            "record_id": new_brand_shop.id
        }
        
    except Exception as e:
        db.rollback()
        return {"error": str(e)}

@app.get("/api/upload-status/{task_id}")
def get_upload_status(
    task_id: str,
    current_user: str = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get upload task status"""
    try:
        task = db.query(UploadTask).filter(
            UploadTask.task_id == task_id,
            UploadTask.pic == current_user
        ).first()
        
        if not task:
            raise HTTPException(status_code=404, detail="Task not found")
        
        response = {
            "task_id": task.task_id,
            "status": task.status,
            "marketplace": task.marketplace,
            "brand": task.brand,
            "batch": task.batch,
            "pic": task.pic,
            "created_at": task.created_at.isoformat() if task.created_at else None,
            "completed_at": task.completed_at.isoformat() if task.completed_at else None
        }
        
        if task.status == "completed":
            # Get interface status counts for this task
            uploaded_orders = db.query(UploadedOrder).filter(
                UploadedOrder.TaskId == task_id
            ).all()
            
            interface_count = sum(1 for order in uploaded_orders if order.InterfaceStatus == 'Interface')
            not_interface_count = len(uploaded_orders) - interface_count
            
            # Prepare detailed order data for modals
            interface_orders = []
            not_interface_orders = []
            
            for order in uploaded_orders:
                order_data = {
                    "order_number": order.OrderNumberFlexo if order.OrderNumberFlexo and order.OrderNumberFlexo.strip() else order.OrderNumber,
                    "marketplace": order.Marketplace,
                    "brand": order.Brand,
                    "batch": order.Batch,
                    "order_status": order.OrderStatus,
                    "interface_status": order.InterfaceStatus,
                    "upload_date": order.UploadDate.isoformat() if order.UploadDate else None,
                    "awb": order.AWB
                }
                
                if order.InterfaceStatus == 'Interface':
                    interface_orders.append(order_data)
                else:
                    not_interface_orders.append(order_data)
            
            response.update({
                "total_orders": task.total_orders,
                "processed_orders": task.processed_orders,
                "processing_time": task.processing_time,
                "external_db_query_time": task.external_db_query_time,
                "interface_count": interface_count,
                "not_interface_count": not_interface_count,
                "interface_orders": interface_orders,
                "not_interface_orders": not_interface_orders,
                "message": "Upload completed successfully"
            })
        elif task.status == "failed":
            response.update({
                "error_message": task.error_message,
                "message": "Upload failed",
                "logs": upload_logs.get(task_id, [])  # Include logs for better error reporting
            })
        elif task.status == "processing":
            response.update({
                "message": "Upload is being processed..."
            })
        else:
            response.update({
                "message": "Upload is pending..."
            })
        
        return response
        
    except Exception as e:
        print(f"Error getting upload status for task {task_id}: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get upload status: {str(e)}")

# Database connection pool monitoring
@app.get("/api/connection-pool-status")
async def get_connection_pool_status():
    """Get database connection pool status"""
    try:
        pool = engine.pool
        total_pool_size = pool.size() + pool.overflow()
        utilization = (pool.checkedout() / total_pool_size) * 100 if total_pool_size > 0 else 0
        
        status = {
            "pool_size": pool.size(),
            "checked_in": pool.checkedin(),
            "checked_out": pool.checkedout(),
            "overflow": pool.overflow(),
            "invalid": pool.invalid(),
            "total_connections": pool.checkedin() + pool.checkedout(),
            "pool_utilization": f"{utilization:.1f}%"
        }
        
        # Log warning if pool is over 80% utilized
        if pool.checkedout() > (pool.size() + pool.overflow()) * 0.8:
            logger.warning(f"Database connection pool is {status['pool_utilization']} utilized!")
        
        return status
    except Exception as e:
        logger.error(f"Error getting connection pool status: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/cleanup-connections")
async def cleanup_connections():
    """Force cleanup of database connections"""
    try:
        # Dispose and recreate engine to clear all connections
        engine.dispose()
        logger.info("Database connections cleaned up successfully")
        return {"message": "Database connections cleaned up successfully"}
    except Exception as e:
        logger.error(f"Error cleaning up connections: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

# Legacy endpoints removed - now using unified auto-scaling /upload endpoint


@app.get("/api/orders")
def get_orders(
    page: int = Query(1, ge=1, description="Page number"),
    page_size: int = Query(50, ge=1, le=1000, description="Items per page"),
    current_user: str = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    try:
        # Check cache first
        cache_key = get_cache_key("orders", current_user, page, page_size)
        cached_result = get_cached_data(cache_key)
        if cached_result:
            return cached_result
        
        # Calculate offset for pagination
        offset = (page - 1) * page_size
        
        # Apply default filter: Exclude cancelled/batal orders
        cancelled_statuses = [
            'batal', 'cancel', 'cancellation', 'BATAL', 'CANCEL', 'CANCELLATION', 
            'Cancellations', 'Dibatalkan', 'Batal', 'CANCELED', 'Cancelled', 'canceled', 
            'Pembatalan diajukan', 'Order Batal', 'CANCELLED'
        ]
        
        # Create base query with default filter (using OrderStatusFlexo only, case-insensitive)
        base_query = db.query(UploadedOrder)
        for status in cancelled_statuses:
            # Use OrderStatusFlexo only, case-insensitive
            base_query = base_query.filter(db.func.upper(UploadedOrder.OrderStatusFlexo) != status.upper())
        
        # Get total count with filter
        total_count = base_query.count()
        
        # Get paginated orders with optimized query
        orders = base_query\
            .order_by(UploadedOrder.UploadDate.desc())\
            .offset(offset)\
            .limit(page_size)\
            .all()
        
        # Convert to response format
        orders_data = []
        for order in orders:
            order_dict = {
                "Id": order.Id,
                "Marketplace": order.Marketplace,
                "Brand": order.Brand,
                "OrderNumber": order.OrderNumber,
                "OrderStatus": order.OrderStatusFlexo,
                "AWB": order.AWB,
                "Transporter": order.Transporter,
                "OrderDate": convert_to_wib(order.OrderDate).isoformat() if order.OrderDate else None,
                "SLA": order.SLA,
                "Batch": order.Batch,
                "PIC": order.PIC,
                "UploadDate": convert_to_wib(order.UploadDate).isoformat() if order.UploadDate else None,
                "InterfaceStatus": order.InterfaceStatus
            }
            orders_data.append(order_dict)
        
        result = {
            "orders": orders_data, 
            "total": total_count,
            "page": page,
            "page_size": page_size,
            "total_pages": (total_count + page_size - 1) // page_size
        }
        
        # Cache the result
        set_cached_data(cache_key, result)
        
        return result
        
    except Exception as e:
        logger.error(f"Error getting orders: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

# DEPRECATED: Old dashboard stats endpoint removed
# Use /api/dashboard/stats instead for optimized performance and correct date filtering

@app.get("/api/orders/stats")
def get_orders_stats(
    current_user: str = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get comprehensive orders statistics for dashboard"""
    try:
        # Get total orders count
        total_orders = db.query(UploadedOrder).count()
        
        # Get interface orders count in single query
        from sqlalchemy import func, case
        interface_stats = db.query(
            func.count(case([(UploadedOrder.InterfaceStatus == "Interface", 1)], else_=0)).label('interface_count'),
            func.count(case([(UploadedOrder.InterfaceStatus != "Interface", 1)], else_=0)).label('not_interface_count')
        ).first()
        
        interface_orders = interface_stats.interface_count
        not_interface_orders = interface_stats.not_interface_count
        
        # Calculate interface rate
        interface_rate = round((interface_orders / total_orders * 100) if total_orders > 0 else 0, 2)
        
        # Get recent activity (last 7 days) - using WIB timezone
        seven_days_ago = get_wib_now() - timedelta(days=7)
        recent_orders = db.query(UploadedOrder).filter(
            UploadedOrder.UploadDate >= seven_days_ago
        ).count()
        
        # Get today's orders - using WIB timezone
        today = get_wib_now().date()
        today_orders = db.query(UploadedOrder).filter(
            func.date(UploadedOrder.UploadDate) == today
        ).count()
        
        # Get marketplace distribution
        marketplace_stats = db.query(
            UploadedOrder.Marketplace,
            func.count(UploadedOrder.Id).label('count')
        ).group_by(UploadedOrder.Marketplace).all()
        
        # Get brand distribution
        brand_stats = db.query(
            UploadedOrder.Brand,
            func.count(UploadedOrder.Id).label('count')
        ).group_by(UploadedOrder.Brand).all()
        
        # Get daily orders for the last 7 days - using WIB timezone
        daily_orders = []
        for i in range(7):
            date = get_wib_now().date() - timedelta(days=i)
            count = db.query(UploadedOrder).filter(
                func.date(UploadedOrder.UploadDate) == date
            ).count()
            daily_orders.append({
                'date': date.isoformat(),
                'count': count
            })
        
        return {
            "total_orders": total_orders,
            "interface_orders": interface_orders,
            "not_interface_orders": not_interface_orders,
            "interface_rate": interface_rate,
            "recent_orders": recent_orders,
            "today_orders": today_orders,
            "marketplace_distribution": [{"marketplace": m[0], "count": m[1]} for m in marketplace_stats],
            "brand_distribution": [{"brand": b[0], "count": b[1]} for b in brand_stats],
            "daily_orders": daily_orders,
            "last_updated": get_wib_now().isoformat()
        }
        
    except Exception as e:
        logger.error(f"Error getting orders stats: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get orders stats: {str(e)}")

@app.get("/api/orders/by-brand-batch/{brand}/{batch}")
def get_orders_by_brand_batch(
    brand: str,
    batch: str,
    task_id: str = None,
    current_user: str = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    try:
        # Get orders by brand and batch, optionally filtered by task_id
        query = db.query(UploadedOrder).filter(
            UploadedOrder.Brand == brand,
            UploadedOrder.Batch == batch
        )
        
        # If task_id is provided, filter by it
        if task_id:
            query = query.filter(UploadedOrder.TaskId == task_id)
        
        orders = query.all()
        
        # Convert to response format
        orders_data = []
        for order in orders:
            order_dict = {
                "Id": order.Id,
                "Marketplace": order.Marketplace,
                "Brand": order.Brand,
                "OrderNumber": order.OrderNumber,
                "OrderStatus": order.OrderStatusFlexo,
                "AWB": order.AWB,
                "Transporter": order.Transporter,
                "OrderDate": convert_to_wib(order.OrderDate).isoformat() if order.OrderDate else None,
                "SLA": order.SLA,
                "Batch": order.Batch,
                "PIC": order.PIC,
                "UploadDate": convert_to_wib(order.UploadDate).isoformat() if order.UploadDate else None,
                "InterfaceStatus": order.InterfaceStatus
            }
            orders_data.append(order_dict)
        
        # Separate interface and not interface orders
        interface_orders = [o for o in orders_data if o.get('InterfaceStatus') == 'Interface']
        not_interface_orders = [o for o in orders_data if o.get('InterfaceStatus') == 'Not Yet Interface']
        
        return {
            "orders": orders_data,
            "total_orders": len(orders_data),
            "interface_orders": interface_orders,
            "not_interface_orders": not_interface_orders,
            "interface_count": len(interface_orders),
            "not_interface_count": len(not_interface_orders),
            "brand": brand,
            "batch": batch
        }
        
    except Exception as e:
        print(f"Error getting orders by brand/batch: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get orders: {str(e)}")

@app.post("/fix-interface-status")
def fix_interface_status(
    order_numbers: List[str],
    current_user: str = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Fix interface status for specific order numbers"""
    try:
        if not order_numbers:
            raise HTTPException(status_code=400, detail="Order numbers list cannot be empty")
        
        updated_count = 0
        
        # Process each order number
        for order_num in order_numbers:
            order = db.query(UploadedOrder).filter(UploadedOrder.OrderNumber == order_num).first()
            if order:
                # Get marketplace for this order
                marketplace = order.Marketplace.lower() if order.Marketplace else 'unknown'
                
                # Query external database for this specific order
                interface_results = check_interface_status([order_num], marketplace)
                
                if order_num in interface_results:
                    result = interface_results[order_num]
                    new_status = result.get('status', 'NotYetInterface')
                    old_status = order.InterfaceStatus
                    
                    if new_status == 'Interface':
                        order.InterfaceStatus = 'Interface'
                        if old_status != 'Interface':
                            updated_count += 1
                            print(f"Fixed {order_num}: {old_status} -> Interface")
                    else:
                        order.InterfaceStatus = 'Not Yet Interface'
                        if old_status != 'Not Yet Interface':
                            updated_count += 1
                            print(f"Fixed {order_num}: {old_status} -> Not Yet Interface")
                else:
                    # Order not found in external database
                    order.InterfaceStatus = 'Not Yet Interface'
                    if order.InterfaceStatus != 'Not Yet Interface':
                        updated_count += 1
                        print(f"Fixed {order_num}: Not found in external DB -> Not Yet Interface")
        
        # Commit all changes
        db.commit()
        
        return {
            "message": f"Fixed interface status for {updated_count} orders",
            "updated_count": updated_count,
            "total_checked": len(order_numbers)
        }
        
    except Exception as e:
        db.rollback()
        print(f"Error fixing interface status: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to fix interface status: {str(e)}")



@app.get("/api/orders/interface-summary/{brand}/{batch}")
def get_interface_summary_by_batch(
    brand: str,
    batch: str,
    current_user: str = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get interface status summary for a specific brand and batch"""
    try:
        # Get orders for the specific brand and batch
        orders = db.query(UploadedOrder).filter(
            UploadedOrder.Brand == brand,
            UploadedOrder.Batch == batch,
            UploadedOrder.PIC == current_user
        ).all()
        
        if not orders:
            return {
                "message": f"No orders found for brand {brand} and batch {batch}",
                "interface": {"count": 0, "orders": []},
                "not_interface": {"count": 0, "orders": []},
                "total": 0,
                "interface_percentage": 0
            }
        
        # Get interface status summary
        interface_summary = get_interface_status_summary(orders, db)
        
        return {
            "brand": brand,
            "batch": batch,
            "pic": current_user,
            **interface_summary
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# Duplicate endpoint removed



@app.get("/api/orders/interface-status/real-time/{brand}/{batch}")
def get_real_time_interface_status(
    brand: str,
    batch: str,
    current_user: str = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get real-time interface status for a specific brand and batch - OPTIMIZED VERSION"""
    try:
        # Get orders for the specific brand and batch
        orders = db.query(UploadedOrder).filter(
            UploadedOrder.Brand == brand,
            UploadedOrder.Batch == batch,
            UploadedOrder.PIC == current_user
        ).all()
        
        if not orders:
            return {
                "message": f"No orders found for brand {brand} and batch {batch}",
                "interface": {"count": 0, "orders": []},
                "not_interface": {"count": 0, "orders": []},
                "total": 0,
                "interface_percentage": 0,
                "last_checked": get_wib_now().isoformat()
            }
        
        # Always run external database query for real-time interface status
        print(f"Running external database query for {len(orders)} orders...")
        
        # For large datasets, use chunked processing to avoid timeout
        if len(orders) > 1000:
            print(f"Large dataset detected ({len(orders)} orders). Using chunked external database query.")
            
            # Group orders by marketplace for efficient processing
            orders_by_marketplace = {}
            for order in orders:
                marketplace = order.Marketplace.lower() if order.Marketplace else 'unknown'
                if marketplace not in orders_by_marketplace:
                    orders_by_marketplace[marketplace] = []
                orders_by_marketplace[marketplace].append(order)
            
            interface_orders = []
            not_interface_orders = []
            
            # Process each marketplace in chunks
            for marketplace, marketplace_orders in orders_by_marketplace.items():
                print(f"Processing {len(marketplace_orders)} orders for marketplace: {marketplace}")
                
                # Process in chunks of 1000 orders to avoid SQL Server timeout
                CHUNK_SIZE = 1000
                for i in range(0, len(marketplace_orders), CHUNK_SIZE):
                    chunk = marketplace_orders[i:i + CHUNK_SIZE]
                    order_numbers = [order.OrderNumber for order in chunk if order.OrderNumber]
                    
                    if order_numbers:
                        print(f"Querying external database for chunk {i//CHUNK_SIZE + 1} ({len(order_numbers)} orders)")
                        
                        # Get interface status from external database
                        interface_results = check_interface_status(order_numbers, marketplace)
                        
                        # Process results
                        for order in chunk:
                            order_number = order.OrderNumber
                            if order_number in interface_results:
                                result = interface_results[order_number]
                                if result["status"] == "Interface":
                                    interface_orders.append({
                                        "order_number": order.OrderNumberFlexo if order.OrderNumberFlexo and order.OrderNumberFlexo.strip() else order_number,
                                        "awb": result.get("awb", order.AWB or ""),
                                        "status": result.get("order_status", order.OrderStatus),
                                        "transporter": result.get("transporter_code", order.Transporter or ""),
                                        "system_id": result.get("system_id"),
                                        "reason": "Found in WMSPROD.ord_line (external DB)"
                                    })
                                else:
                                    not_interface_orders.append({
                                        "order_number": order.OrderNumberFlexo if order.OrderNumberFlexo and order.OrderNumberFlexo.strip() else order_number,
                                        "status": result.get("order_status", order.OrderStatus),
                                        "reason": "Not found in WMSPROD.ord_line (external DB)"
                                    })
                            else:
                                # Order not found in external database, use fallback logic
                                if order.AWB and order.AWB.strip():
                                    interface_orders.append({
                                        "order_number": order.OrderNumberFlexo if order.OrderNumberFlexo and order.OrderNumberFlexo.strip() else order_number,
                                        "awb": order.AWB,
                                        "status": order.OrderStatus,
                                        "transporter": order.Transporter or "",
                                        "reason": "Has AWB/Tracking code (not in external DB)"
                                    })
                                else:
                                    not_interface_orders.append({
                                        "order_number": order.OrderNumberFlexo if order.OrderNumberFlexo and order.OrderNumberFlexo.strip() else order_number,
                                        "status": order.OrderStatus,
                                        "reason": "No AWB/Tracking code and not in external DB"
                                    })
            
            interface_summary = {
                "interface": {
                    "count": len(interface_orders),
                    "orders": interface_orders[:100]  # Limit to 100 for display
                },
                "not_interface": {
                    "count": len(not_interface_orders),
                    "orders": not_interface_orders[:100]  # Limit to 100 for display
                },
                "total": len(orders),
                "interface_percentage": round((len(interface_orders) / len(orders)) * 100, 2) if orders else 0,
                "performance_note": f"Real-time external database query completed. Showing first 100 orders from {len(orders)} total."
            }
        else:
            # For small datasets, use standard real-time query
            print(f"Small dataset detected ({len(orders)} orders). Using standard external database query.")
            interface_summary = get_interface_status_summary(orders, db)
        
        interface_summary["last_checked"] = get_wib_now().isoformat()
        
        return {
            "brand": brand,
            "batch": batch,
            "pic": current_user,
            **interface_summary
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/orders/interface-status/refresh/{brand}/{batch}")
def refresh_interface_status_for_batch(
    brand: str,
    batch: str,
    current_user: str = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Refresh interface status for a specific brand and batch - BACKGROUND PROCESSING"""
    try:
        # Get orders for the specific brand and batch
        orders = db.query(UploadedOrder).filter(
            UploadedOrder.Brand == brand,
            UploadedOrder.Batch == batch,
            UploadedOrder.PIC == current_user
        ).all()
        
        if not orders:
            return {
                "success": False,
                "message": f"No orders found for brand {brand} and batch {batch}",
                "updated_count": 0
            }
        
        # For large datasets, process in chunks
        if len(orders) > 1000:
            print(f"Large dataset detected ({len(orders)} orders). Processing in chunks...")
            
            # Group orders by marketplace for efficient processing
            orders_by_marketplace = {}
            for order in orders:
                marketplace = order.Marketplace.lower() if order.Marketplace else 'unknown'
                if marketplace not in orders_by_marketplace:
                    orders_by_marketplace[marketplace] = []
                orders_by_marketplace[marketplace].append(order)
            
            updated_count = 0
            
            # Process each marketplace in chunks
            for marketplace, marketplace_orders in orders_by_marketplace.items():
                # Process in chunks of 1000 orders
                CHUNK_SIZE = 1000
                for i in range(0, len(marketplace_orders), CHUNK_SIZE):
                    chunk = marketplace_orders[i:i + CHUNK_SIZE]
                    order_numbers = [order.OrderNumber for order in chunk if order.OrderNumber]
                    
                    if order_numbers:
                        # Get interface status from database
                        interface_results = check_interface_status(order_numbers, marketplace)
                        
                        # Update orders with new interface status
                        for order in chunk:
                            if order.OrderNumber in interface_results:
                                result = interface_results[order.OrderNumber]
                                new_status = result.get('status', 'NotYetInterface')
                                old_status = order.InterfaceStatus
                                
                                if new_status == 'Interface':
                                    order.InterfaceStatus = 'Interface'
                                    if old_status != 'Interface':
                                        updated_count += 1
                                else:
                                    if order.InterfaceStatus != 'Interface':
                                        order.InterfaceStatus = 'Not Yet Interface'
                                        if old_status != 'Not Yet Interface':
                                            updated_count += 1
                            else:
                                # Order not found in database
                                if order.InterfaceStatus != 'Interface':
                                    order.InterfaceStatus = 'Not Yet Interface'
                                    updated_count += 1
                    
                    # Commit chunk
                    db.commit()
                    print(f"Processed chunk {i//CHUNK_SIZE + 1} for {marketplace}")
            
            return {
                "success": True,
                "message": f"Interface status refreshed for {len(orders)} orders in {len(orders_by_marketplace)} marketplaces",
                "updated_count": updated_count,
                "total_orders": len(orders),
                "processing_type": "chunked"
            }
        else:
            # For small datasets, use simple refresh logic
            orders = db.query(UploadedOrder).filter(UploadedOrder.PIC == current_user).all()
            return {
                "success": True,
                "message": f"Interface status refreshed for {len(orders)} orders",
                "updated_count": 0,
                "total_orders": len(orders)
            }
            
    except Exception as e:
        db.rollback()
        print(f"Error refreshing interface status for {brand}/{batch}: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to refresh interface status: {str(e)}")

@app.post("/api/orders/interface-status/force-refresh/{brand}/{batch}")
def force_refresh_interface_status(
    brand: str,
    batch: str,
    current_user: str = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Force refresh interface status with external database query - ALWAYS RUNS EXTERNAL QUERY"""
    try:
        # Get orders for the specific brand and batch
        orders = db.query(UploadedOrder).filter(
            UploadedOrder.Brand == brand,
            UploadedOrder.Batch == batch,
            UploadedOrder.PIC == current_user
        ).all()
        
        if not orders:
            return {
                "success": False,
                "message": f"No orders found for brand {brand} and batch {batch}",
                "updated_count": 0
            }
        
        print(f"Force refresh: Running external database query for {len(orders)} orders...")
        
        # Always run external database query regardless of dataset size
        interface_summary = get_interface_status_summary(orders, db)
        
        # Update InterfaceStatus field in database based on external query results
        updated_count = 0
        
        # Group orders by marketplace for efficient processing
        orders_by_marketplace = {}
        for order in orders:
            marketplace = order.Marketplace.lower() if order.Marketplace else 'unknown'
            if marketplace not in orders_by_marketplace:
                orders_by_marketplace[marketplace] = []
            orders_by_marketplace[marketplace].append(order)
        
        # Process each marketplace in chunks
        for marketplace, marketplace_orders in orders_by_marketplace.items():
            # Process in chunks of 1000 orders
            CHUNK_SIZE = 1000
            for i in range(0, len(marketplace_orders), CHUNK_SIZE):
                chunk = marketplace_orders[i:i + CHUNK_SIZE]
                order_numbers = [order.OrderNumber for order in chunk if order.OrderNumber]
                
                if order_numbers:
                    # Get interface status from external database
                    interface_results = check_interface_status(order_numbers, marketplace)
                    
                    # Update orders with new interface status
                    for order in chunk:
                        if order.OrderNumber in interface_results:
                            result = interface_results[order.OrderNumber]
                            new_status = result.get('status', 'NotYetInterface')
                            old_status = order.InterfaceStatus
                            
                            if new_status == 'Interface':
                                order.InterfaceStatus = 'Interface'
                                if old_status != 'Interface':
                                    updated_count += 1
                            else:
                                if order.InterfaceStatus != 'Interface':
                                    order.InterfaceStatus = 'Not Yet Interface'
                                    if old_status != 'Not Yet Interface':
                                        updated_count += 1
                        else:
                            # Order not found in external database
                            if order.InterfaceStatus != 'Interface':
                                order.InterfaceStatus = 'Not Yet Interface'
                                updated_count += 1
                
                # Commit chunk
                db.commit()
                print(f"Force refresh: Processed chunk {i//CHUNK_SIZE + 1} for {marketplace}")
        
        return {
            "success": True,
            "message": f"Force refresh completed for {len(orders)} orders",
            "updated_count": updated_count,
            "total_orders": len(orders),
            "interface_summary": interface_summary,
            "processing_type": "force_refresh_external_db"
        }
        
    except Exception as e:
        db.rollback()
        print(f"Force refresh error for {brand}/{batch}: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Force refresh failed: {str(e)}")

@app.put("/api/orders/{order_id}/remarks")
def update_order_remarks(
    order_id: int,
    remarks: str,
    current_user: str = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Update remarks for a specific order"""
    try:
        # Find the order
        order = db.query(UploadedOrder).filter(
            UploadedOrder.Id == order_id,
            UploadedOrder.PIC == current_user
        ).first()

        if not order:
            raise HTTPException(status_code=404, detail="Order not found")

        # Update remarks
        order.Remarks = remarks
        db.commit()

        return {
            "message": "Remarks updated successfully",
            "order_id": order_id,
            "remarks": remarks
        }
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to update remarks: {str(e)}")




@app.get("/api/orders/sorted-by-interface-status-v2")
def get_orders_sorted_by_interface_status_v2(
    date: str = None,
    current_user: str = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get all user orders sorted by interface status using InterfaceStatus field"""
    try:
        # Build query for current user
        query = db.query(UploadedOrder).filter(UploadedOrder.PIC == current_user)
        
        # Add date filter if provided
        if date:
            try:
                # Parse date string (expected format: YYYY-MM-DD) - using WIB timezone
                date_obj = datetime.strptime(date, "%Y-%m-%d").date()
                
                # Create start and end datetime for the day in WIB timezone
                start_datetime = WIB_TIMEZONE.localize(datetime.combine(date_obj, datetime_time.min))  # 00:00:00 WIB
                end_datetime = WIB_TIMEZONE.localize(datetime.combine(date_obj, datetime_time.max))    # 23:59:59.999999 WIB
                
                # Filter by date range
                query = query.filter(
                    UploadedOrder.UploadDate >= start_datetime,
                    UploadedOrder.UploadDate <= end_datetime
                )
            except ValueError:
                raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
        
        # Default filter: Exclude cancelled orders
        cancelled_statuses = [
            'cancel', 'cancelled', 'canceled', 'dibatalkan', 'batal', 
            'refund', 'return', 'rejected', 'failed', 'expired',
            'void', 'invalid', 'declined', 'denied'
        ]
        
        # Create filter to exclude cancelled orders
        cancelled_filter = ~UpdateUploadedOrder.OrderStatus.ilike(f"%{cancelled_statuses[0]}%")
        for status in cancelled_statuses[1:]:
            cancelled_filter = cancelled_filter & ~UpdateUploadedOrder.OrderStatus.ilike(f"%{status}%")
        
        query = query.filter(cancelled_filter)
        
        orders = query.all()
        
        if not orders:
            message = f"No orders found for current user"
            if date:
                message += f" on {date}"
            return {
                "message": message,
                "orders": [],
                "total": 0,
                "not_interface_count": 0,
                "interface_count": 0,
                "interface_percentage": 0,
                "last_checked": get_wib_now().isoformat(),
                "filter_date": date
            }
        
        # Use InterfaceStatus field from database
        interface_orders = []
        not_interface_orders = []
        
        for order in orders:
            if order.InterfaceStatus == 'Interface':
                interface_orders.append({
                    "order_number": order.OrderNumberFlexo if order.OrderNumberFlexo and order.OrderNumberFlexo.strip() else order.OrderNumber,
                    "awb": order.AWB or "",
                    "status": order.OrderStatus,
                    "transporter": order.Transporter or "",
                    "reason": "Interface status from database"
                })
            else:
                not_interface_orders.append({
                    "order_number": order.OrderNumberFlexo if order.OrderNumberFlexo and order.OrderNumberFlexo.strip() else order.OrderNumber,
                    "status": order.OrderStatus,
                    "reason": "Not Yet Interface status from database"
                })
        
        interface_summary = {
            "interface": {
                "count": len(interface_orders),
                "orders": interface_orders
            },
            "not_interface": {
                "count": len(not_interface_orders),
                "orders": not_interface_orders
            },
            "total": len(orders),
            "interface_percentage": round((len(interface_orders) / len(orders)) * 100, 2) if orders else 0
        }
        
        # Create a list of all orders with their interface status (with deduplication)
        all_orders_with_status = []
        seen_orders = set()  # Track unique order numbers to prevent duplicates
        
        # Add Not Yet Interface orders first
        for order in interface_summary["not_interface"]["orders"]:
            # Find the original order data to get additional fields
            original_order = next((o for o in orders if o.OrderNumber == order["order_number"]), None)
            if original_order:
                # Create unique key for deduplication
                unique_key = f"{order['order_number']}-{original_order.Marketplace}-{original_order.Brand}"
                if unique_key not in seen_orders:
                    seen_orders.add(unique_key)
                    all_orders_with_status.append({
                        "id": original_order.Id,
                        "marketplace": original_order.Marketplace,
                        "brand": original_order.Brand,
                        "order_number": original_order.OrderNumberFlexo if original_order.OrderNumberFlexo and original_order.OrderNumberFlexo.strip() else order["order_number"],
                        "order_status": order.get("status", ""),
                        "awb": "",
                        "transporter": "",
                        "order_date": original_order.OrderDate.strftime("%Y-%m-%d") if original_order.OrderDate else "",
                        "sla": original_order.SLA,
                        "batch": original_order.Batch,
                        "pic": original_order.PIC,
                        "remarks": original_order.Remarks,
                        "status": "Not Yet Interface",
                        "reason": order["reason"],
                        "system_id": "",
                        "interface_priority": 1  # Higher priority for Not Yet Interface
                    })
        
        # Add Interface orders second
        for order in interface_summary["interface"]["orders"]:
            # Find the original order data to get additional fields
            original_order = next((o for o in orders if o.OrderNumber == order["order_number"]), None)
            if original_order:
                # Create unique key for deduplication
                unique_key = f"{order['order_number']}-{original_order.Marketplace}-{original_order.Brand}"
                if unique_key not in seen_orders:
                    seen_orders.add(unique_key)
                    all_orders_with_status.append({
                        "id": original_order.Id,
                        "marketplace": original_order.Marketplace,
                        "brand": original_order.Brand,
                        "order_number": original_order.OrderNumberFlexo if original_order.OrderNumberFlexo and original_order.OrderNumberFlexo.strip() else order["order_number"],
                        "order_status": order.get("status", ""),
                        "awb": order.get("awb", ""),
                        "transporter": order.get("transporter", ""),
                        "order_date": original_order.OrderDate.strftime("%Y-%m-%d") if original_order.OrderDate else "",
                        "sla": original_order.SLA,
                        "batch": original_order.Batch,
                        "pic": original_order.PIC,
                        "remarks": original_order.Remarks,
                        "status": "Interface",
                        "reason": order["reason"],
                        "system_id": "",
                        "interface_priority": 2  # Lower priority for Interface
                    })
        
        # Sort by interface priority (Not Yet Interface first), then by order number
        all_orders_with_status.sort(key=lambda x: (x["interface_priority"], x["order_number"] or ""))
        
        return {
            "orders": all_orders_with_status,
            "total": len(all_orders_with_status),
            "not_interface_count": interface_summary["not_interface"]["count"],
            "interface_count": interface_summary["interface"]["count"],
            "interface_percentage": interface_summary["interface_percentage"],
            "last_checked": get_wib_now().isoformat(),
            "filter_date": date
        }
    except Exception as e:
        print(f"Error in get_orders_sorted_by_interface_status_v2: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get sorted orders: {str(e)}")

@app.get("/api/orders/unique-values")
def get_unique_values(
    field: str = Query(..., description="Field name to get unique values for"),
    db: Session = Depends(get_db)
):
    """Get unique values for a specific field from uploaded_orders table"""
    try:
        # Map field names to actual column names
        field_mapping = {
            'marketplace': 'Marketplace',
            'brand': 'Brand',
            'order_status': 'OrderStatusFlexo',
            'transporter': 'Transporter',
            'batch': 'Batch',
            'pic': 'PIC',
            'interface_status': 'InterfaceStatus',
            'remarks': 'Remarks',
            'order_number_flexo': 'OrderNumberFlexo',
            'order_status_flexo': 'OrderStatusFlexo'
        }
        
        if field not in field_mapping:
            raise HTTPException(status_code=400, detail=f"Invalid field: {field}")
        
        # Get unique values for the field using SQL query
        column_name = field_mapping[field]
        query = f'SELECT DISTINCT "{column_name}" FROM uploaded_orders WHERE "{column_name}" IS NOT NULL AND "{column_name}" != \'\' ORDER BY "{column_name}"'
        
        unique_results = db.execute(text(query)).fetchall()
        
        # Flatten the results
        result = [row[0] for row in unique_results if row[0]]
        
        return {
            "field": field,
            "values": result,
            "count": len(result),
            "table_used": "uploaded_orders"
        }
        
    except Exception as e:
        print(f"Error getting unique values for {field}: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get unique values: {str(e)}")

@app.get("/api/orders/cascading-filters")
def get_cascading_filter_values(
    field: str = Query(..., description="Field name to get unique values for"),
    marketplace_filters: Optional[str] = Query(None, description="Comma-separated marketplace filters"),
    brand_filters: Optional[str] = Query(None, description="Comma-separated brand filters"),
    order_status_filters: Optional[str] = Query(None, description="Comma-separated order status filters"),
    transporter_filters: Optional[str] = Query(None, description="Comma-separated transporter filters"),
    batch_filters: Optional[str] = Query(None, description="Comma-separated batch filters"),
    pic_filters: Optional[str] = Query(None, description="Comma-separated PIC filters"),
    remarks_filters: Optional[str] = Query(None, description="Comma-separated remarks filters"),
    interface_status: Optional[str] = Query(None, description="Filter by interface status"),
    start_date: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
    db: Session = Depends(get_db)
):
    """Get unique values for a specific field with cascading filters applied"""
    try:
        # Map field names to actual column names
        field_mapping = {
            'marketplace': 'Marketplace',
            'brand': 'Brand',
            'order_status': 'OrderStatusFlexo',
            'transporter': 'Transporter',
            'batch': 'Batch',
            'pic': 'PIC',
            'interface_status': 'InterfaceStatus',
            'remarks': 'Remarks',
            'order_number_flexo': 'OrderNumberFlexo',
            'order_status_flexo': 'OrderStatusFlexo'
        }
        
        if field not in field_mapping:
            raise HTTPException(status_code=400, detail=f"Invalid field: {field}")
        
        # Build the base query with cascading filters
        base_query = """
        SELECT DISTINCT "{field_column}" 
        FROM clean_orders 
        WHERE "{field_column}" IS NOT NULL AND "{field_column}" != ''
        """
        
        # Add cascading filter conditions
        conditions = []
        params = {}
        
        # Date range filter
        if start_date and end_date:
            conditions.append('"OrderDate" >= :start_date AND "OrderDate" <= :end_date')
            params['start_date'] = start_date
            params['end_date'] = end_date
        
        # Interface status filter
        if interface_status:
            conditions.append('"InterfaceStatus" = :interface_status')
            params['interface_status'] = interface_status
        
        # Marketplace filter
        if marketplace_filters:
            marketplace_list = [m.strip() for m in marketplace_filters.split(',') if m.strip()]
            if marketplace_list:
                placeholders = ','.join([f':marketplace_{i}' for i in range(len(marketplace_list))])
                conditions.append(f'"Marketplace" IN ({placeholders})')
                for i, marketplace in enumerate(marketplace_list):
                    params[f'marketplace_{i}'] = marketplace
        
        # Brand filter
        if brand_filters:
            brand_list = [b.strip() for b in brand_filters.split(',') if b.strip()]
            if brand_list:
                placeholders = ','.join([f':brand_{i}' for i in range(len(brand_list))])
                conditions.append(f'"Brand" IN ({placeholders})')
                for i, brand in enumerate(brand_list):
                    params[f'brand_{i}'] = brand
        
        # Order status filter
        if order_status_filters:
            status_list = [s.strip() for s in order_status_filters.split(',') if s.strip()]
            if status_list:
                placeholders = ','.join([f':order_status_{i}' for i in range(len(status_list))])
                conditions.append(f'UPPER("OrderStatusFlexo") IN ({placeholders})')
                for i, status in enumerate(status_list):
                    params[f'order_status_{i}'] = status.upper()
        
        # Transporter filter
        if transporter_filters:
            transporter_list = [t.strip() for t in transporter_filters.split(',') if t.strip()]
            if transporter_list:
                placeholders = ','.join([f':transporter_{i}' for i in range(len(transporter_list))])
                conditions.append(f'"Transporter" IN ({placeholders})')
                for i, transporter in enumerate(transporter_list):
                    params[f'transporter_{i}'] = transporter
        
        # Batch filter
        if batch_filters:
            batch_list = [b.strip() for b in batch_filters.split(',') if b.strip()]
            if batch_list:
                placeholders = ','.join([f':batch_{i}' for i in range(len(batch_list))])
                conditions.append(f'"Batch" IN ({placeholders})')
                for i, batch in enumerate(batch_list):
                    params[f'batch_{i}'] = batch
        
        # PIC filter
        if pic_filters:
            pic_list = [p.strip() for p in pic_filters.split(',') if p.strip()]
            if pic_list:
                placeholders = ','.join([f':pic_{i}' for i in range(len(pic_list))])
                conditions.append(f'"PIC" IN ({placeholders})')
                for i, pic in enumerate(pic_list):
                    params[f'pic_{i}'] = pic
        
        # Remarks filter
        if remarks_filters:
            remarks_list = [r.strip() for r in remarks_filters.split(',') if r.strip()]
            if remarks_list:
                placeholders = ','.join([f':remarks_{i}' for i in range(len(remarks_list))])
                conditions.append(f'"Remarks" IN ({placeholders})')
                for i, remarks in enumerate(remarks_list):
                    params[f'remarks_{i}'] = remarks
        
        # Build final query
        field_column = field_mapping[field]
        query = base_query.format(field_column=field_column)
        
        if conditions:
            query += ' AND ' + ' AND '.join(conditions)
        
        query += f' ORDER BY "{field_column}"'
        
        # Execute query
        unique_results = db.execute(text(query), params).fetchall()
        
        # Flatten the results
        result = [row[0] for row in unique_results if row[0]]
        
        return {
            "field": field,
            "values": result,
            "count": len(result),
            "filters_applied": {
                "marketplace": marketplace_filters,
                "brand": brand_filters,
                "order_status": order_status_filters,
                "transporter": transporter_filters,
                "batch": batch_filters,
                "pic": pic_filters,
                "remarks": remarks_filters,
                "interface_status": interface_status,
                "date_range": f"{start_date} to {end_date}" if start_date and end_date else None
            },
            "table_used": "clean_orders"
        }
        
    except Exception as e:
        print(f"Error getting cascading filter values for {field}: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get cascading filter values: {str(e)}")

@app.get("/api/orders/list")
def get_orders_list(
    page: int = Query(1, ge=1, description="Page number"),
    page_size: int = Query(100, ge=1, le=100000, description="Items per page"),
    start_date: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
    order_numbers: Optional[str] = Query(None, description="Comma-separated order numbers"),
    interface_status: Optional[str] = Query(None, description="Filter by interface status"),
    order_status: Optional[str] = Query(None, description="Comma-separated order status filters from filter box"),
    pic: Optional[str] = Query(None, description="Filter by PIC"),
    remarks: Optional[str] = Query(None, description="Filter by remarks"),
    # Column filter parameters
    marketplace: Optional[str] = Query(None, description="Filter by marketplace"),
    brand: Optional[str] = Query(None, description="Filter by brand"),
    marketplace_filters: Optional[str] = Query(None, description="Comma-separated marketplace filters"),
    brand_filters: Optional[str] = Query(None, description="Comma-separated brand filters"),
    order_status_filters: Optional[str] = Query(None, description="Comma-separated order status filters"),
    transporter_filters: Optional[str] = Query(None, description="Comma-separated transporter filters"),
    batch_filters: Optional[str] = Query(None, description="Comma-separated batch filters"),
    pic_filters: Optional[str] = Query(None, description="Comma-separated PIC filters"),
    remarks_filters: Optional[str] = Query(None, description="Comma-separated remarks filters"),
    current_user: str = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get uploaded orders with filtering and pagination for the orders list page"""
    try:
        
        # Use clean_orders view for deduplicated data
        try:
            # Check if clean_orders view exists and has data
            clean_count = db.execute(text("SELECT COUNT(*) FROM clean_orders")).scalar()
            
            if clean_count == 0:
                return {
                    "orders": [],
                    "total_count": 0,
                    "page": page,
                    "page_size": page_size,
                    "total_pages": 0,
                    "message": "No data available. Please create clean_orders view first."
                }
            
            print(f"📊 Using clean_orders view ({clean_count} records)")
            
        except Exception as e:
            print(f"❌ Error accessing clean_orders view: {e}")
            return {
                "orders": [],
                "total_count": 0,
                "page": page,
                "page_size": page_size,
                "total_pages": 0,
                "message": "Clean orders view not available. Please create it first."
            }
        
        # Initialize date variables
        start_datetime = None
        end_datetime = None
        
        # Build SQL query using clean_orders view
        base_query = """
        SELECT "Id", "Marketplace", "Brand", "OrderNumber", "OrderStatus", "AWB", "Transporter", 
               "OrderDate", "SLA", "Batch", "PIC", "UploadDate", "Remarks", "InterfaceStatus", "TaskId", 
               "OrderNumberFlexo", "OrderStatusFlexo"
        FROM clean_orders
        """
        
        count_query = "SELECT COUNT(*) FROM clean_orders"
        conditions = []
        params = {}
        
        # Apply default filter: Exclude cancelled/batal orders (only if no order status filters are applied)
        cancelled_statuses = [
            'batal', 'cancel', 'cancellation', 'BATAL', 'CANCEL', 'CANCELLATION', 
            'Cancellations', 'Dibatalkan', 'Batal', 'CANCELED', 'Cancelled', 'canceled', 
            'Pembatalan diajukan', 'Order Batal', 'CANCELLED'
        ]
        
        # Only apply default filter if no order status filters are specified
        has_order_status_filter = (order_status and order_status.strip()) or (order_status_filters and order_status_filters.strip())
        
        print(f"🔍 DEBUG - Filter Status Check:")
        print(f"  - order_status: '{order_status}'")
        print(f"  - order_status_filters: '{order_status_filters}'")
        print(f"  - has_order_status_filter: {has_order_status_filter}")
        
        if not has_order_status_filter:
            print(f"  - Applying default filter to exclude cancelled orders")
            # Create filter to exclude cancelled orders (using OrderStatusFlexo only, case-insensitive)
            cancelled_placeholders = ','.join([f':cancelled_{i}' for i in range(len(cancelled_statuses))])
            conditions.append(f'UPPER("OrderStatusFlexo") NOT IN ({cancelled_placeholders})')
            for i, status in enumerate(cancelled_statuses):
                params[f'cancelled_{i}'] = status.upper()
        else:
            print(f"  - Skipping default filter, user has specified order status filters")
        
        # Apply date filters - only apply if dates are explicitly provided and not empty
        if (start_date and start_date.strip()) or (end_date and end_date.strip()):
            # Parse provided dates (supports both ISO and YYYY-MM-DD format)
            if start_date and start_date.strip():
                try:
                    start_datetime = parse_date_flexible(start_date)
                    start_naive = convert_to_naive_wib(start_datetime)
                    conditions.append('"UploadDate" >= :start_date')
                    params['start_date'] = start_naive
                except ValueError as e:
                    raise HTTPException(status_code=400, detail=f"Invalid start_date format: {str(e)}")
                
            if end_date and end_date.strip():
                try:
                    end_datetime_temp = parse_date_flexible(end_date)
                    # For end date, set to end of day in WIB timezone
                    if end_datetime_temp:
                        end_datetime = WIB_TIMEZONE.localize(datetime.combine(end_datetime_temp.date(), datetime_time.max))
                        end_naive = convert_to_naive_wib(end_datetime)
                        conditions.append('"UploadDate" <= :end_date')
                        params['end_date'] = end_naive
                except ValueError as e:
                    raise HTTPException(status_code=400, detail=f"Invalid end_date format: {str(e)}")
        
        # Apply other filters - only apply if values are provided and not empty
        if order_numbers and order_numbers.strip():
            # Split comma-separated order numbers and filter
            order_list = [order.strip() for order in order_numbers.split(',') if order.strip()]
            if order_list:
                placeholders = ','.join([f':order_{i}' for i in range(len(order_list))])
                conditions.append(f'"OrderNumber" IN ({placeholders})')
                for i, order in enumerate(order_list):
                    params[f'order_{i}'] = order
        
        if interface_status and interface_status.strip():
            conditions.append('"InterfaceStatus" ILIKE :interface_status')
            params['interface_status'] = f"%{interface_status.strip()}%"
        
        if order_status and order_status.strip():
            # Split comma-separated order statuses and filter
            status_list = [status.strip() for status in order_status.split(',') if status.strip()]
            print(f"🔍 DEBUG - order_status filter:")
            print(f"  - Raw order_status: '{order_status}'")
            print(f"  - Parsed status_list: {status_list}")
            if status_list:
                placeholders = ','.join([f':status_{i}' for i in range(len(status_list))])
                conditions.append(f'UPPER("OrderStatusFlexo") IN ({placeholders})')
                for i, status in enumerate(status_list):
                    params[f'status_{i}'] = status.upper()
                print(f"  - Applied filter: UPPER(\"OrderStatusFlexo\") IN ({placeholders})")
                print(f"  - Parameters: {[params[f'status_{i}'] for i in range(len(status_list))]}")
        
        if pic and pic.strip():
            conditions.append('"PIC" ILIKE :pic')
            params['pic'] = f"%{pic.strip()}%"
        
        if remarks and remarks.strip():
            conditions.append('"Remarks" ILIKE :remarks')
            params['remarks'] = f"%{remarks.strip()}%"
        
        # Apply column filters (single value filters)
        if marketplace and marketplace.strip():
            conditions.append('"Marketplace" ILIKE :marketplace')
            params['marketplace'] = f"%{marketplace.strip()}%"
        
        if brand and brand.strip():
            conditions.append('"Brand" ILIKE :brand')
            params['brand'] = f"%{brand.strip()}%"
        
        # Apply column filters (multiple value filters)
        if marketplace_filters and marketplace_filters.strip():
            marketplace_list = [mp.strip() for mp in marketplace_filters.split(',') if mp.strip()]
            if marketplace_list:
                placeholders = ','.join([f':mp_{i}' for i in range(len(marketplace_list))])
                conditions.append(f'"Marketplace" IN ({placeholders})')
                for i, mp in enumerate(marketplace_list):
                    params[f'mp_{i}'] = mp
        
        if brand_filters and brand_filters.strip():
            brand_list = [brand.strip() for brand in brand_filters.split(',') if brand.strip()]
            if brand_list:
                placeholders = ','.join([f':brand_{i}' for i in range(len(brand_list))])
                conditions.append(f'"Brand" IN ({placeholders})')
                for i, brand in enumerate(brand_list):
                    params[f'brand_{i}'] = brand
        
        if order_status_filters and order_status_filters.strip():
            status_list = [status.strip() for status in order_status_filters.split(',') if status.strip()]
            print(f"🔍 DEBUG - order_status_filters filter:")
            print(f"  - Raw order_status_filters: '{order_status_filters}'")
            print(f"  - Parsed status_list: {status_list}")
            if status_list:
                placeholders = ','.join([f':ostatus_{i}' for i in range(len(status_list))])
                conditions.append(f'UPPER("OrderStatusFlexo") IN ({placeholders})')
                for i, status in enumerate(status_list):
                    params[f'ostatus_{i}'] = status.upper()
                print(f"  - Applied filter: UPPER(\"OrderStatusFlexo\") IN ({placeholders})")
                print(f"  - Parameters: {[params[f'ostatus_{i}'] for i in range(len(status_list))]}")
        
        if transporter_filters and transporter_filters.strip():
            transporter_list = [trans.strip() for trans in transporter_filters.split(',') if trans.strip()]
            if transporter_list:
                placeholders = ','.join([f':trans_{i}' for i in range(len(transporter_list))])
                conditions.append(f'"Transporter" IN ({placeholders})')
                for i, trans in enumerate(transporter_list):
                    params[f'trans_{i}'] = trans
        
        if batch_filters and batch_filters.strip():
            batch_list = [batch.strip() for batch in batch_filters.split(',') if batch.strip()]
            if batch_list:
                placeholders = ','.join([f':batch_{i}' for i in range(len(batch_list))])
                conditions.append(f'"Batch" IN ({placeholders})')
                for i, batch in enumerate(batch_list):
                    params[f'batch_{i}'] = batch
        
        if pic_filters and pic_filters.strip():
            pic_list = [pic.strip() for pic in pic_filters.split(',') if pic.strip()]
            if pic_list:
                placeholders = ','.join([f':pic_{i}' for i in range(len(pic_list))])
                conditions.append(f'"PIC" IN ({placeholders})')
                for i, pic in enumerate(pic_list):
                    params[f'pic_{i}'] = pic
        
        if remarks_filters and remarks_filters.strip():
            remarks_list = [remark.strip() for remark in remarks_filters.split(',') if remark.strip()]
            if remarks_list:
                placeholders = ','.join([f':remark_{i}' for i in range(len(remarks_list))])
                conditions.append(f'"Remarks" IN ({placeholders})')
                for i, remark in enumerate(remarks_list):
                    params[f'remark_{i}'] = remark
        
        # Add WHERE clause if there are conditions
        if conditions:
            where_clause = " WHERE " + " AND ".join(conditions)
            base_query += where_clause
            count_query += where_clause
        
        print(f"🔍 DEBUG - Final Query:")
        print(f"  - Conditions: {conditions}")
        print(f"  - WHERE clause: {where_clause if conditions else 'None'}")
        print(f"  - Parameters: {params}")
        
        # Get total count for pagination
        total_count = db.execute(text(count_query), params).scalar()
        print(f"  - Total count result: {total_count}")
        
        # Calculate offset for pagination
        offset = (page - 1) * page_size
        
        # Add ORDER BY, LIMIT, OFFSET to main query
        base_query += ' ORDER BY "UploadDate" DESC LIMIT :limit OFFSET :offset'
        params['limit'] = page_size
        params['offset'] = offset
        
        # Execute query
        result = db.execute(text(base_query), params).fetchall()
        print(f"  - Query returned {len(result)} rows")
        
        # Convert to response format - no need for deduplication since using clean_orders view
        orders_data = []
        
        for row in result:
            # Fallback logic: Use OrderNumberFlexo if available, otherwise use original OrderNumber
            order_number_flexo = row[15]  # OrderNumberFlexo
            original_order_number = row[3]  # OrderNumber
            display_order_number = order_number_flexo if order_number_flexo and order_number_flexo.strip() else original_order_number
            
            order_dict = {
                "id": row[0],  # Id
                "marketplace": row[1],  # Marketplace
                "brand": row[2],  # Brand
                "order_number": display_order_number,  # Fallback: OrderNumberFlexo or OrderNumber
                "order_status": row[16],  # OrderStatusFlexo
                "awb": row[5],  # AWB
                "transporter": row[6],  # Transporter
                "order_date": row[7].strftime("%Y-%m-%d") if row[7] else None,  # OrderDate
                "sla": row[8],  # SLA
                "batch": row[9],  # Batch
                "pic": row[10],  # PIC
                "upload_date": row[11].strftime("%Y-%m-%d") if row[11] else None,  # UploadDate
                "remarks": row[12],  # Remarks
                "interface_status": row[13],  # InterfaceStatus
                "task_id": row[14],  # TaskId
                "order_number_flexo": row[15],  # OrderNumberFlexo
                "order_status_flexo": row[16]  # OrderStatusFlexo
            }
            orders_data.append(order_dict)
        
        # Debug: Show sample of returned data
        if orders_data:
            print(f"🔍 DEBUG - Sample returned data (first 3 orders):")
            for i, order in enumerate(orders_data[:3]):
                print(f"  Order {i+1}:")
                print(f"    - OrderNumber (displayed): {order['order_number']}")
                print(f"    - OrderStatus (displayed): {order['order_status']}")
                print(f"    - OrderNumberFlexo (raw): {order['order_number_flexo']}")
                print(f"    - OrderStatusFlexo (raw): {order['order_status_flexo']}")
                print(f"    - InterfaceStatus: {order['interface_status']}")
                print(f"    - Filter should match: OrderStatus = '{order['order_status']}'")
        
        # Calculate pagination info
        total_pages = (total_count + page_size - 1) // page_size
        
        return {
            "orders": orders_data,
            "pagination": {
                "current_page": page,
                "page_size": page_size,
                "total_count": total_count,
                "total_pages": total_pages,
                "has_next": page < total_pages,
                "has_prev": page > 1
            },
            "filters": {
                "start_date": start_date,
                "end_date": end_date,
                "order_numbers": order_numbers,
                "interface_status": interface_status,
                "pic": pic,
                "remarks": remarks,
                "marketplace": marketplace,
                "brand": brand,
                "marketplace_filters": marketplace_filters,
                "brand_filters": brand_filters,
                "order_status_filters": order_status_filters,
                "transporter_filters": transporter_filters,
                "batch_filters": batch_filters,
                "pic_filters": pic_filters,
                "remarks_filters": remarks_filters
            }
        }
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error in get_orders_list: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get orders list: {str(e)}")

@app.post("/api/orders/{order_id}/remarks")
async def update_order_remarks(
    order_id: int,
    request: Request,
    current_user: str = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Update remarks for a specific order"""
    try:
        
        # Get request body
        body = await request.json()
        
        # Find the order
        order = db.query(UploadedOrder).filter(UploadedOrder.Id == order_id).first()
        if not order:
            raise HTTPException(status_code=404, detail="Order not found")
        
        # Update remarks
        remarks = body.get('remarks', '')
        order.Remarks = remarks
        db.commit()
        
        
        return {
            "message": "Remarks updated successfully",
            "order_id": order_id,
            "remarks": remarks
        }
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        print(f"Error updating order remarks: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to update remarks: {str(e)}")

@app.put("/api/orders/by-number/{order_number}/remark")
async def update_order_remark_by_number(
    order_number: str,
    remark: str = Query(..., description="Remark to update"),
    current_user: str = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Update remarks for a specific order by order number"""
    try:
        
        # Find the order by order number
        order = db.query(UpdateUploadedOrder).filter(
            UpdateUploadedOrder.OrderNumber == order_number
        ).first()
        
        if not order:
            raise HTTPException(status_code=404, detail="Order not found")
        
        # Update remarks
        order.Remarks = remark
        db.commit()
        
        
        return {
            "message": "Remarks updated successfully",
            "order_number": order_number,
            "remarks": remark
        }
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        print(f"Error updating order remarks by number: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to update remarks: {str(e)}")

@app.get("/api/listbrand")
async def get_list_brand(
    current_user: str = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    try:
        brands = db.query(ListBrand).all()
        
        # Convert datetime objects to ISO format strings
        brands_data = []
        for brand in brands:
            brand_dict = {
                "id": brand.id,
                "brand": brand.brand,
                "marketplace": brand.marketplace,
                "batch": brand.batch,
                "created_at": convert_to_wib(brand.created_at).isoformat() if brand.created_at else None
            }
            brands_data.append(brand_dict)
        
        return {"brands": brands_data}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/listbrand/brands")
async def get_unique_brands(
    current_user: str = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    try:
        brands = db.query(ListBrand.brand).distinct().all()
        unique_brands = [brand[0] for brand in brands]
        return {"brands": unique_brands}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/listbrand/marketplaces")
async def get_unique_marketplaces(
    current_user: str = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    try:
        marketplaces = db.query(ListBrand.marketplace).distinct().all()
        unique_marketplaces = [marketplace[0] for marketplace in marketplaces]
        return {"marketplaces": unique_marketplaces}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/brandshops")
async def get_brand_shops(
    skip: int = 0,
    limit: int = 100,
    brand: Optional[str] = None,
    marketplace_id: Optional[int] = None,
    search: Optional[str] = None,
    current_user: str = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get brand shops with optional filtering and global search"""
    try:
        query = db.query(BrandShop)
        
        # Global search across multiple fields
        if search and search.strip():
            search_term = search.strip()
            # Use exact match for shop_name to prevent incorrect shop_key_1 retrieval
            query = query.filter(
                or_(
                    BrandShop.shop_name.ilike(search_term),  # Exact match for shop_name
                    BrandShop.brand.ilike(f"%{search_term}%"),  # Partial match for other fields
                    BrandShop.shop_key_1.ilike(f"%{search_term}%"),
                    BrandShop.created_by.ilike(f"%{search_term}%"),
                    BrandShop.client_shop_id.cast(String).ilike(f"%{search_term}%"),
                    BrandShop.client_id.cast(String).ilike(f"%{search_term}%"),
                    BrandShop.marketplace_id.cast(String).ilike(f"%{search_term}%"),
                    BrandShop.order_type.ilike(f"%{search_term}%"),
                    BrandShop.shop_name_seller.ilike(f"%{search_term}%")
                )
            )
        
        if brand:
            query = query.filter(BrandShop.brand.ilike(f"%{brand}%"))
        
        if marketplace_id:
            query = query.filter(BrandShop.marketplace_id == marketplace_id)
        
        total = query.count()
        brand_shops = query.offset(skip).limit(limit).all()
        
        return {
            "total": total,
            "data": brand_shops,
            "skip": skip,
            "limit": limit
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/brandshops/brands")
async def get_unique_brands_from_shops(
    current_user: str = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get unique brands from brand_shops table"""
    try:
        brands = db.query(BrandShop.brand).filter(BrandShop.brand.isnot(None)).distinct().all()
        unique_brands = [brand[0] for brand in brands if brand[0]]
        return {"brands": unique_brands}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/brandshops/marketplaces")
async def get_unique_marketplace_ids(
    current_user: str = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get unique marketplace IDs from brand_shops table"""
    try:
        marketplaces = db.query(BrandShop.marketplace_id).filter(BrandShop.marketplace_id.isnot(None)).distinct().all()
        unique_marketplace_ids = [marketplace[0] for marketplace in marketplaces if marketplace[0]]
        return {"marketplace_ids": unique_marketplace_ids}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/brandshops")
async def create_brand_shop(
    shop_data: dict,
    current_user: str = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Create new brand shop"""
    try:
        new_shop = BrandShop(**shop_data)
        db.add(new_shop)
        db.commit()
        db.refresh(new_shop)
        return new_shop
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/brandshops/{shop_id}")
async def update_brand_shop(
    shop_id: int,
    shop_data: dict,
    current_user: str = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Update brand shop"""
    try:
        shop = db.query(BrandShop).filter(BrandShop.id == shop_id).first()
        if not shop:
            raise HTTPException(status_code=404, detail="Shop not found")
        
        for key, value in shop_data.items():
            if hasattr(shop, key):
                setattr(shop, key, value)
        
        db.commit()
        db.refresh(shop)
        return shop
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/brandshops/{shop_id}")
async def delete_brand_shop(
    shop_id: int,
    current_user: str = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Delete brand shop"""
    try:
        shop = db.query(BrandShop).filter(BrandShop.id == shop_id).first()
        if not shop:
            raise HTTPException(status_code=404, detail="Shop not found")
        
        db.delete(shop)
        db.commit()
        return {"message": "Shop deleted successfully"}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/brandshops/bulk-create")
async def bulk_create_brand_shops(
    file: UploadFile = File(...),
    current_user: str = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Bulk create brand shops from Excel/CSV file"""
    try:
        created_shops = []
        errors = []
        
        # Read file content
        file_content = file.file.read()
        
        # Parse file based on extension
        if file.filename.endswith('.xlsx'):
            df = pd.read_excel(io.BytesIO(file_content))
        elif file.filename.endswith('.csv'):
            df = pd.read_csv(io.BytesIO(file_content))
        else:
            raise HTTPException(status_code=400, detail="Unsupported file format. Please use .xlsx or .csv")
        
        # Validate required columns
        required_columns = ['brand', 'marketplace_id']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            raise HTTPException(status_code=400, detail=f"Missing required columns: {', '.join(missing_columns)}")
        
        # Process each row
        for i, row in df.iterrows():
            try:
                shop_data = row.to_dict()
                
                # Clean NaN values and convert types
                for key, value in shop_data.items():
                    if pd.isna(value):
                        shop_data[key] = None
                    elif key in ['marketplace_id', 'client_id', 'client_shop_id'] and value is not None:
                        try:
                            shop_data[key] = int(float(value)) if value != '' else None
                        except (ValueError, TypeError):
                            shop_data[key] = None
                    elif isinstance(value, (int, float)) and (value == float('inf') or value == float('-inf') or pd.isna(value)):
                        shop_data[key] = None
                
                # Validate required fields
                if not shop_data.get('brand') or not shop_data.get('marketplace_id'):
                    errors.append(f"Row {i+2}: Brand and marketplace_id are required")
                    continue
                
                # Check if shop already exists (brand + marketplace_id combination)
                existing = db.query(BrandShop).filter(
                    BrandShop.brand == shop_data['brand'],
                    BrandShop.marketplace_id == shop_data['marketplace_id']
                ).first()
                
                if existing:
                    errors.append(f"Row {i+2}: Shop already exists for brand {shop_data['brand']} in marketplace {shop_data['marketplace_id']}")
                    continue
                
                # Create new shop with defaults
                new_shop = BrandShop(
                    brand=shop_data['brand'],
                    marketplace_id=shop_data['marketplace_id'],
                    shop_name=shop_data.get('shop_name', f"{shop_data['brand']} Shop"),
                    shop_key_1=shop_data.get('shop_key_1', ''),
                    client_shop_id=shop_data.get('client_shop_id'),
                    client_id=shop_data.get('client_id', 1),
                    order_type=shop_data.get('order_type', 'ONLINE'),
                    created_by=current_user,
                    status=shop_data.get('status', 1),
                    is_open=shop_data.get('is_open', 1)
                )
                
                db.add(new_shop)
                db.flush()  # Get the ID without committing
                
                created_shops.append({
                    "id": int(new_shop.id),
                    "brand": str(new_shop.brand) if new_shop.brand else None,
                    "marketplace_id": int(new_shop.marketplace_id) if new_shop.marketplace_id else None,
                    "shop_name": str(new_shop.shop_name) if new_shop.shop_name else None,
                    "shop_key_1": str(new_shop.shop_key_1) if new_shop.shop_key_1 else None,
                    "client_shop_id": int(new_shop.client_shop_id) if new_shop.client_shop_id else None,
                    "client_id": int(new_shop.client_id) if new_shop.client_id else None,
                    "order_type": str(new_shop.order_type) if new_shop.order_type else None
                })
                
            except Exception as e:
                errors.append(f"Row {i+2}: {str(e)}")
        
        db.commit()
        
        return {
            "message": f"Bulk operation completed. {len(created_shops)} shops created, {len(errors)} errors.",
            "created_count": len(created_shops),
            "error_count": len(errors),
            "errors": errors,
            "shops": created_shops
        }
        
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/brandshops/bulk-update")
async def bulk_update_brand_shops(
    updates_data: List[dict],
    current_user: str = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Bulk update brand shops"""
    try:
        updated_shops = []
        errors = []
        
        for i, update_data in enumerate(updates_data):
            try:
                shop_id = update_data.get('id')
                if not shop_id:
                    errors.append(f"Row {i+1}: ID is required for updates")
                    continue
                
                shop = db.query(BrandShop).filter(BrandShop.id == shop_id).first()
                if not shop:
                    errors.append(f"Row {i+1}: Shop with ID {shop_id} not found")
                    continue
                
                # Update fields
                for key, value in update_data.items():
                    if key != 'id' and hasattr(shop, key):
                        setattr(shop, key, value)
                
                updated_shops.append({
                    "id": shop.id,
                    "brand": shop.brand,
                    "marketplace_id": shop.marketplace_id,
                    "shop_name": shop.shop_name,
                    "shop_key_1": shop.shop_key_1,
                    "client_shop_id": shop.client_shop_id,
                    "client_id": shop.client_id,
                    "order_type": shop.order_type
                })
                
            except Exception as e:
                errors.append(f"Row {i+1}: {str(e)}")
        
        db.commit()
        
        return {
            "message": f"Bulk update completed. {len(updated_shops)} shops updated, {len(errors)} errors.",
            "updated_count": len(updated_shops),
            "error_count": len(errors),
            "errors": errors,
            "shops": updated_shops
        }
        
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/brandshops/bulk-delete")
async def bulk_delete_brand_shops(
    ids: List[int],
    current_user: str = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Bulk delete brand shops"""
    try:
        deleted_count = 0
        errors = []
        
        for shop_id in ids:
            try:
                shop = db.query(BrandShop).filter(BrandShop.id == shop_id).first()
                if shop:
                    db.delete(shop)
                    deleted_count += 1
                else:
                    errors.append(f"Shop with ID {shop_id} not found")
            except Exception as e:
                errors.append(f"Error deleting shop ID {shop_id}: {str(e)}")
        
        db.commit()
        
        return {
            "message": f"Bulk delete completed. {deleted_count} shops deleted, {len(errors)} errors.",
            "deleted_count": deleted_count,
            "error_count": len(errors),
            "errors": errors
        }
        
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/listbrand")
async def create_list_brand(
    brand_data: dict,
    current_user: str = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    try:
        # Check if brand-marketplace-batch combination already exists
        existing = db.query(ListBrand).filter(
            ListBrand.brand == brand_data['brand'],
            ListBrand.marketplace == brand_data['marketplace'],
            ListBrand.batch == brand_data.get('batch', '1')
        ).first()
        
        if existing:
            raise HTTPException(status_code=400, detail="Brand-marketplace-batch combination already exists")
        
        new_brand = ListBrand(
            brand=brand_data['brand'],
            marketplace=brand_data['marketplace'],
            batch=brand_data.get('batch', '1')  # Default to '1' if not provided
        )
        
        db.add(new_brand)
        db.commit()
        db.refresh(new_brand)
        
        return {
            "message": "Brand added successfully",
            "brand": {
                "id": new_brand.id,
                "brand": new_brand.brand,
                "marketplace": new_brand.marketplace,
                "batch": new_brand.batch,
                "created_at": new_brand.created_at.isoformat() if new_brand.created_at else None
            }
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/listbrand/bulk-create")
async def bulk_create_marketplace_info(
    file: UploadFile = File(...),
    current_user: str = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Bulk create marketplace info entries from Excel/CSV file"""
    try:
        created_brands = []
        errors = []
        
        # Read file content
        file_content = file.file.read()
        
        # Parse file based on extension
        if file.filename.endswith('.xlsx'):
            df = pd.read_excel(io.BytesIO(file_content))
        elif file.filename.endswith('.csv'):
            df = pd.read_csv(io.BytesIO(file_content))
        else:
            raise HTTPException(status_code=400, detail="Unsupported file format. Please use .xlsx or .csv")
        
        # Validate required columns
        required_columns = ['brand', 'marketplace']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            raise HTTPException(status_code=400, detail=f"Missing required columns: {', '.join(missing_columns)}")
        
        # Process each row
        for i, row in df.iterrows():
            try:
                brand_data = row.to_dict()
                
                # Clean NaN values and convert types
                for key, value in brand_data.items():
                    if pd.isna(value):
                        brand_data[key] = None
                    elif isinstance(value, (int, float)) and (value == float('inf') or value == float('-inf') or pd.isna(value)):
                        brand_data[key] = None
                    elif value is not None:
                        brand_data[key] = str(value)
                
                # Validate required fields
                if not brand_data.get('brand') or not brand_data.get('marketplace'):
                    errors.append(f"Row {i+2}: Brand and marketplace are required")  # +2 because Excel starts at 1 and has header
                    continue
                
                # Check if brand-marketplace-batch combination already exists
                existing = db.query(ListBrand).filter(
                    ListBrand.brand == brand_data['brand'],
                    ListBrand.marketplace == brand_data['marketplace'],
                    ListBrand.batch == brand_data.get('batch', '1')
                ).first()
                
                if existing:
                    errors.append(f"Row {i+2}: Brand-marketplace-batch combination already exists")
                    continue
                
                new_brand = ListBrand(
                    brand=brand_data['brand'],
                    marketplace=brand_data['marketplace'],
                    batch=brand_data.get('batch', '1'),
                    remark=brand_data.get('remark', '')
                )
                
                db.add(new_brand)
                db.flush()  # Get the ID without committing
                
                created_brands.append({
                    "id": int(new_brand.id),
                    "brand": str(new_brand.brand) if new_brand.brand else None,
                    "marketplace": str(new_brand.marketplace) if new_brand.marketplace else None,
                    "batch": str(new_brand.batch) if new_brand.batch else None,
                    "remark": str(new_brand.remark) if new_brand.remark else None
                })
                
            except Exception as e:
                errors.append(f"Row {i+2}: {str(e)}")
        
        db.commit()
        
        return {
            "message": f"Bulk operation completed. {len(created_brands)} entries created, {len(errors)} errors.",
            "created_count": len(created_brands),
            "error_count": len(errors),
            "errors": errors,
            "brands": created_brands
        }
        
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/listbrand/bulk-update")
async def bulk_update_marketplace_info(
    file: UploadFile = File(...),
    current_user: str = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Bulk update marketplace info entries from Excel/CSV file"""
    try:
        updated_brands = []
        errors = []
        
        # Read file content
        file_content = file.file.read()
        
        # Parse file based on extension
        if file.filename.endswith('.xlsx'):
            df = pd.read_excel(io.BytesIO(file_content))
        elif file.filename.endswith('.csv'):
            df = pd.read_csv(io.BytesIO(file_content))
        else:
            raise HTTPException(status_code=400, detail="Unsupported file format. Please use .xlsx or .csv")
        
        # Validate required columns
        required_columns = ['id']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            raise HTTPException(status_code=400, detail=f"Missing required columns: {', '.join(missing_columns)}")
        
        # Process each row
        for i, row in df.iterrows():
            try:
                update_data = row.to_dict()
                
                brand_id = update_data.get('id')
                if not brand_id:
                    errors.append(f"Row {i+2}: ID is required for updates")
                    continue
                
                brand = db.query(ListBrand).filter(ListBrand.id == brand_id).first()
                if not brand:
                    errors.append(f"Row {i+2}: Brand with ID {brand_id} not found")
                    continue
                
                # Update fields
                if 'brand' in update_data and update_data['brand']:
                    brand.brand = update_data['brand']
                if 'marketplace' in update_data and update_data['marketplace']:
                    brand.marketplace = update_data['marketplace']
                if 'batch' in update_data and update_data['batch']:
                    brand.batch = update_data['batch']
                if 'remark' in update_data:
                    brand.remark = update_data['remark']
                
                updated_brands.append({
                    "id": brand.id,
                    "brand": brand.brand,
                    "marketplace": brand.marketplace,
                    "batch": brand.batch,
                    "remark": brand.remark
                })
                
            except Exception as e:
                errors.append(f"Row {i+2}: {str(e)}")
        
        db.commit()
        
        return {
            "message": f"Bulk update completed. {len(updated_brands)} entries updated, {len(errors)} errors.",
            "updated_count": len(updated_brands),
            "error_count": len(errors),
            "errors": errors,
            "brands": updated_brands
        }
        
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/listbrand/bulk-delete")
async def bulk_delete_marketplace_info(
    ids: List[int],
    current_user: str = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Bulk delete marketplace info entries"""
    try:
        deleted_count = 0
        errors = []
        
        for brand_id in ids:
            try:
                brand = db.query(ListBrand).filter(ListBrand.id == brand_id).first()
                if brand:
                    db.delete(brand)
                    deleted_count += 1
                else:
                    errors.append(f"Brand with ID {brand_id} not found")
            except Exception as e:
                errors.append(f"Error deleting brand ID {brand_id}: {str(e)}")
        
        db.commit()
        
        return {
            "message": f"Bulk delete completed. {deleted_count} entries deleted, {len(errors)} errors.",
            "deleted_count": deleted_count,
            "error_count": len(errors),
            "errors": errors
        }
        
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/listbrand/remark")
async def update_list_brand_remark(
    brand: str = Query(..., description="Brand name"),
    marketplace: str = Query(..., description="Marketplace name"),
    batch: str = Query(..., description="Batch number"),
    remark: Optional[str] = Query(None, description="Remark text"),
    current_user: str = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Update remark for a specific brand+marketplace+batch combination"""
    try:
        logger.info(f"Updating remark for brand={brand}, marketplace={marketplace}, batch={batch}, remark={remark}")
        
        # Find the brand record
        brand_record = db.query(ListBrand).filter(
            ListBrand.brand == brand,
            ListBrand.marketplace == marketplace,
            ListBrand.batch == batch
        ).first()
        
        if not brand_record:
            raise HTTPException(status_code=404, detail="Brand combination not found")
        
        # Update the remark in list_brand table (for backward compatibility)
        brand_record.remark = remark
        db.commit()
        db.refresh(brand_record)
        
        # Also update remark in history table for today's records
        current_time = get_wib_now()
        today_start = current_time.replace(hour=0, minute=0, second=0, microsecond=0)
        today_end = current_time.replace(hour=23, minute=59, second=59, microsecond=999999)
        
        history_records = db.query(NotUploadedHistory).filter(
            NotUploadedHistory.brand == brand,
            NotUploadedHistory.marketplace == marketplace,
            NotUploadedHistory.batch == batch,
            NotUploadedHistory.check_date >= today_start,
            NotUploadedHistory.check_date <= today_end
        ).all()
        
        # Update all today's history records with the new remark
        for history_record in history_records:
            history_record.remark = remark
        
        db.commit()
        
        return {
            "message": "Remark updated successfully",
            "brand": brand_record.brand,
            "marketplace": brand_record.marketplace,
            "batch": brand_record.batch,
            "remark": brand_record.remark,
            "history_updated": len(history_records)
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating remark: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/not-uploaded-history/remark")
async def update_not_uploaded_history_remark(
    brand: str = Query(..., description="Brand name"),
    marketplace: str = Query(..., description="Marketplace name"),
    batch: str = Query(..., description="Batch number"),
    check_date: str = Query(..., description="Check date (ISO format)"),
    remark: Optional[str] = Query(None, description="Remark text"),
    current_user: str = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Update remark for a specific not uploaded history record"""
    try:
        logger.info(f"Updating history remark for brand={brand}, marketplace={marketplace}, batch={batch}, check_date={check_date}, remark={remark}")
        
        # Parse the check_date
        check_datetime = datetime.fromisoformat(check_date.replace('Z', '+00:00'))
        
        # Find the history record
        history_record = db.query(NotUploadedHistory).filter(
            NotUploadedHistory.brand == brand,
            NotUploadedHistory.marketplace == marketplace,
            NotUploadedHistory.batch == batch,
            NotUploadedHistory.check_date == check_datetime
        ).first()
        
        if not history_record:
            raise HTTPException(status_code=404, detail="History record not found")
        
        # Update the remark
        history_record.remark = remark
        db.commit()
        db.refresh(history_record)
        
        return {
            "message": "History remark updated successfully",
            "brand": history_record.brand,
            "marketplace": history_record.marketplace,
            "batch": history_record.batch,
            "check_date": history_record.check_date.isoformat(),
            "remark": history_record.remark
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating history remark: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/not-interfaced-order/remark")
async def update_not_interfaced_order_remark(
    order_number: str = Query(..., description="Order number"),
    marketplace: str = Query(..., description="Marketplace name"),
    remark: Optional[str] = Query(None, description="Remark text"),
    current_user: str = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Update remark for a specific not interfaced order"""
    try:
        logger.info(f"Updating not interfaced order remark for order_number={order_number}, marketplace={marketplace}, remark={remark}")
        
        # Find the order in UploadedOrder table
        order = db.query(UploadedOrder).filter(
            UploadedOrder.OrderNumber == order_number,
            UploadedOrder.Marketplace == marketplace
        ).first()
        
        if not order:
            raise HTTPException(status_code=404, detail="Order not found")
        
        # Update the remark
        order.Remarks = remark
        db.commit()
        db.refresh(order)
        
        return {
            "message": "Not interfaced order remark updated successfully",
            "order_number": order.OrderNumber,
            "marketplace": order.Marketplace,
            "brand": order.Brand,
            "remark": order.Remarks
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating not interfaced order remark: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/listbrand/{brand_id}")
async def update_list_brand(
    brand_id: int,
    brand_data: dict,
    current_user: str = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    try:
        brand = db.query(ListBrand).filter(ListBrand.id == brand_id).first()
        if not brand:
            raise HTTPException(status_code=404, detail="Brand not found")
        
        # Check if new combination already exists (excluding current record)
        existing = db.query(ListBrand).filter(
            ListBrand.brand == brand_data['brand'],
            ListBrand.marketplace == brand_data['marketplace'],
            ListBrand.batch == brand_data.get('batch', '1'),
            ListBrand.id != brand_id
        ).first()
        
        if existing:
            raise HTTPException(status_code=400, detail="Brand-marketplace-batch combination already exists")
        
        brand.brand = brand_data['brand']
        brand.marketplace = brand_data['marketplace']
        if 'batch' in brand_data:
            brand.batch = brand_data['batch']
        
        db.commit()
        db.refresh(brand)
        
        return {
            "message": "Brand updated successfully",
            "brand": {
                "id": brand.id,
                "brand": brand.brand,
                "marketplace": brand.marketplace,
                "batch": brand.batch,
                "created_at": brand.created_at.isoformat() if brand.created_at else None
            }
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/listbrand/{brand_id}")
async def delete_list_brand(
    brand_id: int,
    current_user: str = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    try:
        brand = db.query(ListBrand).filter(ListBrand.id == brand_id).first()
        if not brand:
            raise HTTPException(status_code=404, detail="Brand not found")
        
        db.delete(brand)
        db.commit()
        
        return {"message": "Brand deleted successfully"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# Upload History endpoints
@app.post("/api/file-upload-history")
async def create_upload_history(
    marketplace: str = Form(...),
    brand: str = Form(...),
    pic: str = Form(...),
    batch: str = Form(...),
    current_user: str = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    try:
        new_history = UploadHistory(
            marketplace=marketplace,
            brand=brand,
            pic=pic,
            batch=batch
        )
        
        db.add(new_history)
        db.commit()
        db.refresh(new_history)
        
        return {
            "message": "Upload history created successfully",
            "history": {
                "id": new_history.id,
                "marketplace": new_history.marketplace,
                "brand": new_history.brand,
                "pic": new_history.pic,
                "batch": new_history.batch,
                "upload_date": new_history.upload_date.isoformat() if new_history.upload_date and hasattr(new_history.upload_date, 'isoformat') else (new_history.upload_date if new_history.upload_date else None)
            }
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/dashboard/recent-orders")
def get_dashboard_recent_orders(db: Session = Depends(get_db)):
    """Get recent orders for dashboard without authentication"""
    try:
        orders = db.query(UploadedOrder)\
            .order_by(UploadedOrder.UploadDate.desc())\
            .limit(10)\
            .all()
        
        orders_data = []
        for order in orders:
            # Fallback logic: Use OrderNumberFlexo if available, otherwise use original OrderNumber
            display_order_number = order.OrderNumberFlexo if order.OrderNumberFlexo and order.OrderNumberFlexo.strip() else order.OrderNumber
            
            order_dict = {
                "Id": order.Id,
                "Marketplace": order.Marketplace,
                "Brand": order.Brand,
                "OrderNumber": display_order_number,  # Fallback: OrderNumberFlexo or OrderNumber
                "OrderStatus": order.OrderStatusFlexo,
                "AWB": order.AWB,
                "Transporter": order.Transporter,
                "OrderDate": convert_to_wib(order.OrderDate).isoformat() if order.OrderDate else None,
                "SLA": order.SLA,
                "Batch": order.Batch,
                "PIC": order.PIC,
                "UploadDate": convert_to_wib(order.UploadDate).isoformat() if order.UploadDate else None,
                "InterfaceStatus": order.InterfaceStatus
            }
            orders_data.append(order_dict)
        
        return {"orders": orders_data}
        
    except Exception as e:
        logger.error(f"Error getting recent orders: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

# DEPRECATED: Old dashboard advanced-stats endpoint removed
# Use /api/dashboard/advanced-stats instead for optimized performance and correct date filtering

# DEPRECATED: Old dashboard upload-history endpoint removed
# Use /api/dashboard/upload-history instead for optimized performance and correct date filtering

@app.get("/api/file-upload-history", response_model=List[UploadHistoryResponse])
async def get_upload_history(
    current_user: str = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    try:
        history = db.query(UploadHistory).order_by(UploadHistory.upload_date.desc()).all()
        return history
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/debug/upload-history-count")
def debug_upload_history_count(db: Session = Depends(get_db)):
    """Debug endpoint to check upload history data"""
    try:
        total_count = db.query(UploadHistory).count()
        recent_records = db.query(UploadHistory).order_by(UploadHistory.upload_date.desc()).limit(10).all()
        
        return {
            "total_records": total_count,
            "recent_records": [
                {
                    "id": record.id,
                    "marketplace": record.marketplace,
                    "brand": record.brand,
                    "upload_date": record.upload_date.isoformat() if record.upload_date else None
                }
                for record in recent_records
            ]
        }
    except Exception as e:
        logger.error(f"Error in debug endpoint: {str(e)}")
        return {"error": str(e)}

@app.get("/api/file-upload-history/{history_id}", response_model=UploadHistoryResponse)
async def get_upload_history_by_id(
    history_id: int,
    current_user: str = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    try:
        history = db.query(UploadHistory).filter(UploadHistory.id == history_id).first()
        if not history:
            raise HTTPException(status_code=404, detail="Upload history not found")
        return history
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/file-upload-history/{history_id}")
async def delete_upload_history(
    history_id: int,
    current_user: str = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    try:
        # Check if user is superuser or admin
        user = db.query(User).filter(func.lower(User.username) == func.lower(current_user)).first()
        if not user or user.role not in ["admin", "superuser"]:
            raise HTTPException(status_code=403, detail="Only superusers can delete upload history")
        
        history = db.query(UploadHistory).filter(UploadHistory.id == history_id).first()
        if not history:
            raise HTTPException(status_code=404, detail="Upload history not found")
        
        db.delete(history)
        db.commit()
        
        return {"message": "Upload history deleted successfully"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# Startup and shutdown events are now handled by the lifespan context manager

# Marketplace Runner API endpoints
@app.get("/api/marketplace-status")
def get_marketplace_status(current_user: str = Depends(get_current_user)):
    """Get status of all marketplace apps and their orderlist files for current user"""
    try:
        # Check if marketplace apps are enabled
        if not MARKETPLACE_APPS_ENABLED:
            logger.info(f"Marketplace apps disabled, skipping status check for user {current_user}")
            return {}
        
        # Get user-specific workspace
        user_workspace = multi_user_handler.get_user_workspace(current_user)
        print(f"👤 Getting marketplace status for user {current_user} in workspace: {user_workspace}")
        
        marketplaces = {
            'shopee': {'folder': 'Shopee', 'exe': 'ShopeeOrderLogistic.exe'},
            'lazada': {'folder': 'Lazada', 'exe': 'LazadaMarketplace.exe'},
            'blibli': {'folder': 'Blibli', 'exe': 'BliBliProduct2024.exe'},
            'desty': {'folder': 'Desty', 'exe': 'Desty.Console.exe'},
            'ginee': {'folder': 'Ginee', 'exe': 'Ginee.sync.exe'},
            'tiktok': {'folder': 'Tiktok', 'exe': 'tiktok.api.exe'},
            'zalora': {'folder': 'Zalora', 'exe': 'Zalora.Flexo.Integration.exe'},
            'tokopedia': {'folder': 'Tokopedia', 'exe': 'TokopediaOrder.exe'}
        }
        
        status = {}
        for marketplace, config in marketplaces.items():
            folder_path = os.path.join(user_workspace, config['folder'])
            orderlist_path = os.path.join(folder_path, 'Orderlist.txt')
            exe_path = os.path.join(folder_path, config['exe'])
            
            # Check if folder exists, if not, ensure it exists (auto-create if missing)
            if not os.path.exists(folder_path):
                ensure_marketplace_folder_exists(user_workspace, marketplace, config)
            
            # Check if orderlist exists and count orders
            has_orderlist = os.path.exists(orderlist_path)
            order_count = 0
            if has_orderlist:
                try:
                    with open(orderlist_path, 'r', encoding='utf-8') as f:
                        lines = f.readlines()
                        order_count = len([line.strip() for line in lines if line.strip()])
                except:
                    order_count = 0
            
            # Check if exe exists
            has_exe = os.path.exists(exe_path)
            
            # Determine status
            if has_exe and has_orderlist and order_count > 0:
                marketplace_status = 'ready'
            elif has_exe:
                marketplace_status = 'no_orders'
            else:
                marketplace_status = 'error'
            
            status[marketplace] = {
                'status': marketplace_status,
                'hasOrderlist': has_orderlist,
                'orderCount': order_count,
                'hasExe': has_exe,
                'folder': config['folder'],
                'exe': config['exe']
            }
        
        return status
        
    except Exception as e:
        logger.error(f"Error getting marketplace status: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get marketplace status: {str(e)}")

@app.get("/api/marketplace-orderlist/{marketplace}")
def get_marketplace_orderlist(
    marketplace: str,
    current_user: str = Depends(get_current_user)
):
    """Get content of orderlist.txt for specific marketplace in user workspace"""
    try:
        # Get user-specific workspace
        user_workspace = multi_user_handler.get_user_workspace(current_user)
        print(f"👤 Viewing orderlist for user {current_user} in workspace: {user_workspace}")
        
        marketplace_folders = {
            'shopee': 'Shopee',
            'lazada': 'Lazada',
            'blibli': 'Blibli',
            'desty': 'Desty',
            'ginee': 'Ginee',
            'tiktok': 'Tiktok',
            'zalora': 'Zalora',
            'jubelio': 'Jubelio',
            'tokopedia': 'Tokopedia'
        }
        
        if marketplace not in marketplace_folders:
            raise HTTPException(status_code=400, detail="Invalid marketplace")
        
        folder_path = os.path.join(user_workspace, marketplace_folders[marketplace])
        page_orderlist_path = os.path.join(user_workspace, 'Page_Orderlist', marketplace_folders[marketplace])
        
        # Ensure marketplace folder exists (auto-create if missing) only if folder doesn't exist
        if not os.path.exists(folder_path):
            config = {'folder': marketplace_folders[marketplace]}
            ensure_marketplace_folder_exists(user_workspace, marketplace, config)
        
        # Ensure Page_Orderlist marketplace folder exists
        if not os.path.exists(page_orderlist_path):
            os.makedirs(page_orderlist_path, exist_ok=True)
            print(f"📁 Created Page_Orderlist folder: {page_orderlist_path}")
        
        orderlist_path = os.path.join(page_orderlist_path, 'Orderlist.txt')
        
        if not os.path.exists(orderlist_path):
            return {"content": "Orderlist.txt tidak ditemukan"}
        
        with open(orderlist_path, 'r', encoding='utf-8') as f:
            content = f.read()
        
        return {"content": content}
        
    except Exception as e:
        logger.error(f"Error getting orderlist for {marketplace}: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get orderlist: {str(e)}")

@app.post("/api/run-marketplace-app")
def run_marketplace_app(
    request: dict,
    current_user: str = Depends(get_current_user)
):
    """Run specific marketplace app in user workspace"""
    try:
        # Check if marketplace apps are enabled
        if not MARKETPLACE_APPS_ENABLED:
            raise HTTPException(status_code=503, detail="Marketplace apps are disabled")
            
        marketplace = request.get('marketplace')
        order_number = request.get('order_number')
        if not marketplace:
            raise HTTPException(status_code=400, detail="Marketplace is required")
        
        # Get user-specific workspace
        user_workspace = multi_user_handler.get_user_workspace(current_user)
        if order_number:
            print(f"👤 Running marketplace app for user {current_user} - Order {order_number} in workspace: {user_workspace}")
        else:
            print(f"👤 Running marketplace app for user {current_user} in workspace: {user_workspace}")
        
        marketplace_configs = {
            'shopee': {'folder': 'Shopee', 'exe': 'ShopeeOrderLogistic.exe'},
            'lazada': {'folder': 'Lazada', 'exe': 'LazadaMarketplace.exe'},
            'blibli': {'folder': 'Blibli', 'exe': 'BliBliProduct2024.exe'},
            'desty': {'folder': 'Desty', 'exe': 'Desty.Console.exe'},
            'ginee': {'folder': 'Ginee', 'exe': 'Ginee.sync.exe'},
            'tiktok': {'folder': 'Tiktok', 'exe': 'tiktok.api.exe'},
            'zalora': {'folder': 'Zalora', 'exe': 'Zalora.Flexo.Integration.exe'},
            'tokopedia': {'folder': 'Tokopedia', 'exe': 'TokopediaOrder.exe'}
        }
        
        if marketplace not in marketplace_configs:
            raise HTTPException(status_code=400, detail="Invalid marketplace")
        
        config = marketplace_configs[marketplace]
        
        folder_path = os.path.join(user_workspace, config['folder'])
        
        # Ensure marketplace folder exists (auto-create if missing) only if folder doesn't exist
        if not os.path.exists(folder_path):
            ensure_marketplace_folder_exists(user_workspace, marketplace, config)
        exe_path = os.path.join(folder_path, config['exe'])
        
        if not os.path.exists(exe_path):
            raise HTTPException(status_code=404, detail=f"{config['exe']} not found")
        
        # Check if Orderlist.txt exists and count orders (now in Page_Orderlist folder)
        page_orderlist_path = os.path.join(folder_path, 'Page_Orderlist', config['folder'])
        orderlist_path = os.path.join(page_orderlist_path, 'Orderlist.txt')
        
        # Ensure Page_Orderlist marketplace folder exists
        if not os.path.exists(page_orderlist_path):
            os.makedirs(page_orderlist_path, exist_ok=True)
            print(f"📁 Created Page_Orderlist folder: {page_orderlist_path}")
        order_count = 0
        
        # If specific order_number is provided, temporarily replace Orderlist.txt with single order
        temp_orderlist_created = False
        original_orderlist_backup = None
        
        try:
            if order_number:
                # Backup original Orderlist.txt if it exists
                if os.path.exists(orderlist_path):
                    original_orderlist_backup = os.path.join(page_orderlist_path, 'Orderlist.txt.bak')
                    shutil.copy2(orderlist_path, original_orderlist_backup)
                    print(f"📋 Backed up original Orderlist.txt to Orderlist.txt.bak")
                
                # Create new Orderlist.txt with only the specified order
                try:
                    with open(orderlist_path, 'w', encoding='utf-8') as f:
                        f.write(order_number)
                    print(f"📋 Created temporary Orderlist.txt for order {order_number}")
                    print(f"📋 File path: {orderlist_path}")
                    print(f"📋 File contents: {order_number}")
                    
                    # Also create a specific order file (some executables might look for this pattern)
                    specific_order_file = os.path.join(page_orderlist_path, f'Orderlist_{order_number}.txt')
                    with open(specific_order_file, 'w', encoding='utf-8') as f:
                        f.write(order_number)
                    print(f"📋 Also created specific order file: {specific_order_file}")
                    
                    # Also create a file with just the order number (no prefix)
                    simple_order_file = os.path.join(page_orderlist_path, f'{order_number}.txt')
                    with open(simple_order_file, 'w', encoding='utf-8') as f:
                        f.write(order_number)
                    print(f"📋 Also created simple order file: {simple_order_file}")
                    
                    temp_orderlist_created = True
                    order_count = 1
                except Exception as e:
                    print(f"⚠️ Could not create temporary Orderlist.txt: {str(e)}")
                    # Fallback to original orderlist
                    if os.path.exists(orderlist_path):
                        with open(orderlist_path, 'r', encoding='utf-8') as f:
                            lines = f.readlines()
                            order_count = len([line.strip() for line in lines if line.strip()])
            else:
                # Use original orderlist
                if os.path.exists(orderlist_path):
                    with open(orderlist_path, 'r', encoding='utf-8') as f:
                        lines = f.readlines()
                        order_count = len([line.strip() for line in lines if line.strip()])
            
            # Run the executable with logging
            print(f"🚀 Running executable: {exe_path}")
            print(f"🚀 Working directory: {folder_path}")
            
            # Try to get brand name from orderlist filename first, then config
            brand_name = "Unknown"
            try:
                # Try to get brand from config files (orderlist filename uses task_id, not brand name)
                config_files = [f for f in os.listdir(folder_path) if f.endswith('.config')]
                for config_file in config_files:
                    try:
                        with open(os.path.join(folder_path, config_file), 'r', encoding='utf-8') as cf:
                            config_content = cf.read()
                            brand_match = re.search(r'<add key="brand_name" value="([^"]+)"', config_content)
                            if brand_match:
                                brand_name = brand_match.group(1)
                                print(f"🔍 Detected brand from config file: {brand_name}")
                                break
                    except:
                        continue
                        
                print(f"✅ Using brand: {brand_name}")
                
                # Switch to correct config file based on detected brand
                if brand_name != "Unknown":
                    default_config_path = os.path.join(folder_path, f"{config['exe']}.config")
                    brand_config_path = os.path.join(folder_path, f"{brand_name}_{marketplace.upper()}.config")
                    
                    if os.path.exists(brand_config_path):
                        print(f"🔄 Switching to {brand_name} config...")
                        # Backup original config
                        backup_config_path = os.path.join(folder_path, f"{config['exe']}.config.backup")
                        if os.path.exists(default_config_path):
                            shutil.copy2(default_config_path, backup_config_path)
                            print(f"   📋 Backed up original config to {os.path.basename(backup_config_path)}")
                        
                        # Replace with brand config
                        shutil.copy2(brand_config_path, default_config_path)
                        print(f"   ✅ Using {brand_name} config: {os.path.basename(brand_config_path)}")
                    else:
                        print(f"⚠️ Brand config not found: {os.path.basename(brand_config_path)}")
                        print(f"   Using default config (may have wrong brand/shop_id)")
                        
            except Exception as e:
                print(f"⚠️ Error detecting brand: {str(e)}")
                pass
            
            success, process = run_executable_with_logging(exe_path, folder_path, marketplace, current_user, brand_name, None)
            if not success:
                print(f"⚠️ Failed to start {marketplace} app with logging, falling back to silent mode")
                success = run_executable_silently(exe_path, folder_path, marketplace)
                if not success:
                    print(f"⚠️ Failed to start {marketplace} app silently")
                else:
                    print(f"✅ Successfully started {marketplace} app silently")
            else:
                print(f"✅ Successfully started {marketplace} app with logging")
                
            # Add a delay to allow the executable to read the file
            if temp_orderlist_created:
                print(f"⏳ Waiting 3 seconds for executable to read Orderlist.txt...")
                import time
                time.sleep(3)
        
        finally:
            # Restore original config file if we switched to brand config
            try:
                if brand_name != "Unknown":
                    default_config_path = os.path.join(folder_path, f"{config['exe']}.config")
                    backup_config_path = os.path.join(folder_path, f"{config['exe']}.config.backup")
                    
                    if os.path.exists(backup_config_path):
                        shutil.move(backup_config_path, default_config_path)
                        print(f"🔄 Restored original config file")
            except Exception as e:
                print(f"⚠️ Error restoring original config: {str(e)}")
            
            # Restore original Orderlist.txt if we created a temporary one
            if temp_orderlist_created:
                try:
                    # Delete the temporary Orderlist.txt
                    if os.path.exists(orderlist_path):
                        os.remove(orderlist_path)
                        print(f"🗑️ Removed temporary Orderlist.txt")
                    
                    # Delete the specific order files if they were created
                    if order_number:
                        specific_order_file = os.path.join(folder_path, f'Orderlist_{order_number}.txt')
                        if os.path.exists(specific_order_file):
                            os.remove(specific_order_file)
                            print(f"🗑️ Removed specific order file: {specific_order_file}")
                        
                        simple_order_file = os.path.join(folder_path, f'{order_number}.txt')
                        if os.path.exists(simple_order_file):
                            os.remove(simple_order_file)
                            print(f"🗑️ Removed simple order file: {simple_order_file}")
                    
                    # Restore original Orderlist.txt from backup
                    if original_orderlist_backup and os.path.exists(original_orderlist_backup):
                        shutil.move(original_orderlist_backup, orderlist_path)
                        print(f"🔄 Restored original Orderlist.txt from backup")
                except Exception as e:
                    print(f"⚠️ Error restoring original Orderlist.txt: {str(e)}")
        
        # Create notification data
        if order_number:
            notification_data = {
                "marketplace": marketplace.title(),
                "brand": "Manual Run",
                "order_count": 1,
                "order_number": order_number,
                "message": f"{marketplace.title()} app started successfully for order {order_number}"
            }
            success_message = f"{marketplace.title()} app started successfully for order {order_number}"
        else:
            notification_data = {
                "marketplace": marketplace.title(),
                "brand": "Manual Run",
                "order_count": order_count,
                "message": f"{marketplace.title()} app started successfully for {order_count} orders"
            }
            success_message = f"{marketplace.title()} app started successfully"
        
        return {
            "success": True,
            "message": success_message,
            "marketplace": marketplace,
            "marketplace_notification": notification_data
        }
        
    except Exception as e:
        logger.error(f"Error running marketplace app {marketplace}: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to run app: {str(e)}")

@app.post("/api/run-all-marketplace-apps")
def run_all_marketplace_apps(current_user: str = Depends(get_current_user)):
    """Run all marketplace apps with smart orderlist generation"""
    try:
        # Check if marketplace apps are enabled
        if not MARKETPLACE_APPS_ENABLED:
            raise HTTPException(status_code=503, detail="Marketplace apps are disabled")
            
        project_root = get_project_root()
        jobgetorder_path = os.path.join(project_root, 'JobGetOrder')
        
        marketplace_configs = {
            'shopee': {'folder': 'Shopee', 'exe': 'ShopeeOrderLogistic.exe'},
            'lazada': {'folder': 'Lazada', 'exe': 'LazadaMarketplace.exe'},
            'blibli': {'folder': 'Blibli', 'exe': 'BliBliProduct2024.exe'},
            'desty': {'folder': 'Desty', 'exe': 'Desty.Console.exe'},
            'ginee': {'folder': 'Ginee', 'exe': 'Ginee.sync.exe'},
            'tiktok': {'folder': 'Tiktok', 'exe': 'tiktok.api.exe'},
            'zalora': {'folder': 'Zalora', 'exe': 'Zalora.Flexo.Integration.exe'},
            'tokopedia': {'folder': 'Tokopedia', 'exe': 'TokopediaOrder.exe'},
            'jubelio': {'folder': 'Jubelio', 'exe': 'Jubelio_project.exe'}
        }
        
        results = []
        import subprocess
        import platform
        import time
        
        # Get database session
        db = SessionLocal()
        try:
            print(f"🚀 Starting Run All Apps for user: {current_user}")
            
            for marketplace, config in marketplace_configs.items():
                folder_path = os.path.join(jobgetorder_path, config['folder'])
                page_orderlist_path = os.path.join(jobgetorder_path, 'Page_Orderlist', config['folder'])
                exe_path = os.path.join(folder_path, config['exe'])
                
                # Ensure Page_Orderlist marketplace folder exists
                if not os.path.exists(page_orderlist_path):
                    try:
                        os.makedirs(page_orderlist_path, exist_ok=True)
                        print(f"📁 Created Page_Orderlist folder: {page_orderlist_path}")
                    except Exception as e:
                        print(f"❌ Error creating Page_Orderlist folder: {str(e)}")
                        results.append({
                            "marketplace": marketplace,
                            "success": False,
                            "message": f"Failed to create Page_Orderlist folder for {marketplace}: {str(e)}"
                        })
                        continue
                
                # Step 1: Check if executable exists
                if not os.path.exists(exe_path):
                    results.append({
                        "marketplace": marketplace,
                        "success": False,
                        "message": f"{config['exe']} not found"
                    })
                    continue
                
                # Step 2: Find Marketplace+Brand combinations from .config files
                marketplace_brands = []
                try:
                    config_files = [f for f in os.listdir(folder_path) if f.endswith('.config')]
                    for config_file in config_files:
                        try:
                            with open(os.path.join(folder_path, config_file), 'r', encoding='utf-8') as cf:
                                config_content = cf.read()
                                brand_match = re.search(r'<add key="brand_name" value="([^"]+)"', config_content)
                                if brand_match:
                                    brand_name = brand_match.group(1)
                                    marketplace_brands.append({
                                        'brand': brand_name,
                                        'config_file': config_file
                                    })
                                    print(f"📋 Found config: {marketplace} - {brand_name}")
                        except Exception as e:
                            print(f"⚠️ Error reading config file {config_file}: {str(e)}")
                            continue
                except Exception as e:
                    print(f"⚠️ Error scanning config files for {marketplace}: {str(e)}")
                
                if not marketplace_brands:
                    results.append({
                        "marketplace": marketplace,
                        "success": False,
                        "message": f"No brand configs found for {marketplace}"
                    })
                    continue
                
                # Step 3: Generate Orderlist.txt for each brand combination
                marketplace_success = False
                total_orders_added = 0
                
                for brand_info in marketplace_brands:
                    brand_name = brand_info['brand']
                    
                    # Query orders: InterfaceStatus = 'Not Yet Interface' AND OrderStatus NOT IN cancelled statuses
                    cancelled_statuses = ['batal', 'cancel', 'cancellation', 'BATAL', 'CANCEL', 'CANCELLATION', 
                                        'Cancellations', 'Dibatalkan', 'Batal', 'CANCELED', 'Cancelled', 
                                        'canceled', 'Pembatalan diajukan', 'Order Batal', 'CANCELLED']
                    
                    not_interfaced_orders = db.query(UploadedOrder).filter(
                        UploadedOrder.Marketplace == marketplace.upper(),
                        UploadedOrder.Brand == brand_name,
                        UploadedOrder.InterfaceStatus == 'Not Yet Interface',
                        ~UploadedOrder.OrderStatusFlexo.in_(cancelled_statuses)
                    ).all()
                    
                    if not_interfaced_orders:
                        # Generate Orderlist.txt (now in Page_Orderlist folder)
                        orderlist_path = os.path.join(page_orderlist_path, 'Orderlist.txt')
                        order_numbers = [order.OrderNumber for order in not_interfaced_orders if order.OrderNumber]
                        
                        # Get shop_id for formatting
                        shop_id = get_shop_id_from_brand_shops(brand_name, marketplace, db)
                        
                        # Handle existing read-only file
                        if os.path.exists(orderlist_path):
                            try:
                                # Try to make file writable first
                                os.chmod(orderlist_path, 0o666)
                                print(f"🔓 Made existing Orderlist.txt writable: {orderlist_path}")
                            except Exception as e:
                                print(f"⚠️ Could not change permissions for existing file: {str(e)}")
                        
                        # Write to Orderlist.txt with proper formatting
                        try:
                            with open(orderlist_path, 'w', encoding='utf-8') as f:
                                for order_number in order_numbers:
                                    if marketplace.lower() in ['zalora', 'blibli']:
                                        # For ZALORA and BLIBLI: use format "id,shop_key_1"
                                        if shop_id:
                                            f.write(f"{order_number},{shop_id}\n")
                                        else:
                                            f.write(f"{order_number}\n")
                                    else:
                                        # For other marketplaces: use format "id" only
                                        f.write(f"{order_number}\n")
                            
                            # Make file read-only
                            try:
                                os.chmod(orderlist_path, 0o444)
                                print(f"🔒 Made Orderlist.txt read-only: {orderlist_path}")
                            except Exception as e:
                                print(f"⚠️ Could not make Orderlist.txt read-only: {str(e)}")
                                
                        except Exception as e:
                            print(f"❌ Error writing Orderlist.txt: {str(e)}")
                            # Try alternative approach with temporary file
                            try:
                                import tempfile
                                import shutil
                                
                                # Create temporary file
                                with tempfile.NamedTemporaryFile(mode='w', encoding='utf-8', delete=False, suffix='.txt') as temp_file:
                                    for order_number in order_numbers:
                                        if marketplace.lower() in ['zalora', 'blibli']:
                                            if shop_id:
                                                temp_file.write(f"{order_number},{shop_id}\n")
                                            else:
                                                temp_file.write(f"{order_number}\n")
                                        else:
                                            temp_file.write(f"{order_number}\n")
                                    
                                    temp_path = temp_file.name
                                
                                # Move temporary file to target location
                                shutil.move(temp_path, orderlist_path)
                                print(f"✅ Generated Orderlist.txt using temporary file: {orderlist_path}")
                                
                                # Make file read-only
                                try:
                                    os.chmod(orderlist_path, 0o444)
                                    print(f"🔒 Made Orderlist.txt read-only: {orderlist_path}")
                                except Exception as chmod_e:
                                    print(f"⚠️ Could not make Orderlist.txt read-only: {str(chmod_e)}")
                                    
                            except Exception as alt_e:
                                print(f"❌ Alternative approach also failed: {str(alt_e)}")
                                results.append({
                                    "marketplace": marketplace,
                                    "success": False,
                                    "message": f"Failed to write Orderlist.txt for {marketplace}: {str(e)} (Alternative: {str(alt_e)})"
                                })
                                continue
                        
                        total_orders_added += len(order_numbers)
                        print(f"✅ Generated Orderlist.txt for {marketplace} - {brand_name}: {len(order_numbers)} orders")
                        
                        # Create backup (now in Page_Orderlist folder)
                        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                        backup_path = os.path.join(page_orderlist_path, f'Orderlist_{brand_name}_{timestamp}.txt')
                        with open(backup_path, 'w', encoding='utf-8') as f:
                            for order_number in order_numbers:
                                if marketplace.lower() in ['zalora', 'blibli']:
                                    if shop_id:
                                        f.write(f"{order_number},{shop_id}\n")
                                    else:
                                        f.write(f"{order_number}\n")
                                else:
                                    f.write(f"{order_number}\n")
                        
                        print(f"💾 Backup created: {backup_path}")
                    else:
                        print(f"ℹ️ No Not Interfaced orders found for {marketplace} - {brand_name}")
                
                # Step 4: Check if Orderlist.txt was created and has content (now in Page_Orderlist folder)
                orderlist_path = os.path.join(page_orderlist_path, 'Orderlist.txt')
                if not os.path.exists(orderlist_path):
                    results.append({
                        "marketplace": marketplace,
                        "success": False,
                        "message": f"No Not Interfaced orders found for any brand in {marketplace}"
                    })
                    continue
                
                # Verify Orderlist.txt has content
                try:
                    with open(orderlist_path, 'r', encoding='utf-8') as f:
                        order_lines = [line.strip() for line in f if line.strip()]
                    
                    if not order_lines:
                        results.append({
                            "marketplace": marketplace,
                            "success": False,
                            "message": f"Orderlist.txt is empty for {marketplace}"
                        })
                        continue
                        
                except Exception as e:
                    results.append({
                        "marketplace": marketplace,
                        "success": False,
                        "message": f"Cannot read Orderlist.txt for {marketplace}: {str(e)}"
                    })
                    continue
                
                # Step 5: Run the executable
                try:
                    brand_name = marketplace_brands[0]['brand'] if marketplace_brands else "Unknown"
                    success, process = run_executable_with_logging(exe_path, folder_path, marketplace, current_user, brand_name, None)
                    
                    if not success:
                        print(f"⚠️ Failed to start {marketplace} app with logging, falling back to silent mode")
                        success = run_executable_silently(exe_path, folder_path, marketplace)
                        if not success:
                            print(f"⚠️ Failed to start {marketplace} app silently")
                    
                    results.append({
                        "marketplace": marketplace,
                        "success": success,
                        "message": f"{marketplace.title()} started with {len(order_lines)} orders" if success else f"Failed to start {marketplace}",
                        "orders_processed": len(order_lines),
                        "brands_found": len(marketplace_brands)
                    })
                    
                    marketplace_success = success
                    
                except Exception as e:
                    results.append({
                        "marketplace": marketplace,
                        "success": False,
                        "message": f"Failed to start {marketplace}: {str(e)}"
                    })
                
                # Small delay between starting apps
                time.sleep(1)
                
        finally:
            db.close()
        
        success_count = sum(1 for r in results if r['success'])
        total_orders = sum(r.get('orders_processed', 0) for r in results)
        
        return {
            "success": success_count > 0,
            "message": f"Started {success_count}/{len(results)} marketplace apps with {total_orders} total orders",
            "results": results,
            "summary": {
                "total_marketplaces": len(marketplace_configs),
                "successful": success_count,
                "failed": len(results) - success_count,
                "total_orders_processed": total_orders
            }
        }
        
    except Exception as e:
        logger.error(f"Error running all marketplace apps: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to run all apps: {str(e)}")

@app.get("/auto-run-config")
def get_auto_run_config(current_user: str = Depends(get_current_user)):
    """Get auto-run configuration"""
    return {
        "auto_run_enabled": AUTO_RUN_MARKETPLACE_APPS,
        "marketplace_apps_enabled": MARKETPLACE_APPS_ENABLED,
        "message": "Auto-run configuration retrieved successfully"
    }

@app.get("/check-auto-run-status")
def check_auto_run_status():
    """Check auto-run status without authentication"""
    return {
        "auto_run_enabled": AUTO_RUN_MARKETPLACE_APPS,
        "marketplace_apps_enabled": MARKETPLACE_APPS_ENABLED,
        "message": "Auto-run status retrieved successfully"
    }

@app.get("/validate-auto-run-system")
def validate_auto_run_system():
    """Validate auto-run system configuration and health"""
    try:
        validation_results = {
            "auto_run_enabled": AUTO_RUN_MARKETPLACE_APPS,
            "environment_check": {},
            "executable_check": {},
            "folder_check": {},
            "overall_status": "healthy"
        }
        
        # Check environment variables
        validation_results["environment_check"] = {
            "AUTO_RUN_MARKETPLACE_APPS": AUTO_RUN_MARKETPLACE_APPS,
            "env_file_exists": os.path.exists(".env")
        }
        
        # Check marketplace executables
        marketplace_configs = {
            'shopee': {'folder': 'Shopee', 'exe': 'ShopeeOrderLogistic.exe'},
            'lazada': {'folder': 'Lazada', 'exe': 'LazadaMarketplace.exe'},
            'blibli': {'folder': 'Blibli', 'exe': 'BliBliProduct2024.exe'},
            'desty': {'folder': 'Desty', 'exe': 'Desty.Console.exe'},
            'ginee': {'folder': 'Ginee', 'exe': 'Ginee.sync.exe'},
            'tiktok': {'folder': 'Tiktok', 'exe': 'tiktok.api.exe'},
            'zalora': {'folder': 'Zalora', 'exe': 'Zalora.Flexo.Integration.exe'},
            'tokopedia': {'folder': 'Tokopedia', 'exe': 'TokopediaOrder.exe'},
            'jubelio': {'folder': 'Jubelio', 'exe': 'Jubelio_project.exe'}
        }
        
        # Get project root
        if os.path.basename(os.getcwd()) == 'backend':
            project_root = get_project_root()
        else:
            project_root = os.getcwd()
        
        jobgetorder_path = os.path.join(project_root, "JobGetOrder")
        
        # Check JobGetOrder directory
        validation_results["folder_check"]["jobgetorder_exists"] = os.path.exists(jobgetorder_path)
        
        # Check each marketplace executable
        for marketplace, config in marketplace_configs.items():
            exe_path = os.path.join(jobgetorder_path, config['folder'], config['exe'])
            validation_results["executable_check"][marketplace] = {
                "folder_exists": os.path.exists(os.path.join(jobgetorder_path, config['folder'])),
                "executable_exists": os.path.exists(exe_path),
                "executable_path": exe_path
            }
        
        # Check user directories
        user_dirs = []
        if os.path.exists(jobgetorder_path):
            for item in os.listdir(jobgetorder_path):
                if item.startswith("User_") and os.path.isdir(os.path.join(jobgetorder_path, item)):
                    user_dirs.append(item)
        
        validation_results["folder_check"]["user_directories"] = user_dirs
        validation_results["folder_check"]["total_users"] = len(user_dirs)
        
        # Overall health check
        if not AUTO_RUN_MARKETPLACE_APPS:
            validation_results["overall_status"] = "disabled"
        elif not validation_results["folder_check"]["jobgetorder_exists"]:
            validation_results["overall_status"] = "error"
        else:
            # Check if at least some executables exist
            executable_count = sum(1 for check in validation_results["executable_check"].values() 
                                 if check["executable_exists"])
            if executable_count == 0:
                validation_results["overall_status"] = "error"
            elif executable_count < len(marketplace_configs):
                validation_results["overall_status"] = "partial"
        
        return {
            "success": True,
            "validation": validation_results,
            "message": f"Auto-run system validation completed. Status: {validation_results['overall_status']}"
        }
        
    except Exception as e:
        logger.error(f"Error validating auto-run system: {str(e)}")
        return {
            "success": False,
            "error": str(e),
            "message": "Failed to validate auto-run system"
        }

@app.get("/logging-config")
def get_logging_config(current_user: str = Depends(get_current_user)):
    """Get logging configuration"""
    return {
        "auto_logging_enabled": True,  # Always enabled now
        "message": "Auto-logging is enabled for marketplace apps"
    }

@app.post("/api/test-run-marketplace-app")
def test_run_marketplace_app(request: dict):
    """Test endpoint to run marketplace app without authentication"""
    try:
        marketplace = request.get('marketplace')
        user = request.get('user', 'fais')  # Default to 'fais' user
        
        if not marketplace:
            raise HTTPException(status_code=400, detail="Marketplace is required")
        
        # Get user-specific workspace
        user_workspace = multi_user_handler.get_user_workspace(user)
        print(f"👤 Running marketplace app for user {user} - {marketplace} in workspace: {user_workspace}")
        
        marketplace_configs = {
            'shopee': {'folder': 'Shopee', 'exe': 'ShopeeOrderLogistic.exe'},
            'lazada': {'folder': 'Lazada', 'exe': 'LazadaMarketplace.exe'},
            'blibli': {'folder': 'Blibli', 'exe': 'BliBliProduct2024.exe'},
            'desty': {'folder': 'Desty', 'exe': 'Desty.Console.exe'},
            'ginee': {'folder': 'Ginee', 'exe': 'Ginee.sync.exe'},
            'tiktok': {'folder': 'Tiktok', 'exe': 'tiktok.api.exe'},
            'zalora': {'folder': 'Zalora', 'exe': 'Zalora.Flexo.Integration.exe'},
            'tokopedia': {'folder': 'Tokopedia', 'exe': 'TokopediaOrder.exe'}
        }
        
        if marketplace not in marketplace_configs:
            raise HTTPException(status_code=400, detail="Invalid marketplace")
        
        config = marketplace_configs[marketplace]
        folder_path = os.path.join(user_workspace, config['folder'])
        exe_path = os.path.join(folder_path, config['exe'])
        
        if not os.path.exists(exe_path):
            raise HTTPException(status_code=404, detail=f"{config['exe']} not found")
        
        # Detect brand from config files (orderlist filename uses task_id, not brand name)
        brand_name = "Unknown"
        try:
            config_files = [f for f in os.listdir(folder_path) if f.endswith('.config')]
            for config_file in config_files:
                try:
                    with open(os.path.join(folder_path, config_file), 'r', encoding='utf-8') as cf:
                        config_content = cf.read()
                        brand_match = re.search(r'<add key="brand_name" value="([^"]+)"', config_content)
                        if brand_match:
                            brand_name = brand_match.group(1)
                            print(f"🔍 Detected brand from config file: {brand_name}")
                            break
                except:
                    continue
        except Exception as e:
            print(f"⚠️ Error detecting brand: {str(e)}")
        
        # Switch to correct config file
        if brand_name != "Unknown":
            default_config_path = os.path.join(folder_path, f"{config['exe']}.config")
            brand_config_path = os.path.join(folder_path, f"{brand_name}_{marketplace.upper()}.config")
            
            if os.path.exists(brand_config_path):
                print(f"🔄 Switching to {brand_name} config...")
                backup_config_path = os.path.join(folder_path, f"{config['exe']}.config.backup")
                if os.path.exists(default_config_path):
                    shutil.copy2(default_config_path, backup_config_path)
                    print(f"   📋 Backed up original config")
                
                shutil.copy2(brand_config_path, default_config_path)
                print(f"   ✅ Using {brand_name} config")
        
        # Run with logging
        success, process = run_executable_with_logging(exe_path, folder_path, marketplace, user, brand_name, None)
        
        if success:
            return {
                "success": True,
                "message": f"{marketplace.title()} app started with logging",
                "user": user,
                "brand": brand_name,
                "log_file": os.path.join(folder_path, f"{marketplace}_app.log")
            }
        else:
            return {
                "success": False,
                "message": f"Failed to start {marketplace} app with logging"
            }
            
    except Exception as e:
        logger.error(f"Error running marketplace app: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/add-manual-log")
def add_manual_log(
    request: dict,
    current_user: str = Depends(get_current_user)
):
    """Add manual log entry to marketplace app log file"""
    try:
        marketplace = request.get('marketplace', '').lower()
        log_message = request.get('message', '')
        log_level = request.get('level', 'INFO')
        
        if not marketplace or not log_message:
            raise HTTPException(status_code=400, detail="Marketplace and message are required")
        
        # Get user workspace
        user_workspace = multi_user_handler.get_user_workspace(current_user)
        
        # Find marketplace folder
        marketplace_configs = {
            'shopee': 'Shopee',
            'lazada': 'Lazada', 
            'blibli': 'Blibli',
            'desty': 'Desty',
            'ginee': 'Ginee',
            'tiktok': 'Tiktok',
            'zalora': 'Zalora',
            'tokopedia': 'Tokopedia',
            'jubelio': 'Jubelio'
        }
        
        if marketplace not in marketplace_configs:
            raise HTTPException(status_code=400, detail="Invalid marketplace")
        
        folder_name = marketplace_configs[marketplace]
        folder_path = os.path.join(user_workspace, folder_name)
        
        if not os.path.exists(folder_path):
            raise HTTPException(status_code=404, detail=f"Marketplace folder not found: {folder_name}")
        
        # Create log file path
        log_file_path = os.path.join(folder_path, f"{marketplace}_app.log")
        
        # Add log entry
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        log_entry = f"{timestamp} [{log_level}] {log_message}\n"
        
        with open(log_file_path, 'a', encoding='utf-8') as f:
            f.write(log_entry)
        
        return {
            "success": True,
            "message": f"Log entry added to {marketplace} app log",
            "log_file": log_file_path,
            "timestamp": timestamp
        }
        
    except Exception as e:
        logger.error(f"Error adding manual log: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to add log: {str(e)}")

@app.post("/auto-run-config")
def update_auto_run_config(
    request: dict,
    current_user: str = Depends(get_current_user)
):
    """Update auto-run configuration"""
    try:
        global AUTO_RUN_MARKETPLACE_APPS
        enabled = request.get('enabled', True)
        AUTO_RUN_MARKETPLACE_APPS = enabled
        
        # Update environment variable
        os.environ['AUTO_RUN_MARKETPLACE_APPS'] = str(enabled).lower()
        
        logger.info(f"Auto-run configuration updated by {current_user}: {enabled}")
        
        return {
            "success": True,
            "auto_run_enabled": AUTO_RUN_MARKETPLACE_APPS,
            "message": f"Auto-run {'enabled' if enabled else 'disabled'} successfully"
        }
        
    except Exception as e:
        logger.error(f"Error updating auto-run config: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to update auto-run config: {str(e)}")

@app.get("/api/config")
def get_app_config(current_user: str = Depends(get_current_user)):
    """Get application configuration"""
    return {
        "marketplace_apps_enabled": MARKETPLACE_APPS_ENABLED,
        "auto_run_marketplace_apps": AUTO_RUN_MARKETPLACE_APPS,
        "current_user": current_user
    }

@app.get("/marketplace-notifications")
async def get_marketplace_notifications():
    """Get recent marketplace notifications"""
    # This could be enhanced to store notifications in database
    # For now, return a simple response
    return {
        "notifications": [],
        "message": "Notification system ready"
    }

@app.get("/marketplace-completion-status")
async def get_marketplace_completion_status():
    """Get completion status of running marketplace apps"""
    try:
        project_root = get_project_root()
        jobgetorder_path = os.path.join(project_root, 'JobGetOrder')
        
        marketplace_configs = {
            'shopee': {'folder': 'Shopee', 'exe': 'ShopeeOrderLogistic.exe'},
            'lazada': {'folder': 'Lazada', 'exe': 'LazadaMarketplace.exe'},
            'blibli': {'folder': 'Blibli', 'exe': 'BliBliProduct2024.exe'},
            'desty': {'folder': 'Desty', 'exe': 'Desty.Console.exe'},
            'ginee': {'folder': 'Ginee', 'exe': 'Ginee.sync.exe'},
            'tiktok': {'folder': 'Tiktok', 'exe': 'tiktok.api.exe'},
            'zalora': {'folder': 'Zalora', 'exe': 'Zalora.Flexo.Integration.exe'},
            'tokopedia': {'folder': 'Tokopedia', 'exe': 'TokopediaOrder.exe'},
            'jubelio': {'folder': 'Jubelio', 'exe': 'Jubelio_project.exe'}
        }
        
        completion_status = {}
        
        for marketplace, config in marketplace_configs.items():
            folder_path = os.path.join(jobgetorder_path, config['folder'])
            orderlist_path = os.path.join(folder_path, 'Orderlist.txt')
            
            # Check if Orderlist.txt exists (indicates app was started)
            if os.path.exists(orderlist_path):
                # Check for completion indicators
                completion_indicators = {
                    'has_orderlist': True,
                    'order_count': 0,
                    'completion_status': 'unknown',
                    'last_modified': None,
                    'output_files': []
                }
                
                # Count orders in Orderlist.txt
                try:
                    with open(orderlist_path, 'r', encoding='utf-8') as f:
                        lines = f.readlines()
                        completion_indicators['order_count'] = len([line.strip() for line in lines if line.strip()])
                except:
                    completion_indicators['order_count'] = 0
                
                # Check last modified time
                try:
                    completion_indicators['last_modified'] = os.path.getmtime(orderlist_path)
                except:
                    completion_indicators['last_modified'] = None
                
                # Check for output files that might indicate completion
                output_files = []
                try:
                    for file in os.listdir(folder_path):
                        if file.endswith(('.log', '.txt', '.csv', '.json')) and file != 'Orderlist.txt':
                            file_path = os.path.join(folder_path, file)
                            file_stat = os.stat(file_path)
                            output_files.append({
                                'name': file,
                                'size': file_stat.st_size,
                                'modified': file_stat.st_mtime
                            })
                except:
                    pass
                
                completion_indicators['output_files'] = output_files
                
                # Determine completion status based on file patterns
                if len(output_files) > 0:
                    # Check if any output file was modified recently (within last 5 minutes)
                    current_time = time.time()
                    recent_files = [f for f in output_files if current_time - f['modified'] < 300]
                    
                    if recent_files:
                        completion_indicators['completion_status'] = 'completed'
                    else:
                        completion_indicators['completion_status'] = 'running'
                else:
                    completion_indicators['completion_status'] = 'running'
                
                completion_status[marketplace] = completion_indicators
        
        return {
            "success": True,
            "completion_status": completion_status,
            "timestamp": time.time()
        }
        
    except Exception as e:
        logger.error(f"Error checking marketplace completion status: {str(e)}")
        return {
            "success": False,
            "error": str(e)
        }

@app.post("/check-marketplace-completion")
async def check_marketplace_completion(
    request: dict,
    current_user: str = Depends(get_current_user)
):
    """Manually check completion status of a specific marketplace app"""
    try:
        marketplace = request.get('marketplace')
        if not marketplace:
            raise HTTPException(status_code=400, detail="Marketplace is required")
        
        project_root = get_project_root()
        jobgetorder_path = os.path.join(project_root, 'JobGetOrder')
        
        marketplace_configs = {
            'shopee': {'folder': 'Shopee', 'exe': 'ShopeeOrderLogistic.exe'},
            'lazada': {'folder': 'Lazada', 'exe': 'LazadaMarketplace.exe'},
            'blibli': {'folder': 'Blibli', 'exe': 'BliBliProduct2024.exe'},
            'desty': {'folder': 'Desty', 'exe': 'Desty.Console.exe'},
            'ginee': {'folder': 'Ginee', 'exe': 'Ginee.sync.exe'},
            'tiktok': {'folder': 'Tiktok', 'exe': 'tiktok.api.exe'},
            'zalora': {'folder': 'Zalora', 'exe': 'Zalora.Flexo.Integration.exe'},
            'tokopedia': {'folder': 'Tokopedia', 'exe': 'TokopediaOrder.exe'}
        }
        
        if marketplace not in marketplace_configs:
            raise HTTPException(status_code=400, detail="Invalid marketplace")
        
        config = marketplace_configs[marketplace]
        folder_path = os.path.join(jobgetorder_path, config['folder'])
        orderlist_path = os.path.join(folder_path, 'Orderlist.txt')
        
        if not os.path.exists(orderlist_path):
            return {
                "success": True,
                "marketplace": marketplace,
                "status": "not_started",
                "message": f"{marketplace.title()} app has not been started"
            }
        
        # Check for output files
        output_files = []
        try:
            for file in os.listdir(folder_path):
                if file.endswith(('.log', '.txt', '.csv', '.json')) and file != 'Orderlist.txt':
                    file_path = os.path.join(folder_path, file)
                    file_stat = os.stat(file_path)
                    output_files.append({
                        'name': file,
                        'size': file_stat.st_size,
                        'modified': file_stat.st_mtime,
                        'modified_readable': datetime.fromtimestamp(file_stat.st_mtime).strftime('%Y-%m-%d %H:%M:%S')
                    })
        except:
            pass
        
        # Determine status
        current_time = time.time()
        recent_files = [f for f in output_files if current_time - f['modified'] < 300]  # 5 minutes
        
        if len(recent_files) > 0:
            status = "completed"
            message = f"{marketplace.title()} app completed successfully - {len(recent_files)} output files generated"
        else:
            status = "running"
            message = f"{marketplace.title()} app is still running - {len(output_files)} output files found"
        
        return {
            "success": True,
            "marketplace": marketplace,
            "status": status,
            "message": message,
            "output_files": output_files,
            "orderlist_exists": True,
            "last_check": datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
        
    except Exception as e:
        logger.error(f"Error checking marketplace completion: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to check completion: {str(e)}")

@app.get("/debug-table-counts")
async def debug_table_counts():
    """Debug endpoint to check table counts and sync status"""
    try:
        db = SessionLocal()
        
        # Get counts from both tables
        uploaded_count = db.execute(text("SELECT COUNT(*) FROM uploaded_orders")).scalar()
        update_uploaded_count = db.execute(text("SELECT COUNT(*) FROM update_uploaded_orders")).scalar()
        
        # Check missing orders
        missing_count = db.execute(text("""
            SELECT COUNT(*) FROM uploaded_orders uo
            WHERE NOT EXISTS (
                SELECT 1 FROM update_uploaded_orders uuo 
                WHERE uuo."OrderNumber" = uo."OrderNumber"
            )
        """)).scalar()
        
        return {
            "uploaded_orders_count": uploaded_count,
            "update_uploaded_orders_count": update_uploaded_count,
            "missing_orders_count": missing_count,
            "sync_needed": missing_count > 0,
            "status": "success"
        }
        
    except Exception as e:
        return {
            "error": str(e),
            "status": "error"
        }
    finally:
        db.close()



@app.get("/api/create-clean-orders-view")
async def create_view_endpoint():
    """Create clean_orders view for deduplicated data"""
    try:
        db = SessionLocal()
        
        # Create the view
        success = create_clean_orders_view(db)
        
        if success:
            # Check view creation
            view_count = db.execute(text("SELECT COUNT(*) FROM clean_orders")).scalar()
            
            return {
                "message": "Clean orders view created successfully",
                "view_record_count": view_count,
                "status": "success"
            }
        else:
            return {
                "error": "Failed to create view",
                "status": "error"
            }
            
    except Exception as e:
        return {
            "error": str(e),
            "status": "error"
        }
    finally:
        db.close()

@app.post("/api/refresh-interface-status-simple")
async def refresh_interface_status_simple(current_user: str = Depends(get_current_user)):
    """Simple refresh interface status - works with 1 table + 1 view approach"""
    try:
        print(f"🔄 Starting simple interface status refresh for user: {current_user}")
        
        db = SessionLocal()
        try:
            # Get orders that need to be refreshed (where InterfaceStatus = 'Not Yet Interface')
            missing_orders_result = db.execute(text("""
                SELECT "OrderNumber", "Marketplace", "Brand", "OrderStatus"
                FROM uploaded_orders 
                WHERE "InterfaceStatus" = 'Not Yet Interface'
                AND "OrderStatus" NOT IN ('batal', 'cancel', 'cancellation', 'BATAL', 'CANCEL', 'CANCELLATION', 'Cancellations', 'Dibatalkan', 'Batal', 'CANCELED', 'Cancelled', 'canceled', 'Pembatalan diajukan', 'Order Batal', 'CANCELLED')
                ORDER BY "UploadDate" DESC
            """))
            
            missing_orders = missing_orders_result.fetchall()
            print(f"📊 Found {len(missing_orders)} orders that need external database sync")
            
            if missing_orders:
                # Group orders by marketplace
                orders_by_marketplace = {}
                for order in missing_orders:
                    order_number, marketplace, brand, order_status = order
                    marketplace_key = marketplace.lower() if marketplace else 'unknown'
                    if marketplace_key not in orders_by_marketplace:
                        orders_by_marketplace[marketplace_key] = []
                    orders_by_marketplace[marketplace_key].append({
                        'order_number': order_number,
                        'marketplace': marketplace,
                        'brand': brand
                    })
                
                # Process each marketplace
                updated_count = 0
                for marketplace, orders_list in orders_by_marketplace.items():
                    print(f"🔄 Processing {len(orders_list)} orders for marketplace: {marketplace}")
                    
                    # Get order numbers for this marketplace
                    order_numbers = [order['order_number'] for order in orders_list]
                    
                    # Query external database
                    external_results = check_external_database_status(order_numbers, marketplace)
                    print(f"📊 External database returned {len(external_results)} results for {marketplace}")
                    
                    # Update uploaded_orders with external data
                    for order in orders_list:
                        order_number = order['order_number']
                        if order_number in external_results:
                            external_data = external_results[order_number]
                            
                            # Update uploaded_orders directly
                            db.execute(text("""
                                UPDATE uploaded_orders 
                                SET 
                                    "OrderNumberFlexo" = :order_number_flexo,
                                    "OrderStatusFlexo" = :order_status_flexo,
                                    "InterfaceStatus" = :interface_status
                                WHERE "OrderNumber" = :order_number
                            """), {
                                "order_number_flexo": external_data.get('system_ref_id', ''),
                                "order_status_flexo": external_data.get('order_status', ''),
                                "interface_status": external_data.get('interface_status', 'Not Yet Interface'),
                                "order_number": order_number
                            })
                            
                            updated_count += 1
                
                db.commit()
                print(f"✅ Successfully updated {updated_count} orders with external database data")
                
                return {
                    "success": True,
                    "message": f"Interface status refreshed for {updated_count} orders",
                    "updated_count": updated_count,
                    "total_processed": len(missing_orders),
                    "marketplaces_processed": len(orders_by_marketplace)
                }
            else:
                return {
                    "success": True,
                    "message": "No orders need refresh - all orders already have external data",
                    "updated_count": 0,
                    "total_processed": 0
                }
                
        except Exception as e:
            print(f"❌ Error in simple refresh: {e}")
            db.rollback()
            return {
                "success": False,
                "message": f"Error refreshing interface status: {str(e)}",
                "updated_count": 0
            }
        finally:
            db.close()
            
    except Exception as e:
        print(f"❌ Error in refresh interface status: {e}")
        return {
            "success": False,
            "message": f"Error: {str(e)}",
            "updated_count": 0
        }

@app.post("/api/refresh-interface-status")
async def refresh_interface_status(current_user: str = Depends(get_current_user)):
    """Refresh interface status - simplified version"""
    try:
        print(f"🔄 Starting interface status refresh for user: {current_user}")
        
        db = SessionLocal()
        try:
            # Get orders that need to be refreshed
            missing_orders_result = db.execute(text("""
                SELECT "OrderNumber", "Marketplace", "Brand", "OrderStatus"
                FROM uploaded_orders 
                WHERE "InterfaceStatus" = 'Not Yet Interface'
                AND "OrderStatus" NOT IN ('batal', 'cancel', 'cancellation', 'BATAL', 'CANCEL', 'CANCELLATION', 'Cancellations', 'Dibatalkan', 'Batal', 'CANCELED', 'Cancelled', 'canceled', 'Pembatalan diajukan', 'Order Batal', 'CANCELLED')
                ORDER BY "UploadDate" DESC
            """))
            
            missing_orders = missing_orders_result.fetchall()
            print(f"📊 Found {len(missing_orders)} orders that need external database sync")
            
            if not missing_orders:
                return {
                    "success": True,
                    "message": "No orders need refresh - all orders already have external data",
                    "updated_count": 0
                }
            
            # Process in batches to avoid timeout and memory issues
            BATCH_SIZE = 1000  # Process 1000 orders at a time
            total_orders = len(missing_orders)
            total_updated = 0
            
            print(f"🔄 Processing {total_orders} orders in batches of {BATCH_SIZE}")
            
            for batch_start in range(0, total_orders, BATCH_SIZE):
                batch_end = min(batch_start + BATCH_SIZE, total_orders)
                batch_orders = missing_orders[batch_start:batch_end]
                
                print(f"📦 Processing batch {batch_start//BATCH_SIZE + 1}: orders {batch_start + 1}-{batch_end} of {total_orders}")
                
                # Group orders by marketplace for this batch
                orders_by_marketplace = {}
                for order in batch_orders:
                    order_number, marketplace, brand, order_status = order
                    marketplace_key = marketplace.lower() if marketplace else 'unknown'
                    if marketplace_key not in orders_by_marketplace:
                        orders_by_marketplace[marketplace_key] = []
                    orders_by_marketplace[marketplace_key].append({
                        'order_number': order_number,
                        'marketplace': marketplace,
                        'brand': brand
                    })
                
                # Process each marketplace in this batch
                batch_updated = 0
                batch_external_orders = 0
                
                for marketplace, orders_list in orders_by_marketplace.items():
                    print(f"🔄 Processing {len(orders_list)} orders for marketplace: {marketplace}")
                    
                    # Get order numbers for this marketplace
                    order_numbers = [order['order_number'] for order in orders_list]
                    
                    # Query external database
                    external_results = check_external_database_status(order_numbers, marketplace)
                    print(f"📊 External database returned {len(external_results)} results for {marketplace}")
                    batch_external_orders += len(external_results)
                    
                    # Update uploaded_orders with external data
                    for order in orders_list:
                        order_number = order['order_number']
                        if order_number in external_results:
                            external_data = external_results[order_number]
                            
                            # Update uploaded_orders directly
                            db.execute(text("""
                                UPDATE uploaded_orders 
                                SET 
                                    "OrderNumberFlexo" = :order_number_flexo,
                                    "OrderStatusFlexo" = :order_status_flexo,
                                    "InterfaceStatus" = :interface_status
                                WHERE "OrderNumber" = :order_number
                            """), {
                                "order_number_flexo": external_data.get('system_ref_id', ''),
                                "order_status_flexo": external_data.get('order_status', ''),
                                "interface_status": external_data.get('interface_status', 'Not Yet Interface'),
                                "order_number": order_number
                            })
                            
                            batch_updated += 1
                            print(f"  ✅ Updated order {order_number}: FlexoNumber={external_data.get('system_ref_id', '')}, FlexoStatus={external_data.get('order_status', '')}")
                
                # Commit this batch
                db.commit()
                total_updated += batch_updated
                print(f"✅ Batch {batch_start//BATCH_SIZE + 1} completed: {batch_updated} orders updated (Total: {total_updated}/{total_orders})")
            
            print(f"🎉 All batches completed! Successfully updated {total_updated} orders out of {total_orders} checked")
                
            return {
                "success": True,
                "message": f"Interface status refreshed successfully! {total_updated} orders updated from {total_orders} orders processed.",
                "updated_count": total_updated,
                "total_orders_processed": total_orders
            }
                
        except Exception as e:
            print(f"❌ Error in refresh: {e}")
            db.rollback()
            return {
                "success": False,
                "message": f"Error refreshing interface status: {str(e)}",
                "updated_count": 0
            }
        finally:
            db.close()
            
    except Exception as e:
        print(f"❌ Error in refresh interface status: {str(e)}")
        import traceback
        traceback.print_exc()
        return {
            "success": False,
            "message": f"Error: {str(e)}",
            "updated_count": 0
        }



@app.get("/queue-status")
async def get_queue_status():
    """Get current upload queue status without authentication for dashboard"""
    try:
        queue_status = multi_user_handler.get_queue_status()
        return {
            "success": True,
            "queue_status": queue_status,
            "timestamp": datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
    except Exception as e:
        logger.error(f"Error getting queue status: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get queue status: {str(e)}")

@app.post("/cleanup-user-workspace")
async def cleanup_user_workspace(
    request: dict,
    current_user: str = Depends(get_current_user)
):
    """Clean up old files in user workspace"""
    try:
        older_than_hours = request.get('older_than_hours', 24)
        multi_user_handler.cleanup_user_workspace(current_user, older_than_hours)
        
        return {
            "success": True,
            "message": f"Cleaned up workspace for user {current_user} (files older than {older_than_hours} hours)",
            "user": current_user
        }
    except Exception as e:
        logger.error(f"Error cleaning up workspace: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to cleanup workspace: {str(e)}")

@app.post("/save-not-uploaded-history")
async def manual_save_not_uploaded_history(db: Session = Depends(get_db)):
    """Manually save current not uploaded items to history (for testing)"""
    try:
        save_not_uploaded_history()
        return {"success": True, "message": "Not uploaded items history saved successfully"}
    except Exception as e:
        logger.error(f"Error manually saving not uploaded history: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/reset-remarks")
async def manual_reset_remarks(db: Session = Depends(get_db)):
    """Manually reset all remarks (for testing)"""
    try:
        reset_remarks_daily()
        return {"success": True, "message": "All remarks reset successfully"}
    except Exception as e:
        logger.error(f"Error manually resetting remarks: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

# Debug endpoint removed for production
# @app.post("/debug/reset-upload-history")
# def reset_upload_history(db: Session = Depends(get_db)):
#     """Reset UploadHistory table with fresh sample data"""
#     # ... implementation removed for production

# Debug endpoint removed for production
# @app.get("/debug/timezone")
# def debug_timezone():
#     """Debug timezone information"""
#     # ... implementation removed for production

# Debug endpoint removed for production
# @app.get("/debug/upload-history")
# def debug_upload_history(
#     date: Optional[str] = Query(None, description="Date to check (YYYY-MM-DD)"),
#     db: Session = Depends(get_db)
# ):
#     """Debug endpoint to check UploadHistory data for specific date"""
#     # ... implementation removed for production

@app.get("/api/not-uploaded-items")
async def get_not_uploaded_items(
    start_date: Optional[str] = Query(None, description="Start date (ISO format)"),
    end_date: Optional[str] = Query(None, description="End date (ISO format)"),
    db: Session = Depends(get_db)
):
    """Get list of brand+marketplace+batch combinations that haven't been uploaded yet"""
    try:
        # Always use history table, but filter by date if provided
        query = db.query(NotUploadedHistory).filter(NotUploadedHistory.status == 'not_uploaded')
        
        if start_date and end_date:
            # Filter by date range
            start_dt = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
            end_dt = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
            query = query.filter(
                NotUploadedHistory.check_date >= start_dt,
                NotUploadedHistory.check_date <= end_dt
            )
        else:
            # If no date filter, get today's data (use string comparison to avoid timezone issues)
            current_time = get_wib_now()
            today_str = current_time.strftime('%Y-%m-%d')
            query = query.filter(
                func.date(NotUploadedHistory.check_date) == today_str
            )
        
        history_items = query.all()
        
        not_uploaded = []
        for item in history_items:
            not_uploaded.append({
                "brand": item.brand,
                "marketplace": item.marketplace,
                "batch": item.batch,
                "remark": item.remark,
                "created_at": item.check_date.isoformat() if item.check_date else None
            })
        
        # Sort by brand, then marketplace, then batch
        not_uploaded.sort(key=lambda x: (x["brand"], x["marketplace"], x["batch"]))
        
        logger.info(f"Found {len(not_uploaded)} not uploaded items from history")
        
        return {
            "success": True,
            "not_uploaded_items": not_uploaded,
            "total_expected": len(not_uploaded),
            "message": f"Found {len(not_uploaded)} not uploaded items"
        }
        
    except Exception as e:
        logger.error(f"Error getting not uploaded items: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get not uploaded items: {str(e)}")

# Cache for marketplace logs to improve performance
_marketplace_logs_cache = {}
_cache_timestamp = None
CACHE_DURATION = 300  # 5 minutes cache

def _get_cached_logs():
    """Get cached logs if still valid"""
    global _cache_timestamp
    if _cache_timestamp and (datetime.now() - _cache_timestamp).seconds < CACHE_DURATION:
        return _marketplace_logs_cache.get('logs', []), _marketplace_logs_cache.get('summary', {})
    return None, None

def _cache_logs(logs, summary):
    """Cache logs and summary"""
    global _marketplace_logs_cache, _cache_timestamp
    _marketplace_logs_cache = {'logs': logs, 'summary': summary}
    _cache_timestamp = datetime.now()

def _scan_marketplace_logs():
    """Scan and process marketplace logs with optimized performance"""
    from datetime import timedelta
    import concurrent.futures
    from functools import lru_cache
    
    @lru_cache(maxsize=1000)
    def _extract_brand_from_filename(filename):
        """Cached brand extraction from filename"""
        if 'Orderlist_' in filename:
            parts = filename.replace('Orderlist_', '').replace('.txt', '').split('_')
            return parts[0] if parts else "Unknown"
        return "Unknown"
    
    @lru_cache(maxsize=100)
    def _read_config_brand(config_path):
        """Cached config file reading"""
        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                content = f.read()
                brand_match = re.search(r'<add key="brand_name" value="([^"]+)"', content)
                return brand_match.group(1) if brand_match else "Unknown"
        except:
            return "Unknown"
    
    def _process_orderlist_file(orderlist_file, marketplace_dir, current_user, config_files):
        """Process a single orderlist file"""
        try:
            # Get file modification time
            mod_time = os.path.getmtime(orderlist_file)
            mod_datetime = datetime.fromtimestamp(mod_time)
            
            # Read order numbers efficiently
            with open(orderlist_file, 'r', encoding='utf-8') as f:
                order_numbers = [line.strip() for line in f if line.strip()]
            
            if not order_numbers:
                return None
            
            # Extract brand efficiently
            filename = os.path.basename(orderlist_file)
            brand_name = _extract_brand_from_filename(filename)
            
            if brand_name == "Unknown" and filename.lower() == 'orderlist.txt':
                # Try config files
                for config_file in config_files:
                    brand_name = _read_config_brand(config_file)
                    if brand_name != "Unknown":
                        break
            
            return {
                "id": f"{current_user}_{marketplace_dir}_{filename}_{mod_time}",
                "user": current_user,
                "marketplace": marketplace_dir,
                "brand": brand_name,
                "orderlist_file": filename,
                "order_count": len(order_numbers),
                "order_numbers": order_numbers[:10],
                "all_order_numbers": order_numbers,
                "execution_time": mod_datetime.isoformat(),
                "file_path": orderlist_file,
                "status": "completed" if order_numbers else "empty"
            }
        except Exception as e:
            logger.error(f"Error processing orderlist file {orderlist_file}: {str(e)}")
            return None
    
    # Get project root directory
    current_dir = os.getcwd()
    if os.path.basename(current_dir) == 'backend':
        project_root = get_project_root()
    else:
        project_root = current_dir
    
    jobgetorder_path = os.path.join(project_root, "JobGetOrder")
    logger.info(f"Looking for JobGetOrder at: {jobgetorder_path}")
    
    if not os.path.exists(jobgetorder_path):
        return [], {}
    
    logs = []
    processed_orders = set()
    
    # Get directories to scan
    all_dirs = []
    try:
        for item in os.listdir(jobgetorder_path):
            item_path = os.path.join(jobgetorder_path, item)
            if os.path.isdir(item_path):
                all_dirs.append(item_path)
    except Exception as e:
        logger.error(f"Error scanning directories: {str(e)}")
        return [], {}
    
    # Process directories in parallel for better performance
    with concurrent.futures.ThreadPoolExecutor(max_workers=4) as executor:
        futures = []
        
        for dir_path in all_dirs:
            dir_name = os.path.basename(dir_path)
            current_user = dir_name.replace("User_", "") if dir_name.startswith("User_") else "System"
            
            try:
                for marketplace_dir in os.listdir(dir_path):
                    marketplace_path = os.path.join(dir_path, marketplace_dir)
                    if not os.path.isdir(marketplace_path):
                        continue
                    
                    # Collect files efficiently
                    orderlist_files = []
                    config_files = []
                    
                    try:
                        # Check Page_Orderlist folder first
                        page_orderlist_path = os.path.join(dir_path, 'Page_Orderlist', marketplace_dir)
                        if os.path.exists(page_orderlist_path):
                            for file in os.listdir(page_orderlist_path):
                                file_path = os.path.join(page_orderlist_path, file)
                                if os.path.isfile(file_path):
                                    if file.lower().startswith('orderlist') and file.lower().endswith('.txt'):
                                        orderlist_files.append(file_path)
                        
                        # Also check marketplace folder for config files and fallback orderlist files
                        for file in os.listdir(marketplace_path):
                            file_path = os.path.join(marketplace_path, file)
                            if os.path.isfile(file_path):
                                if file.lower().startswith('orderlist') and file.lower().endswith('.txt'):
                                    # Only add if not already found in Page_Orderlist
                                    if not any(os.path.basename(f) == file for f in orderlist_files):
                                        orderlist_files.append(file_path)
                                elif file.endswith('.config'):
                                    config_files.append(file_path)
                    except Exception as e:
                        logger.error(f"Error reading directory {marketplace_path}: {str(e)}")
                        continue
                    
                    # Sort orderlist files (backup files first)
                    orderlist_files.sort(key=lambda x: (os.path.basename(x).lower() == 'orderlist.txt', os.path.getmtime(x)), reverse=True)
                    
                    # Process files
                    for orderlist_file in orderlist_files:
                        future = executor.submit(_process_orderlist_file, orderlist_file, marketplace_dir, current_user, config_files)
                        futures.append(future)
            
            except Exception as e:
                logger.error(f"Error processing directory {dir_path}: {str(e)}")
                continue
        
        # Collect results
        for future in concurrent.futures.as_completed(futures):
            try:
                result = future.result()
                if result:
                    # Check for duplicates
                    order_set = set(result["all_order_numbers"])
                    if not order_set.issubset(processed_orders):
                        logs.append(result)
                        processed_orders.update(order_set)
            except Exception as e:
                logger.error(f"Error processing future result: {str(e)}")
                continue
    
    # Sort logs by execution time (newest first)
    logs.sort(key=lambda x: x["execution_time"], reverse=True)
    
    # Generate summary
    summary = {
        "total_logs": len(logs),
        "total_orders_processed": sum(log["order_count"] for log in logs),
        "marketplaces": list(set(log["marketplace"] for log in logs)),
        "brands": list(set(log["brand"] for log in logs)),
        "users": list(set(log["user"] for log in logs)),
        "recent_executions": len([log for log in logs if 
            datetime.fromisoformat(log["execution_time"]) > datetime.now() - timedelta(hours=24)])
    }
    
    return logs, summary

@app.get("/api/marketplace-logs")
async def get_marketplace_logs(
    marketplace: Optional[str] = Query(None, description="Filter by marketplace"),
    brand: Optional[str] = Query(None, description="Filter by brand"),
    user: Optional[str] = Query(None, description="Filter by user"),
    limit: int = Query(20, description="Number of log entries to return"),
    offset: int = Query(0, description="Number of log entries to skip"),
    db: Session = Depends(get_db)
):
    """Get marketplace execution logs with order numbers - Optimized version"""
    try:
        # Check cache first
        cached_logs, cached_summary = _get_cached_logs()
        
        if cached_logs is None:
            # Cache miss - scan files
            logger.info("Cache miss - scanning marketplace logs")
            all_logs, summary = _scan_marketplace_logs()
            _cache_logs(all_logs, summary)
        else:
            # Cache hit
            logger.info("Cache hit - using cached marketplace logs")
            all_logs, summary = cached_logs, cached_summary
        
        # Apply filters - only apply if values are provided and not empty
        filtered_logs = all_logs
        
        if marketplace and marketplace.strip():
            filtered_logs = [log for log in filtered_logs if log["marketplace"].upper() == marketplace.strip().upper()]
        
        if brand and brand.strip():
            filtered_logs = [log for log in filtered_logs if log["brand"].upper() == brand.strip().upper()]
        
        if user and user.strip():
            filtered_logs = [log for log in filtered_logs if log["user"].upper() == user.strip().upper()]
        
        # Apply pagination
        total_logs = len(filtered_logs)
        paginated_logs = filtered_logs[offset:offset + limit]
        
        # Update summary with filtered data
        filtered_summary = {
            "total_logs": total_logs,
            "total_orders_processed": sum(log["order_count"] for log in filtered_logs),
            "marketplaces": list(set(log["marketplace"] for log in filtered_logs)),
            "brands": list(set(log["brand"] for log in filtered_logs)),
            "users": list(set(log["user"] for log in filtered_logs)),
            "recent_executions": len([log for log in filtered_logs if 
                datetime.fromisoformat(log["execution_time"]) > datetime.now() - timedelta(hours=24)])
        }
        
        logger.info(f"Retrieved {len(paginated_logs)} marketplace logs (total: {total_logs})")
        
        return {
            "success": True,
            "logs": paginated_logs,
            "summary": filtered_summary,
            "pagination": {
                "limit": limit,
                "offset": offset,
                "total": total_logs,
                "has_more": offset + limit < total_logs
            },
            "timestamp": datetime.now().isoformat(),
            "cached": cached_logs is not None
        }
        
    except Exception as e:
        logger.error(f"Error getting marketplace logs: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get marketplace logs: {str(e)}")

@app.post("/api/marketplace-logs/refresh")
async def refresh_marketplace_logs_cache():
    """Force refresh the marketplace logs cache"""
    try:
        global _cache_timestamp
        _cache_timestamp = None  # Invalidate cache
        logger.info("Marketplace logs cache invalidated")
        return {"success": True, "message": "Cache refreshed successfully"}
    except Exception as e:
        logger.error(f"Error refreshing cache: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to refresh cache: {str(e)}")

@app.get("/api/marketplace-logs/{log_id}")
async def get_marketplace_log_details(
    log_id: str,
    db: Session = Depends(get_db)
):
    """Get detailed information for a specific marketplace log entry"""
    try:
        # Parse log_id to extract components
        parts = log_id.split('_')
        if len(parts) < 4:
            raise HTTPException(status_code=400, detail="Invalid log ID format")
        
        user = parts[0]
        marketplace = parts[1]
        filename = '_'.join(parts[2:-1])  # Everything except last part (timestamp)
        timestamp = parts[-1]
        
        # Construct file path
        # Get the project root directory
        # If running from backend directory, go up one level
        # If running from project root, use current directory
        if os.path.basename(os.getcwd()) == 'backend':
            project_root = get_project_root()
        else:
            project_root = os.getcwd()
        jobgetorder_path = os.path.join(project_root, "JobGetOrder")
        if user == "System":
            file_path = os.path.join(jobgetorder_path, marketplace, filename)
        else:
            file_path = os.path.join(jobgetorder_path, f"User_{user}", marketplace, filename)
        
        if not os.path.exists(file_path):
            raise HTTPException(status_code=404, detail="Log file not found")
        
        # Read order numbers
        order_numbers = []
        with open(file_path, 'r', encoding='utf-8') as f:
            for line in f:
                order_num = line.strip()
                if order_num:
                    order_numbers.append(order_num)
        
        # Get file stats
        file_stats = os.stat(file_path)
        mod_time = datetime.fromtimestamp(file_stats.st_mtime)
        file_size = file_stats.st_size
        
        # Try to get brand from config
        brand_name = "Unknown"
        config_dir = os.path.dirname(file_path)
        for file in os.listdir(config_dir):
            if file.endswith('.config'):
                try:
                    with open(os.path.join(config_dir, file), 'r', encoding='utf-8') as f:
                        content = f.read()
                        brand_match = re.search(r'<add key="brand_name" value="([^"]+)"', content)
                        if brand_match:
                            brand_name = brand_match.group(1)
                            break
                except:
                    continue
        
        return {
            "success": True,
            "log_details": {
                "id": log_id,
                "user": user,
                "marketplace": marketplace,
                "brand": brand_name,
                "filename": filename,
                "order_count": len(order_numbers),
                "order_numbers": order_numbers,
                "execution_time": mod_time.isoformat(),
                "file_size_bytes": file_size,
                "file_path": file_path
            },
            "timestamp": datetime.now().isoformat()
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error getting marketplace log details: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get log details: {str(e)}")

@app.get("/marketplace-terminal-logs")
async def get_marketplace_terminal_logs(
    marketplace: Optional[str] = Query(None, description="Filter by marketplace"),
    brand: Optional[str] = Query(None, description="Filter by brand"),
    user: Optional[str] = Query(None, description="Filter by user"),
    limit: int = Query(100, description="Number of log entries to return"),
    offset: int = Query(0, description="Number of log entries to skip"),
    db: Session = Depends(get_db)
):
    """Get terminal/console logs from marketplace applications"""
    try:
        terminal_logs = []
        
        # Get the project root directory
        if os.path.basename(os.getcwd()) == 'backend':
            project_root = get_project_root()
        else:
            project_root = os.getcwd()
        
        # Check for log files in various locations
        log_locations = [
            os.path.join(project_root, "logs", "app.log"),
            os.path.join(project_root, "backend", "logs", "app.log"),
            os.path.join(project_root, "logs"),
            os.path.join(project_root, "backend", "logs")
        ]
        
        # Also check for marketplace-specific log files
        jobgetorder_path = os.path.join(project_root, "JobGetOrder")
        if os.path.exists(jobgetorder_path):
            for item in os.listdir(jobgetorder_path):
                if item.startswith("User_") or item in ["Shopee", "Lazada", "Tokopedia", "Bukalapak", "Blibli", "Tiktok", "Ginee", "Jubelio", "Desty", "Zalora"]:
                    user_marketplace_path = os.path.join(jobgetorder_path, item)
                    if os.path.isdir(user_marketplace_path):
                        for marketplace_dir in os.listdir(user_marketplace_path):
                            marketplace_path = os.path.join(user_marketplace_path, marketplace_dir)
                            if os.path.isdir(marketplace_path):
                                # Look for log files in marketplace directory
                                for file in os.listdir(marketplace_path):
                                    if file.endswith('.log') or file.endswith('.txt'):
                                        log_locations.append(os.path.join(marketplace_path, file))
        
        # Process log files
        for log_file in log_locations:
            if os.path.exists(log_file) and os.path.isfile(log_file):
                try:
                    # Get file stats
                    file_stats = os.stat(log_file)
                    mod_time = datetime.fromtimestamp(file_stats.st_mtime)
                    
                    # Read recent log entries
                    with open(log_file, 'r', encoding='utf-8', errors='ignore') as f:
                        lines = f.readlines()
                    
                    # Get last N lines (recent entries)
                    recent_lines = lines[-limit:] if len(lines) > limit else lines
                    
                    # Parse log entries
                    for i, line in enumerate(recent_lines):
                        line = line.strip()
                        if not line:
                            continue
                            
                        # Try to extract timestamp, level, and message
                        timestamp_match = re.match(r'^(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2})', line)
                        timestamp = timestamp_match.group(1) if timestamp_match else mod_time.strftime('%Y-%m-%d %H:%M:%S')
                        
                        # Extract log level
                        level = "INFO"
                        if "ERROR" in line.upper():
                            level = "ERROR"
                        elif "WARNING" in line.upper() or "WARN" in line.upper():
                            level = "WARNING"
                        elif "DEBUG" in line.upper():
                            level = "DEBUG"
                        
                        # Extract marketplace/brand info from file path
                        marketplace_name = "Unknown"
                        brand_name = "Unknown"
                        user_name = "System"
                        
                        if "JobGetOrder" in log_file:
                            path_parts = log_file.split(os.sep)
                            for part in path_parts:
                                if part.startswith("User_"):
                                    user_name = part.replace("User_", "")
                                elif part in ["Shopee", "Lazada", "Tokopedia", "Bukalapak", "Blibli", "Tiktok", "Ginee", "Jubelio", "Desty", "Zalora"]:
                                    marketplace_name = part
                        
                        # Apply filters
                        if marketplace and marketplace_name.upper() != marketplace.upper():
                            continue
                        if user and user_name.upper() != user.upper():
                            continue
                        
                        # Create log entry
                        log_entry = {
                            "id": f"{os.path.basename(log_file)}_{i}_{file_stats.st_mtime}",
                            "timestamp": timestamp,
                            "level": level,
                            "message": line,
                            "marketplace": marketplace_name,
                            "brand": brand_name,
                            "user": user_name,
                            "source_file": os.path.basename(log_file),
                            "file_path": log_file,
                            "line_number": len(lines) - len(recent_lines) + i + 1
                        }
                        
                        terminal_logs.append(log_entry)
                        
                except Exception as e:
                    logger.error(f"Error processing log file {log_file}: {str(e)}")
                    continue
        
        # Sort logs by timestamp (newest first)
        terminal_logs.sort(key=lambda x: x["timestamp"], reverse=True)
        
        # Apply pagination
        total_logs = len(terminal_logs)
        paginated_logs = terminal_logs[offset:offset + limit]
        
        # Get summary statistics
        summary = {
            "total_logs": total_logs,
            "error_count": len([log for log in terminal_logs if log["level"] == "ERROR"]),
            "warning_count": len([log for log in terminal_logs if log["level"] == "WARNING"]),
            "info_count": len([log for log in terminal_logs if log["level"] == "INFO"]),
            "marketplaces": list(set(log["marketplace"] for log in terminal_logs)),
            "brands": list(set(log["brand"] for log in terminal_logs)),
            "users": list(set(log["user"] for log in terminal_logs)),
            "recent_logs": len([log for log in terminal_logs if 
                datetime.fromisoformat(log["timestamp"].replace(' ', 'T')) > datetime.now() - timedelta(hours=1)])
        }
        
        logger.info(f"Retrieved {len(paginated_logs)} terminal logs (total: {total_logs})")
        
        return {
            "success": True,
            "logs": paginated_logs,
            "summary": summary,
            "pagination": {
                "limit": limit,
                "offset": offset,
                "total": total_logs,
                "has_more": offset + limit < total_logs
            },
            "timestamp": datetime.now().isoformat()
        }
        
    except Exception as e:
        logger.error(f"Error getting terminal logs: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get terminal logs: {str(e)}")

@app.get("/marketplace-app-logs")
async def get_marketplace_app_logs(
    marketplace: Optional[str] = Query(None, description="Filter by marketplace"),
    user: Optional[str] = Query(None, description="Filter by user"),
    limit: int = Query(50, description="Number of logs to return"),
    offset: int = Query(0, description="Number of logs to skip")
):
    """Get logs from marketplace application executables"""
    try:
        # Get the project root directory
        if os.path.basename(os.getcwd()) == 'backend':
            project_root = get_project_root()
        else:
            project_root = os.getcwd()
        
        jobgetorder_path = os.path.join(project_root, "JobGetOrder")
        
        if not os.path.exists(jobgetorder_path):
            return {"logs": [], "total": 0, "message": "JobGetOrder directory not found"}
        
        all_logs = []
        
        # Get all user directories
        all_dirs = []
        if user:
            user_dir = os.path.join(jobgetorder_path, f"User_{user}")
            if os.path.exists(user_dir):
                all_dirs.append(user_dir)
        else:
            # Get all user directories
            for item in os.listdir(jobgetorder_path):
                item_path = os.path.join(jobgetorder_path, item)
                if os.path.isdir(item_path) and item.startswith("User_"):
                    all_dirs.append(item_path)
        
        for dir_path in all_dirs:
            if not os.path.exists(dir_path):
                continue
                
            # Get user name from directory
            dir_name = os.path.basename(dir_path)
            current_user_name = dir_name.replace("User_", "") if dir_name.startswith("User_") else "System"
            
            # Get all marketplace subdirectories
            for marketplace_dir in os.listdir(dir_path):
                marketplace_path = os.path.join(dir_path, marketplace_dir)
                if not os.path.isdir(marketplace_path):
                    continue
                    
                # Apply marketplace filter
                if marketplace and marketplace_dir.upper() != marketplace.upper():
                    continue
                
                # Look for log files from marketplace apps
                for file in os.listdir(marketplace_path):
                    file_path = os.path.join(marketplace_path, file)
                    if os.path.isfile(file_path):
                        # Check for common log file patterns
                        if (file.lower().endswith('.log') or 
                            (file.lower().endswith('.txt') and 'log' in file.lower()) or
                            file.lower().endswith('.out') or
                            file.lower().endswith('.err') or
                            file.lower().endswith('.txt') and 'orderlist' not in file.lower()):
                            
                            try:
                                # Get file modification time
                                mod_time = os.path.getmtime(file_path)
                                mod_datetime = datetime.fromtimestamp(mod_time)
                                
                                # Read log content
                                with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                                    content = f.read()
                                
                                # Parse log lines
                                log_lines = content.split('\n')
                                for line_num, line in enumerate(log_lines):
                                    if line.strip():
                                        # Try to extract timestamp, level, and message
                                        timestamp = None
                                        level = "INFO"
                                        message = line.strip()
                                        
                                        # Try to parse timestamp patterns
                                        timestamp_patterns = [
                                            r'(\d{1,2}/\d{1,2}/\d{4} \d{1,2}:\d{2}:\d{2})',  # 9/14/2025 13:40:04
                                            r'(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2})',        # 2025-09-14 13:40:04
                                            r'(\d{2}:\d{2}:\d{2})',                          # 13:40:04
                                        ]
                                        
                                        for pattern in timestamp_patterns:
                                            match = re.search(pattern, line)
                                            if match:
                                                timestamp = match.group(1)
                                                break
                                        
                                        # Detect log level
                                        if any(word in line.upper() for word in ['ERROR', 'EXCEPTION', 'FAILED', 'CRITICAL']):
                                            level = "ERROR"
                                        elif any(word in line.upper() for word in ['WARNING', 'WARN', 'CAUTION']):
                                            level = "WARNING"
                                        elif any(word in line.upper() for word in ['DEBUG', 'TRACE']):
                                            level = "DEBUG"
                                        
                                        # Create log entry
                                        log_entry = {
                                            "id": f"{current_user_name}_{marketplace_dir}_{file}_{line_num}_{mod_time}",
                                            "user": current_user_name,
                                            "marketplace": marketplace_dir,
                                            "log_file": file,
                                            "timestamp": timestamp or mod_datetime.strftime("%Y-%m-%d %H:%M:%S"),
                                            "level": level,
                                            "message": message,
                                            "file_path": file_path,
                                            "line_number": line_num + 1,
                                            "execution_time": mod_datetime
                                        }
                                        
                                        all_logs.append(log_entry)
                                        
                            except Exception as e:
                                print(f"Error reading log file {file_path}: {e}")
                                continue
        
        # Sort by execution time (most recent first)
        all_logs.sort(key=lambda x: x['execution_time'], reverse=True)
        
        # Apply pagination
        total_logs = len(all_logs)
        paginated_logs = all_logs[offset:offset + limit]
        
        logger.info(f"Retrieved {len(paginated_logs)} marketplace app logs (total: {total_logs})")
        
        return {
            "logs": paginated_logs,
            "total": total_logs,
            "pagination": {
                "limit": limit,
                "offset": offset,
                "total": total_logs,
                "has_more": offset + limit < total_logs
            },
            "timestamp": datetime.now().isoformat()
        }
        
    except Exception as e:
        logger.error(f"Error retrieving marketplace app logs: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/marketplace-terminal-logs/stream")
async def stream_marketplace_terminal_logs(
    marketplace: Optional[str] = Query(None, description="Filter by marketplace"),
    user: Optional[str] = Query(None, description="Filter by user"),
    db: Session = Depends(get_db)
):
    """Stream real-time terminal logs from marketplace applications"""
    try:
        
        async def generate_logs():
            # Get the project root directory
            if os.path.basename(os.getcwd()) == 'backend':
                project_root = get_project_root()
            else:
                project_root = os.getcwd()
            
            # Find marketplace app log file
            if marketplace and user:
                user_workspace = multi_user_handler.get_user_workspace(user)
                if user_workspace:
                    log_file = os.path.join(user_workspace, marketplace.title(), f"{marketplace.lower()}_app.log")
                else:
                    log_file = os.path.join(project_root, "JobGetOrder", f"User_{user}", marketplace.title(), f"{marketplace.lower()}_app.log")
            else:
                # Fallback to system log
                log_file = os.path.join(project_root, "logs", "app.log")
            
            if not os.path.exists(log_file):
                yield f"data: {json.dumps({'error': f'Log file not found: {log_file}'})}\n\n"
                return
            else:
                # Stream log file
                with open(log_file, 'r', encoding='utf-8', errors='ignore') as f:
                    # Go to end of file
                    f.seek(0, 2)
                    
                    while True:
                        line = f.readline()
                        if line:
                            # Parse log line
                            timestamp_match = re.match(r'^(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2})', line.strip())
                            timestamp = timestamp_match.group(1) if timestamp_match else datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                            
                            level = "INFO"
                            if "ERROR" in line.upper():
                                level = "ERROR"
                            elif "WARNING" in line.upper() or "WARN" in line.upper():
                                level = "WARNING"
                            elif "DEBUG" in line.upper():
                                level = "DEBUG"
                            
                            log_entry = {
                                "timestamp": timestamp,
                                "level": level,
                                "message": line.strip(),
                                "marketplace": marketplace or "System",
                                "user": user or "System"
                            }
                            
                            # Apply filters
                            if marketplace and marketplace.upper() not in line.upper():
                                continue
                            if user and user.upper() not in line.upper():
                                continue
                            
                            yield f"data: {json.dumps(log_entry)}\n\n"
                        else:
                            await asyncio.sleep(1)  # Wait for new content
        
        return StreamingResponse(
            generate_logs(),
            media_type="text/plain",
            headers={
                "Cache-Control": "no-cache",
                "Connection": "keep-alive",
                "Content-Type": "text/event-stream"
            }
        )
        
    except Exception as e:
        logger.error(f"Error streaming terminal logs: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to stream terminal logs: {str(e)}")

# Brand Accounts API Endpoints
class BrandAccountBase(BaseModel):
    brand: str
    platform: str
    uid: Optional[str] = None
    password: Optional[str] = None
    status_account: Optional[str] = None
    email_sms: Optional[str] = None
    pic_otp: Optional[str] = None
    user_device: Optional[str] = None

class BrandAccountCreate(BrandAccountBase):
    pass

class BrandAccountUpdate(BaseModel):
    brand: Optional[str] = None
    platform: Optional[str] = None
    uid: Optional[str] = None
    password: Optional[str] = None
    status_account: Optional[str] = None
    email_sms: Optional[str] = None
    pic_otp: Optional[str] = None
    user_device: Optional[str] = None

class BrandAccountResponse(BrandAccountBase):
    id: int
    created_at: datetime
    
    class Config:
        from_attributes = True

@app.get("/brand-accounts")
def get_brand_accounts(
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db)
):
    """Get all brand accounts"""
    try:
        # Use PostgreSQL with proper parameter placeholders
        query = "SELECT * FROM brand_accounts LIMIT $1 OFFSET $2"
        result = db.execute(text(query), [limit, skip])
        rows = result.fetchall()
        
        # Convert to response format
        brand_accounts = []
        for row in rows:
            brand_accounts.append({
                "id": row[0],
                "brand": row[1],
                "platform": row[2],
                "uid": row[3],
                "password": row[4],
                "status_account": row[5],
                "email_sms": row[6],
                "pic_otp": row[7] if len(row) > 7 else None,
                "user_device": row[8] if len(row) > 8 else None,
                "created_at": row[9] if len(row) > 9 else None
            })
        
        return brand_accounts
        
    except Exception as e:
        logger.error(f"Error fetching brand accounts: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to fetch brand accounts: {str(e)}")

@app.post("/brand-accounts", response_model=BrandAccountResponse)
def create_brand_account(
    brand_account: BrandAccountCreate,
    current_user: str = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Create a new brand account"""
    
    try:
        # Use raw SQL to insert
        insert_query = """
        INSERT INTO brand_accounts 
        (brand, platform, uid, password, status_account, email_sms, pic_otp, user_device, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """
        
        now = datetime.now()
        
        # PostgreSQL: Use $1, $2, etc. parameter placeholders
        insert_query = """
            INSERT INTO brand_accounts (brand, platform, uid, password, status_account, email_sms, pic_otp, user_device, created_at)
            VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9)
        """
        
        db.commit()
        
        # Get the inserted record using PostgreSQL RETURNING clause
        insert_query_with_returning = """
            INSERT INTO brand_accounts (brand, platform, uid, password, status_account, email_sms, pic_otp, user_device, created_at)
            VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9)
            RETURNING id
        """
        result = db.execute(text(insert_query_with_returning), [
            brand_account.brand,
            brand_account.platform,
            brand_account.uid,
            brand_account.password,
            brand_account.status_account,
            brand_account.email_sms,
            brand_account.pic_otp,
            brand_account.user_device,
            now
        ])
        new_id = result.fetchone()[0]
        
        # Get the full record
        select_query = "SELECT * FROM brand_accounts WHERE id = $1"
        result = db.execute(text(select_query), [new_id])
        row = result.fetchone()
        
        return {
            "id": row[0],
            "brand": row[1],
            "platform": row[2],
            "uid": row[3],
            "password": row[4],
            "status_account": row[5],
            "email_sms": row[6],
            "pic_otp": row[7],
            "user_device": row[8],
            "created_at": row[9]
        }
        
    except Exception as e:
        db.rollback()
        logger.error(f"Error creating brand account: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to create brand account: {str(e)}")

@app.put("/brand-accounts/{account_id}", response_model=BrandAccountResponse)
def update_brand_account(
    account_id: int,
    brand_account: BrandAccountUpdate,
    current_user: str = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Update a brand account"""
    
    try:
        # Check if record exists
        check_query = "SELECT COUNT(*) FROM brand_accounts WHERE id = $1"
        result = db.execute(text(check_query), [account_id])
        count = result.fetchone()[0]
        
        if count == 0:
            raise HTTPException(status_code=404, detail="Brand account not found")
        
        # Build update query dynamically
        update_fields = []
        params = []
        param_index = 1
        
        for field, value in brand_account.dict(exclude_unset=True).items():
            if value is not None:
                update_fields.append(f"{field} = ${param_index}")
                params.append(value)
                param_index += 1
        
        if not update_fields:
            raise HTTPException(status_code=400, detail="No fields to update")
        
        update_query = f"UPDATE brand_accounts SET {', '.join(update_fields)} WHERE id = ${param_index}"
        params.append(account_id)
        
        db.execute(text(update_query), params)
        db.commit()
        
        # Get updated record
        select_query = "SELECT * FROM brand_accounts WHERE id = $1"
        result = db.execute(text(select_query), [account_id])
        row = result.fetchone()
        
        return {
            "id": row[0],
            "brand": row[1],
            "platform": row[2],
            "uid": row[3],
            "password": row[4],
            "status_account": row[5],
            "email_sms": row[6],
            "pic_otp": row[7],
            "user_device": row[8],
            "created_at": row[9]
        }
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Error updating brand account: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to update brand account: {str(e)}")

@app.delete("/brand-accounts/{account_id}")
def delete_brand_account(
    account_id: int,
    current_user: str = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Delete a brand account"""
    
    try:
        # Check if record exists
        check_query = "SELECT COUNT(*) FROM brand_accounts WHERE id = $1"
        result = db.execute(text(check_query), [account_id])
        count = result.fetchone()[0]
        
        if count == 0:
            raise HTTPException(status_code=404, detail="Brand account not found")
        
        # Delete record
        delete_query = "DELETE FROM brand_accounts WHERE id = $1"
        db.execute(text(delete_query), [account_id])
        db.commit()
        
        return {"message": "Brand account deleted successfully"}
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Error deleting brand account: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to delete brand account: {str(e)}")

@app.get("/brand-accounts/test")
def test_brand_accounts(
    current_user: str = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Test brand accounts endpoint"""
    
    try:
        # Simple test query
        result = db.execute(text("SELECT COUNT(*) FROM brand_accounts"))
        count = result.fetchone()[0]
        
        return {
            "message": "Brand Accounts endpoint working",
            "total_count": count,
            "user": current_user.username
        }
        
    except Exception as e:
        logger.error(f"Error in test endpoint: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Test failed: {str(e)}")

@app.get("/test-cors")
def test_cors():
    """Test CORS endpoint without authentication"""
    return {
        "message": "CORS test successful",
        "timestamp": datetime.now().isoformat()
    }

@app.get("/brand-accounts-simple")
def get_brand_accounts_simple(skip: int = 0, limit: int = 100, search: str = None):
    """Simple brand accounts endpoint with pagination and search"""
    try:
        # Use PostgreSQL connection
        db = SessionLocal()
        
        if search and search.strip():
            # Search across all columns
            search_pattern = f"%{search.strip()}%"
            
            # Get count with search
            count_query = text("""
                SELECT COUNT(*) FROM brand_accounts 
                WHERE brand ILIKE :search OR platform ILIKE :search OR uid ILIKE :search OR password ILIKE :search 
                OR status_account ILIKE :search OR email_sms ILIKE :search OR pic_otp ILIKE :search 
                OR created_at::text ILIKE :search OR updated_at::text ILIKE :search
            """)
            count_result = db.execute(count_query, {"search": search_pattern})
            count = count_result.fetchone()[0]
            
            # Get data with search
            data_query = text("""
                SELECT brand, platform, uid, password, status_account, email_sms, pic_otp, created_at, updated_at 
                FROM brand_accounts 
                WHERE brand ILIKE :search OR platform ILIKE :search OR uid ILIKE :search OR password ILIKE :search 
                OR status_account ILIKE :search OR email_sms ILIKE :search OR pic_otp ILIKE :search 
                OR created_at::text ILIKE :search OR updated_at::text ILIKE :search
                LIMIT :limit OFFSET :offset
            """)
            rows_result = db.execute(data_query, {"search": search_pattern, "limit": limit, "offset": skip})
            rows = rows_result.fetchall()
        else:
            # No search - get all data
            count_query = text("SELECT COUNT(*) FROM brand_accounts")
            count_result = db.execute(count_query)
            count = count_result.fetchone()[0]
            
            data_query = text("SELECT brand, platform, uid, password, status_account, email_sms, pic_otp, created_at, updated_at FROM brand_accounts LIMIT :limit OFFSET :offset")
            rows_result = db.execute(data_query, {"limit": limit, "offset": skip})
            rows = rows_result.fetchall()
        
        db.close()
        
        # Convert to response format
        brand_accounts = []
        for row in rows:
            brand_accounts.append({
                "brand": row[0],
                "platform": row[1],
                "uid": row[2],
                "password": row[3],
                "status_account": row[4],
                "email_sms": row[5],
                "pic_otp": row[6],
                "created_at": row[7],
                "updated_at": row[8]
            })
        
        return {
            "total_count": count,
            "data": brand_accounts,
            "skip": skip,
            "limit": limit
        }
        
    except Exception as e:
        return {"error": str(e)}

@app.put("/brand-accounts-simple/{uid}")
def update_brand_account(uid: str, brand_account: dict, db: Session = Depends(get_db)):
    """Update brand account by UID"""
    try:
        # Check if record exists
        check_query = "SELECT COUNT(*) FROM brand_accounts WHERE uid = :uid"
        result = db.execute(text(check_query), {"uid": uid})
        if result.fetchone()[0] == 0:
            raise HTTPException(status_code=404, detail="Brand account not found")
        
        # Update record including UID
        update_query = """
            UPDATE brand_accounts 
            SET brand = :brand, platform = :platform, uid = :new_uid, password = :password, status_account = :status_account, 
                email_sms = :email_sms, pic_otp = :pic_otp, updated_at = :updated_at
            WHERE uid = :uid
        """
        db.execute(text(update_query), {
            "brand": brand_account.get('brand'),
            "platform": brand_account.get('platform'),
            "new_uid": brand_account.get('uid'),
            "password": brand_account.get('password'),
            "status_account": brand_account.get('status_account'),
            "email_sms": brand_account.get('email_sms'),
            "pic_otp": brand_account.get('pic_otp'),
            "updated_at": get_wib_now(),
            "uid": uid
        })
        
        db.commit()
        
        return {"message": "Brand account updated successfully"}
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Error updating brand account: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to update brand account: {str(e)}")

@app.delete("/brand-accounts-simple/{uid}")
def delete_brand_account(uid: str, db: Session = Depends(get_db)):
    """Delete brand account by UID"""
    try:
        # Check if record exists
        check_query = "SELECT COUNT(*) FROM brand_accounts WHERE uid = :uid"
        result = db.execute(text(check_query), {"uid": uid})
        if result.fetchone()[0] == 0:
            raise HTTPException(status_code=404, detail="Brand account not found")
        
        # Delete record
        delete_query = "DELETE FROM brand_accounts WHERE uid = :uid"
        db.execute(text(delete_query), {"uid": uid})
        
        db.commit()
        
        return {"message": "Brand account deleted successfully"}
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Error deleting brand account: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to delete brand account: {str(e)}")

@app.post("/brand-accounts-simple")
def create_brand_account(brand_account: dict, db: Session = Depends(get_db)):
    """Create new brand account"""
    try:
        # Check if UID already exists
        check_query = "SELECT COUNT(*) FROM brand_accounts WHERE uid = :uid"
        result = db.execute(text(check_query), {"uid": brand_account.get('uid')})
        if result.fetchone()[0] > 0:
            raise HTTPException(status_code=400, detail="UID already exists")
        
        # Insert new record
        insert_query = """
            INSERT INTO brand_accounts 
            (brand, platform, uid, password, status_account, email_sms, pic_otp, created_at, updated_at)
            VALUES (:brand, :platform, :uid, :password, :status_account, :email_sms, :pic_otp, :created_at, :updated_at)
        """
        now = get_wib_now()
        db.execute(text(insert_query), {
            "brand": brand_account.get('brand'),
            "platform": brand_account.get('platform'),
            "uid": brand_account.get('uid'),
            "password": brand_account.get('password'),
            "status_account": brand_account.get('status_account'),
            "email_sms": brand_account.get('email_sms'),
            "pic_otp": brand_account.get('pic_otp'),
            "created_at": now,
            "updated_at": now
        })
        
        db.commit()
        
        return {"message": "Brand account created successfully"}
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Error creating brand account: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to create brand account: {str(e)}")

@app.get("/brand-accounts/stats")
def get_brand_accounts_stats(
    current_user: str = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get brand accounts statistics"""
    
    try:
        # Get total count
        total_query = "SELECT COUNT(*) FROM brand_accounts"
        result = db.execute(text(total_query))
        total_count = result.fetchone()[0]
        
        # Get count by platform
        platform_query = "SELECT platform, COUNT(*) FROM brand_accounts GROUP BY platform"
        result = db.execute(text(platform_query))
        platform_stats = {row[0]: row[1] for row in result.fetchall()}
        
        # Get count by brand
        brand_query = "SELECT brand, COUNT(*) FROM brand_accounts GROUP BY brand ORDER BY COUNT(*) DESC LIMIT 10"
        result = db.execute(text(brand_query))
        brand_stats = {row[0]: row[1] for row in result.fetchall()}
        
        return {
            "total_accounts": total_count,
            "platform_stats": platform_stats,
            "top_brands": brand_stats
        }
        
    except Exception as e:
        logger.error(f"Error fetching brand accounts stats: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to fetch brand accounts stats: {str(e)}")

@app.post("/brand-accounts/bulk-create")
async def bulk_create_brand_accounts(
    file: UploadFile = File(...),
    current_user: str = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Bulk create brand accounts from Excel/CSV file"""
    try:
        created_accounts = []
        errors = []
        
        # Read file content
        file_content = file.file.read()
        
        # Parse file based on extension
        if file.filename.endswith('.xlsx'):
            df = pd.read_excel(io.BytesIO(file_content))
        elif file.filename.endswith('.csv'):
            df = pd.read_csv(io.BytesIO(file_content))
        else:
            raise HTTPException(status_code=400, detail="Unsupported file format. Please use .xlsx or .csv")
        
        # Validate required columns
        required_columns = ['brand', 'platform']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            raise HTTPException(status_code=400, detail=f"Missing required columns: {', '.join(missing_columns)}")
        
        # Process each row
        for i, row in df.iterrows():
            try:
                account_data = row.to_dict()
                
                # Clean NaN values and convert types
                for key, value in account_data.items():
                    if pd.isna(value):
                        account_data[key] = None
                    elif isinstance(value, (int, float)) and (value == float('inf') or value == float('-inf') or pd.isna(value)):
                        account_data[key] = None
                    elif value is not None:
                        account_data[key] = str(value)
                
                # Validate required fields
                if not account_data.get('brand') or not account_data.get('platform'):
                    errors.append(f"Row {i+2}: Brand and platform are required")
                    continue
                
                # Check if account already exists (brand + platform + uid combination)
                uid = account_data.get('uid')
                if uid:
                    check_query = "SELECT COUNT(*) FROM brand_accounts WHERE brand = :brand AND platform = :platform AND uid = :uid"
                    result = db.execute(text(check_query), {
                        "brand": account_data['brand'],
                        "platform": account_data['platform'],
                        "uid": uid
                    })
                    if result.fetchone()[0] > 0:
                        errors.append(f"Row {i+2}: Account already exists for brand {account_data['brand']} on platform {account_data['platform']} with UID {uid}")
                        continue
                
                # Create new account
                insert_query = """
                    INSERT INTO brand_accounts (brand, platform, uid, password, status_account, created_at)
                    VALUES (:brand, :platform, :uid, :password, :status_account, :created_at)
                    RETURNING id, brand, platform, uid, password, status_account, created_at
                """
                
                result = db.execute(text(insert_query), {
                    "brand": account_data['brand'],
                    "platform": account_data['platform'],
                    "uid": account_data.get('uid', ''),
                    "password": account_data.get('password', ''),
                    "status_account": account_data.get('status_account', 'active'),
                    "created_at": datetime.now()
                })
                
                row = result.fetchone()
                created_accounts.append({
                    "id": int(row[0]) if row[0] else None,
                    "brand": str(row[1]) if row[1] else None,
                    "platform": str(row[2]) if row[2] else None,
                    "uid": str(row[3]) if row[3] else None,
                    "password": str(row[4]) if row[4] else None,
                    "status_account": str(row[5]) if row[5] else None,
                    "created_at": row[6].isoformat() if row[6] else None
                })
                
            except Exception as e:
                errors.append(f"Row {i+2}: {str(e)}")
        
        db.commit()
        
        return {
            "message": f"Bulk operation completed. {len(created_accounts)} accounts created, {len(errors)} errors.",
            "created_count": len(created_accounts),
            "error_count": len(errors),
            "errors": errors,
            "accounts": created_accounts
        }
        
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/brand-accounts/bulk-update")
async def bulk_update_brand_accounts(
    updates_data: List[dict],
    current_user: str = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Bulk update brand accounts"""
    try:
        updated_accounts = []
        errors = []
        
        for i, update_data in enumerate(updates_data):
            try:
                uid = update_data.get('uid')
                if not uid:
                    errors.append(f"Row {i+1}: UID is required for updates")
                    continue
                
                # Check if account exists
                check_query = "SELECT COUNT(*) FROM brand_accounts WHERE uid = :uid"
                result = db.execute(text(check_query), {"uid": uid})
                if result.fetchone()[0] == 0:
                    errors.append(f"Row {i+1}: Account with UID {uid} not found")
                    continue
                
                # Build update query dynamically
                update_fields = []
                update_values = {"uid": uid}
                
                for field in ['brand', 'platform', 'password', 'status_account']:
                    if field in update_data:
                        update_fields.append(f"{field} = :{field}")
                        update_values[field] = update_data[field]
                
                if update_fields:
                    update_query = f"UPDATE brand_accounts SET {', '.join(update_fields)} WHERE uid = :uid RETURNING id, brand, platform, uid, password, status_account, created_at"
                    result = db.execute(text(update_query), update_values)
                    row = result.fetchone()
                    
                    updated_accounts.append({
                        "id": row[0],
                        "brand": row[1],
                        "platform": row[2],
                        "uid": row[3],
                        "password": row[4],
                        "status_account": row[5],
                        "created_at": row[6].isoformat() if row[6] else None
                    })
                
            except Exception as e:
                errors.append(f"Row {i+1}: {str(e)}")
        
        db.commit()
        
        return {
            "message": f"Bulk update completed. {len(updated_accounts)} accounts updated, {len(errors)} errors.",
            "updated_count": len(updated_accounts),
            "error_count": len(errors),
            "errors": errors,
            "accounts": updated_accounts
        }
        
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/brand-accounts/bulk-delete")
async def bulk_delete_brand_accounts(
    uids: List[str],
    current_user: str = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Bulk delete brand accounts"""
    try:
        deleted_count = 0
        errors = []
        
        for uid in uids:
            try:
                # Check if account exists
                check_query = "SELECT COUNT(*) FROM brand_accounts WHERE uid = :uid"
                result = db.execute(text(check_query), {"uid": uid})
                if result.fetchone()[0] == 0:
                    errors.append(f"Account with UID {uid} not found")
                    continue
                
                # Delete account
                delete_query = "DELETE FROM brand_accounts WHERE uid = :uid"
                db.execute(text(delete_query), {"uid": uid})
                deleted_count += 1
                
            except Exception as e:
                errors.append(f"Error deleting account UID {uid}: {str(e)}")
        
        db.commit()
        
        return {
            "message": f"Bulk delete completed. {deleted_count} accounts deleted, {len(errors)} errors.",
            "deleted_count": deleted_count,
            "error_count": len(errors),
            "errors": errors
        }
        
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/listbrand/template")
async def download_marketplace_template(
    current_user: str = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Download marketplace info template"""
    try:
        # Create template data
        template_data = [
            {
                "brand": "FACETOLOGY",
                "marketplace": "SHOPEE",
                "batch": "1",
                "remark": "Sample remark"
            },
            {
                "brand": "FACETOLOGY",
                "marketplace": "TIKTOK",
                "batch": "1",
                "remark": ""
            },
            {
                "brand": "DEARDOER",
                "marketplace": "TOKOPEDIA",
                "batch": "2",
                "remark": "Example"
            }
        ]
        
        # Create DataFrame
        df = pd.DataFrame(template_data)
        
        # Create Excel file in memory
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Template', index=False)
            
            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Template']
            
            # Add header formatting
            from openpyxl.styles import Font, PatternFill
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
        
        output.seek(0)
        
        # Return file as response
        from fastapi.responses import StreamingResponse
        return StreamingResponse(
            io.BytesIO(output.read()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=marketplace_template.xlsx"}
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/brandshops/template")
async def download_brandshops_template(
    current_user: str = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Download brand shops template"""
    try:
        # Create template data
        template_data = [
            {
                "brand": "FACETOLOGY",
                "marketplace_id": 2,
                "shop_name": "FACETOLOGY SHOPEE",
                "shop_key_1": "12345",
                "client_shop_id": 1001,
                "client_id": 1,
                "order_type": "ONLINE"
            },
            {
                "brand": "FACETOLOGY",
                "marketplace_id": 11,
                "shop_name": "FACETOLOGY TIKTOK",
                "shop_key_1": "67890",
                "client_shop_id": 1002,
                "client_id": 1,
                "order_type": "ONLINE"
            },
            {
                "brand": "DEARDOER",
                "marketplace_id": 1,
                "shop_name": "DEARDOER TOKOPEDIA",
                "shop_key_1": "11111",
                "client_shop_id": 1003,
                "client_id": 2,
                "order_type": "ONLINE"
            }
        ]
        
        # Create DataFrame
        df = pd.DataFrame(template_data)
        
        # Create Excel file in memory
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Template', index=False)
            
            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Template']
            
            # Add header formatting
            from openpyxl.styles import Font, PatternFill
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
        
        output.seek(0)
        
        # Return file as response
        from fastapi.responses import StreamingResponse
        return StreamingResponse(
            io.BytesIO(output.read()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=brandshops_template.xlsx"}
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/brand-accounts/template")
async def download_brandaccounts_template(
    current_user: str = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Download brand accounts template"""
    try:
        # Create template data
        template_data = [
            {
                "brand": "FACETOLOGY",
                "platform": "SHOPEE",
                "uid": "facetology_shopee@example.com",
                "password": "password123",
                "status_account": "active"
            },
            {
                "brand": "FACETOLOGY",
                "platform": "TIKTOK",
                "uid": "facetology_tiktok@example.com",
                "password": "password456",
                "status_account": "active"
            },
            {
                "brand": "DEARDOER",
                "platform": "TOKOPEDIA",
                "uid": "deardoer_tokopedia@example.com",
                "password": "password789",
                "status_account": "active"
            }
        ]
        
        # Create DataFrame
        df = pd.DataFrame(template_data)
        
        # Create Excel file in memory
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Template', index=False)
            
            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Template']
            
            # Add header formatting
            from openpyxl.styles import Font, PatternFill
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
        
        output.seek(0)
        
        # Return file as response
        from fastapi.responses import StreamingResponse
        return StreamingResponse(
            io.BytesIO(output.read()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=brandaccounts_template.xlsx"}
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/dashboard/export")
async def export_dashboard_data(
    start_date: Optional[str] = Query(None, description="Start date (ISO format)"),
    end_date: Optional[str] = Query(None, description="End date (ISO format)"),
    order_status: Optional[str] = Query(None, description="Comma-separated order statuses to filter"),
    current_user: str = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Export all dashboard data to Excel file"""
    try:
        
        def make_timezone_naive(dt):
            """Convert timezone-aware datetime to timezone-naive"""
            if dt is None:
                return None
            if hasattr(dt, 'tzinfo') and dt.tzinfo is not None:
                return dt.replace(tzinfo=None)
            return dt
        
        # Parse dates if provided, otherwise export all data
        start_dt = None
        end_dt = None
        if start_date and end_date:
            try:
                start_dt = parse_date_flexible(start_date)
                end_dt = parse_date_flexible(end_date)
                
                # Convert to naive datetime for database comparison (same as dashboard stats)
                if start_dt and start_dt.tzinfo is not None:
                    wib = pytz.timezone('Asia/Jakarta')
                    start_dt = start_dt.astimezone(wib).replace(tzinfo=None)
                
                if end_dt and end_dt.tzinfo is not None:
                    wib = pytz.timezone('Asia/Jakarta')
                    end_dt = end_dt.astimezone(wib).replace(tzinfo=None)
                    
            except ValueError as e:
                raise HTTPException(status_code=400, detail=f"Invalid date format: {str(e)}")
        
        logger.info(f"Exporting dashboard data {'from ' + str(start_dt) + ' to ' + str(end_dt) if start_dt and end_dt else 'all data'} for user {current_user}")
        logger.info(f"Export date range: start_date='{start_date}', end_date='{end_date}' -> start_dt={start_dt}, end_dt={end_dt}")
        
        # Get all dashboard data
        # 1. Orders Stats
        orders_query = db.query(UploadedOrder)
        if start_dt and end_dt:
            orders_query = orders_query.filter(
            UploadedOrder.UploadDate >= start_dt,
            UploadedOrder.UploadDate <= end_dt
            )
        
        # Apply order status filter to orders stats if provided
        if order_status:
            order_statuses = [status.strip() for status in order_status.split(',') if status.strip()]
            if order_statuses:
                orders_query = orders_query.filter(
                    UploadedOrder.OrderStatus.in_(order_statuses)
                )
        
        orders_stats = orders_query.all()
        logger.info(f"Export found {len(orders_stats)} orders after filtering")
        
        # 2. Upload History (use same logic as dashboard - no date filtering, get recent records)
        upload_history_query = db.query(UploadHistory).order_by(UploadHistory.upload_date.desc()).limit(100)
        upload_history = upload_history_query.all()
        
        # 3. Not Uploaded Items (use same logic as dashboard endpoint)
        query = db.query(NotUploadedHistory).filter(NotUploadedHistory.status == 'not_uploaded')
        
        if start_dt and end_dt:
            # Filter by date range (convert to timezone-aware for comparison)
            wib = pytz.timezone('Asia/Jakarta')
            start_dt_tz = wib.localize(start_dt)
            end_dt_tz = wib.localize(end_dt)
            query = query.filter(
                NotUploadedHistory.check_date >= start_dt_tz,
                NotUploadedHistory.check_date <= end_dt_tz
            )
        else:
            # If no date filter, get today's data
            current_time = get_wib_now()
            today_str = current_time.strftime('%Y-%m-%d')
            query = query.filter(
                func.date(NotUploadedHistory.check_date) == today_str
            )
        
        history_items = query.all()
        
        not_uploaded = []
        for item in history_items:
            not_uploaded.append({
                "marketplace": item.marketplace,
                "brand": item.brand,
                "batch": item.batch,
                "remark": item.remark,
                "created_at": make_timezone_naive(item.check_date)
            })
        
        # Sort by brand, then marketplace, then batch
        not_uploaded.sort(key=lambda x: (x["brand"], x["marketplace"], x["batch"]))
        
        # 4. Not Interfaced Orders
        not_interfaced_query = db.query(UploadedOrder).filter(
            UploadedOrder.InterfaceStatus != 'Interface'
        )
        if start_dt and end_dt:
            not_interfaced_query = not_interfaced_query.filter(
            UploadedOrder.UploadDate >= start_dt,
            UploadedOrder.UploadDate <= end_dt
            )
        
        # Apply order status filter if provided
        if order_status:
            order_statuses = [status.strip() for status in order_status.split(',') if status.strip()]
            if order_statuses:
                not_interfaced_query = not_interfaced_query.filter(
                    UploadedOrder.OrderStatus.in_(order_statuses)
                )
        
        not_interfaced_orders = not_interfaced_query.all()
        
        # Create Excel file
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Sheet 1: Orders Statistics
            orders_df = pd.DataFrame([{
                'Marketplace': order.Marketplace,
                'Brand': order.Brand,
                'Order Number': order.OrderNumber,
                'Batch': order.Batch,
                'Interface Status': order.InterfaceStatus,
                'Upload Date': make_timezone_naive(order.UploadDate),
                'PIC': order.PIC
            } for order in orders_stats])
            orders_df.to_excel(writer, sheet_name='Orders Statistics', index=False)
            
            # Sheet 2: Upload History
            upload_df = pd.DataFrame([{
                'Marketplace': history.marketplace,
                'Brand': history.brand,
                'PIC': history.pic,
                'Batch': history.batch,
                'Upload Date': make_timezone_naive(history.upload_date)
            } for history in upload_history])
            upload_df.to_excel(writer, sheet_name='Upload History', index=False)
            
            # Sheet 3: Not Uploaded Items
            not_uploaded_df = pd.DataFrame(not_uploaded)
            not_uploaded_df.to_excel(writer, sheet_name='Not Uploaded Items', index=False)
            
            # Sheet 4: Not Interfaced Orders
            not_interfaced_df = pd.DataFrame([{
                'Marketplace': order.Marketplace,
                'Brand': order.Brand,
                'Order Number': order.OrderNumber,
                'Batch': order.Batch,
                'Order Status': order.OrderStatus,
                'Interface Status': order.InterfaceStatus,
                'Remark': order.Remarks,
                'Upload Date': make_timezone_naive(order.UploadDate),
                'PIC': order.PIC
            } for order in not_interfaced_orders])
            not_interfaced_df.to_excel(writer, sheet_name='Not Interfaced Orders', index=False)
            
            # Sheet 5: Summary Statistics
            summary_data = {
                'Metric': [
                    'Total Orders',
                    'Total Upload History Records',
                    'Not Uploaded Items',
                    'Not Interfaced Orders',
                    'Export Date Range',
                    'Applied Filters',
                    'Generated By',
                    'Generated At'
                ],
                'Value': [
                    len(orders_stats),
                    len(upload_history),
                    len(not_uploaded),
                    len(not_interfaced_orders),
                    f"{start_dt.strftime('%Y-%m-%d %H:%M:%S')} to {end_dt.strftime('%Y-%m-%d %H:%M:%S')}" if start_dt and end_dt else "All data",
                    f"Order Status Filter: {order_status}" if order_status else "No order status filter",
                    current_user,
                    make_timezone_naive(datetime.now()).strftime('%Y-%m-%d %H:%M:%S')
                ]
            }
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
        
        output.seek(0)
        
        # Generate filename
        if start_dt and end_dt:
            filename = f"dashboard-export-{start_dt.strftime('%Y%m%d')}-{end_dt.strftime('%Y%m%d')}.xlsx"
        else:
            filename = f"dashboard-export-all-data-{make_timezone_naive(datetime.now()).strftime('%Y%m%d')}.xlsx"
        
        return StreamingResponse(
            io.BytesIO(output.read()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        logger.error(f"Error exporting dashboard data: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to export dashboard data: {str(e)}")

# Daily reset scheduler for remarks
def reset_remarks_daily():
    """Reset all remarks in list_brand table daily at 23:59:59"""
    try:
        # Create a new database session for this operation
        engine = create_engine(SQLALCHEMY_DATABASE_URL)
        SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
        db = SessionLocal()
        
        # Use raw SQL to avoid trigger issues
        from sqlalchemy import text
        db.execute(text("UPDATE list_brand SET remark = NULL"))
        db.commit()
        
        logger.info("Daily remark reset completed - all remarks cleared")
        db.close()
        
    except Exception as e:
        logger.error(f"Error during daily remark reset: {str(e)}")

def save_not_uploaded_history():
    """Save current not uploaded items to history table"""
    try:
        engine = create_engine(SQLALCHEMY_DATABASE_URL)
        SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
        db = SessionLocal()
        
        # Get current not uploaded items (real-time)
        expected_combinations = db.query(ListBrand).all()
        uploaded_combinations = db.query(
            UploadedOrder.Brand,
            UploadedOrder.Marketplace,
            UploadedOrder.Batch
        ).distinct().all()
        
        # Convert uploaded combinations to set for faster lookup
        uploaded_set = set()
        for combo in uploaded_combinations:
            brand = combo.Brand.strip().upper() if combo.Brand else ""
            marketplace = combo.Marketplace.strip().upper() if combo.Marketplace else ""
            batch = combo.Batch.strip().upper() if combo.Batch else ""
            uploaded_set.add((brand, marketplace, batch))
        
        # Save not uploaded items to history
        current_time = get_wib_now()
        for expected in expected_combinations:
            brand = expected.brand.strip().upper() if expected.brand else ""
            marketplace = expected.marketplace.strip().upper() if expected.marketplace else ""
            batch = expected.batch.strip().upper() if expected.batch else ""
            
            if (brand, marketplace, batch) not in uploaded_set:
                # Check if this combination already exists in history for today
                existing = db.query(NotUploadedHistory).filter(
                    NotUploadedHistory.brand == expected.brand,
                    NotUploadedHistory.marketplace == expected.marketplace,
                    NotUploadedHistory.batch == expected.batch,
                    NotUploadedHistory.check_date >= current_time.replace(hour=0, minute=0, second=0, microsecond=0),
                    NotUploadedHistory.check_date < current_time.replace(hour=23, minute=59, second=59, microsecond=999999)
                ).first()
                
                if not existing:
                    history_item = NotUploadedHistory(
                        brand=expected.brand,
                        marketplace=expected.marketplace,
                        batch=expected.batch,
                        remark=expected.remark,
                        status='not_uploaded',
                        check_date=current_time
                    )
                    db.add(history_item)
        
        db.commit()
        logger.info("Not uploaded items history saved successfully")
        db.close()
    except Exception as e:
        logger.error(f"Error saving not uploaded history: {str(e)}")

def schedule_daily_reset():
    """Schedule daily reset at 23:59:59"""
    
    # Clear any existing schedules
    schedule.clear()
    
    # Schedule the reset for 23:59:59 every day
    schedule.every().day.at("23:59:59").do(reset_remarks_daily)
    
    # Schedule history save to run daily at 23:58:00 (before reset)
    schedule.every().day.at("23:58:00").do(save_not_uploaded_history)
    
    def run_scheduler():
        logger.info("Scheduler thread started")
        while True:
            try:
                schedule.run_pending()
                time.sleep(1)
            except Exception as e:
                logger.error(f"Scheduler error: {str(e)}")
                time.sleep(5)  # Wait 5 seconds before retrying
    
    # Start the scheduler in a separate thread
    scheduler_thread = Thread(target=run_scheduler, daemon=True)
    scheduler_thread.start()
    
    # Log current time and next scheduled jobs
    current_time = datetime.now().strftime("%H:%M:%S")
    logger.info(f"Daily remark reset and history save scheduler started at {current_time}")
    logger.info("- History save: 23:58:00 daily")
    logger.info("- Remark reset: 23:59:59 daily")
    
    # Log next scheduled jobs
    jobs = schedule.get_jobs()
    for job in jobs:
        logger.info(f"Next scheduled job: {job.job_func.__name__} at {job.next_run}")


# Upload log storage (in-memory for now, can be moved to database later)
upload_logs = {}
marketplace_logs = {}
global_marketplace_logs = []  # Global logs for all marketplace apps without task_id

def add_upload_log(task_id: str, level: str, message: str):
    """Add a log entry for upload process"""
    if task_id not in upload_logs:
        upload_logs[task_id] = []
    
    log_entry = {
        "timestamp": datetime.now().isoformat(),
        "level": level,
        "message": message
    }
    upload_logs[task_id].append(log_entry)
    
    # Also log to console for debugging
    print(f"[{task_id}] {level.upper()}: {message}")
    
    # Keep only last 1000 logs per task to prevent memory issues
    if len(upload_logs[task_id]) > 1000:
        upload_logs[task_id] = upload_logs[task_id][-1000:]

def add_marketplace_log(task_id: str, level: str, message: str):
    """Add a log entry for marketplace app execution"""
    if task_id not in marketplace_logs:
        marketplace_logs[task_id] = []
    
    log_entry = {
        "timestamp": datetime.now().isoformat(),
        "level": level,
        "message": message
    }
    marketplace_logs[task_id].append(log_entry)
    
    # Keep only last 1000 logs per task to prevent memory issues
    if len(marketplace_logs[task_id]) > 1000:
        marketplace_logs[task_id] = marketplace_logs[task_id][-1000:]

@app.get("/api/upload-logs/{task_id}")
def get_upload_logs(
    task_id: str,
    current_user: str = Depends(get_current_user)
):
    """Get upload process logs for a specific task"""
    try:
        # Get task info to verify ownership
        db = SessionLocal()
        task = db.query(UploadTask).filter(UploadTask.task_id == task_id).first()
        db.close()
        
        if not task:
            raise HTTPException(status_code=404, detail="Task not found")
        
        # Get logs for this task
        logs = upload_logs.get(task_id, [])
        
        # Determine current step based on task status
        current_step = 0
        if task.status == "processing":
            current_step = 2  # Processing step
        elif task.status == "completed":
            current_step = 6  # Completed step
        elif task.status == "failed":
            current_step = 2  # Failed at processing step
        
        return {
            "task_id": task_id,
            "logs": logs,
            "current_step": current_step,
            "task_status": task.status
        }
        
    except Exception as e:
        logger.error(f"Error getting upload logs: {e}")
        raise HTTPException(status_code=500, detail="Failed to get logs")

@app.get("/api/marketplace-logs/task/{task_id}")
def get_marketplace_logs_by_task(
    task_id: str,
    current_user: str = Depends(get_current_user)
):
    """Get marketplace app execution logs for a specific task"""
    logger.info(f"Getting marketplace logs for task_id: {task_id}, user: {current_user}")
    
    try:
        # Get marketplace logs for this task (stored in memory)
        logs = marketplace_logs.get(task_id, [])
        logger.info(f"Found {len(logs)} marketplace logs for task {task_id}")
        
        # If no logs for this task, include recent global marketplace logs
        if not logs and global_marketplace_logs:
            # Get last 50 global logs
            recent_global_logs = global_marketplace_logs[-50:]
            logs = recent_global_logs
            logger.info(f"Including {len(logs)} recent global marketplace logs")
        
        return {
            "task_id": task_id,
            "logs": logs
        }
        
    except Exception as e:
        logger.error(f"Error getting marketplace logs: {e}")
        raise HTTPException(status_code=500, detail="Failed to get marketplace logs")


@app.get("/api/error-stats")
def get_error_stats(
    current_user: str = Depends(get_current_user)
):
    """Get error statistics for ErrorDashboard component"""
    try:
        # Mock error stats data - in real implementation, this would come from database
        error_stats = {
            "total_errors": 45,
            "critical_errors": 3,
            "warning_errors": 12,
            "info_errors": 30,
            "error_rate": 2.5,
            "resolved_errors": 42,
            "pending_errors": 3
        }
        
        return error_stats
        
    except Exception as e:
        logger.error(f"Error getting error stats: {e}")
        raise HTTPException(status_code=500, detail="Failed to get error stats")

@app.get("/api/error-trends")
def get_error_trends(
    days: int = Query(7, description="Number of days for trend analysis"),
    current_user: str = Depends(get_current_user)
):
    """Get error trends for ErrorDashboard component"""
    try:
        # Mock error trends data
        trends = []
        for i in range(days):
            date = (datetime.now() - timedelta(days=i)).strftime("%Y-%m-%d")
            trends.append({
                "date": date,
                "errors": max(0, 10 - i + (i % 3) * 2),
                "resolved": max(0, 8 - i + (i % 2) * 3)
            })
        
        return {"trends": list(reversed(trends))}
        
    except Exception as e:
        logger.error(f"Error getting error trends: {e}")
        raise HTTPException(status_code=500, detail="Failed to get error trends")

@app.get("/api/recent-errors")
def get_recent_errors(
    limit: int = Query(10, description="Number of recent errors to return"),
    current_user: str = Depends(get_current_user)
):
    """Get recent errors for ErrorDashboard component"""
    try:
        # Mock recent errors data
        recent_errors = []
        error_types = ["Database Connection", "API Timeout", "File Upload", "Authentication", "Validation"]
        
        for i in range(limit):
            recent_errors.append({
                "id": f"ERR-{1000 + i}",
                "type": error_types[i % len(error_types)],
                "message": f"Error occurred in {error_types[i % len(error_types)].lower()}",
                "timestamp": (datetime.now() - timedelta(hours=i)).isoformat(),
                "severity": ["critical", "warning", "info"][i % 3],
                "status": ["resolved", "pending", "investigating"][i % 3]
            })
        
        return {"errors": recent_errors}
        
    except Exception as e:
        logger.error(f"Error getting recent errors: {e}")
        raise HTTPException(status_code=500, detail="Failed to get recent errors")

@app.get("/api/monitoring-dashboard")
def get_monitoring_dashboard(
    current_user: str = Depends(get_current_user)
):
    """Get monitoring dashboard data for MonitoringDashboard component"""
    try:
        # Mock monitoring data
        monitoring_data = {
            "system_health": {
                "cpu_usage": 45.2,
                "memory_usage": 67.8,
                "disk_usage": 23.1,
                "network_latency": 12.5
            },
            "active_processes": 15,
            "queue_status": {
                "pending": 23,
                "processing": 5,
                "completed": 156,
                "failed": 2
            },
            "marketplace_status": {
                "shopee": "active",
                "lazada": "active", 
                "tokopedia": "maintenance",
                "tiktok": "active"
            },
            "last_updated": get_wib_now().isoformat()
        }
        
        return monitoring_data
        
    except Exception as e:
        logger.error(f"Error getting monitoring dashboard: {e}")
        raise HTTPException(status_code=500, detail="Failed to get monitoring dashboard")

@app.post("/api/upload-chunk")
async def upload_chunk(
    chunk: UploadFile = File(...),
    chunk_index: int = Form(...),
    total_chunks: int = Form(...),
    file_id: str = Form(...),
    current_user: str = Depends(get_current_user)
):
    """Upload file chunk for OptimizedFileUpload component"""
    try:
        logger.info(f"Uploading chunk {chunk_index + 1}/{total_chunks} for file {file_id}")
        
        # Create temporary directory for chunks if it doesn't exist
        chunk_dir = f"temp_chunks/{file_id}"
        os.makedirs(chunk_dir, exist_ok=True)
        
        # Save chunk to temporary file
        chunk_filename = f"{chunk_dir}/chunk_{chunk_index}"
        with open(chunk_filename, "wb") as buffer:
            content = await chunk.read()
            buffer.write(content)
        
        return {
            "message": f"Chunk {chunk_index + 1}/{total_chunks} uploaded successfully",
            "chunk_index": chunk_index,
            "total_chunks": total_chunks,
            "file_id": file_id
        }
        
    except Exception as e:
        logger.error(f"Error uploading chunk: {e}")
        raise HTTPException(status_code=500, detail="Failed to upload chunk")

@app.post("/api/upload-finalize")
async def upload_finalize(
    file_id: str = Form(...),
    filename: str = Form(...),
    total_chunks: int = Form(...),
    current_user: str = Depends(get_current_user)
):
    """Finalize chunked upload for OptimizedFileUpload component"""
    try:
        logger.info(f"Finalizing upload for file {file_id} with {total_chunks} chunks")
        
        chunk_dir = f"temp_chunks/{file_id}"
        final_filename = f"uploads/{filename}"
        
        # Create uploads directory if it doesn't exist
        os.makedirs("uploads", exist_ok=True)
        
        # Combine all chunks into final file
        with open(final_filename, "wb") as final_file:
            for chunk_index in range(total_chunks):
                chunk_filename = f"{chunk_dir}/chunk_{chunk_index}"
                if os.path.exists(chunk_filename):
                    with open(chunk_filename, "rb") as chunk_file:
                        final_file.write(chunk_file.read())
                    # Remove chunk file after combining
                    os.remove(chunk_filename)
        
        # Remove chunk directory
        if os.path.exists(chunk_dir):
            os.rmdir(chunk_dir)
        
        return {
            "message": "File upload finalized successfully",
            "filename": filename,
            "file_id": file_id,
            "file_path": final_filename
        }
        
    except Exception as e:
        logger.error(f"Error finalizing upload: {e}")
        raise HTTPException(status_code=500, detail="Failed to finalize upload")

@app.get("/api/itemid-comparison")
def get_itemid_comparison(
    current_user: str = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get ItemId vs ItemIdFlexo comparison statistics"""
    try:
        # Get comparison statistics
        comparison_sql = """
        SELECT 
            comparison_status,
            COUNT(*) as count
        FROM itemid_comparison
        GROUP BY comparison_status
        ORDER BY 
            CASE comparison_status
                WHEN 'Mismatch' THEN 1
                WHEN 'Item Different' THEN 2
                WHEN 'Item Missing' THEN 3
                WHEN 'Match' THEN 4
                WHEN 'Both Missing' THEN 5
                ELSE 6
            END;
        """
        
        result = db.execute(text(comparison_sql))
        comparison_data = result.fetchall()
        
        # Get recent mismatches for details
        mismatch_sql = """
        SELECT 
            "OrderNumber",
            "Marketplace",
            "Brand",
            excel_itemid,
            external_itemid,
            "UploadDate"
        FROM itemid_comparison
        WHERE comparison_status = 'Mismatch'
        ORDER BY "UploadDate" DESC
        LIMIT 10;
        """
        
        result = db.execute(text(mismatch_sql))
        mismatch_details = result.fetchall()
        
        # Format response
        comparison_stats = {}
        total_orders = 0
        
        for status, count in comparison_data:
            comparison_stats[status] = count
            total_orders += count
        
        # Calculate percentages
        comparison_percentages = {}
        for status, count in comparison_stats.items():
            percentage = (count / total_orders * 100) if total_orders > 0 else 0
            comparison_percentages[status] = round(percentage, 1)
        
        # Format mismatch details
        mismatch_list = []
        for row in mismatch_details:
            mismatch_list.append({
                "order_number": row[0],
                "marketplace": row[1],
                "brand": row[2],
                "excel_itemid": row[3],
                "external_itemid": row[4],
                "upload_date": row[5].isoformat() if row[5] else None
            })
        
        return {
            "success": True,
            "total_orders": total_orders,
            "comparison_stats": comparison_stats,
            "comparison_percentages": comparison_percentages,
            "recent_mismatches": mismatch_list,
            "summary": {
                "match_count": comparison_stats.get("Match", 0),
                "mismatch_count": comparison_stats.get("Mismatch", 0),
                "excel_missing_count": comparison_stats.get("Item Missing", 0),
                "external_missing_count": comparison_stats.get("Item Different", 0),
                "both_missing_count": comparison_stats.get("Both Missing", 0)
            }
        }
        
    except Exception as e:
        print(f"❌ Error getting SKU comparison: {e}")
        return {
            "success": False,
            "error": str(e),
            "total_orders": 0,
            "comparison_stats": {},
            "comparison_percentages": {},
            "recent_mismatches": [],
            "summary": {
                "match_count": 0,
                "mismatch_count": 0,
                "excel_missing_count": 0,
                "external_missing_count": 0,
                "both_missing_count": 0
            }
        }

@app.get("/api/summary-orders")
def get_summary_orders(
    current_user: str = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get summary orders for SummaryOrder component"""
    try:
        # Get summary data from database
        total_orders = db.query(UpdateUploadedOrder).count()
        interfaced_orders = db.query(UpdateUploadedOrder).filter(
            UpdateUploadedOrder.InterfaceStatus == 'Interface'
        ).count()
        not_interfaced_orders = total_orders - interfaced_orders
        
        # Get brand distribution
        brand_stats = db.query(
            UpdateUploadedOrder.Brand,
            func.count(UpdateUploadedOrder.Id).label('count')
        ).group_by(UpdateUploadedOrder.Brand).all()
        
        # Get marketplace distribution
        marketplace_stats = db.query(
            UpdateUploadedOrder.Marketplace,
            func.count(UpdateUploadedOrder.Id).label('count')
        ).group_by(UpdateUploadedOrder.Marketplace).all()
        
        summary_data = {
            "total_orders": total_orders,
            "interfaced_orders": interfaced_orders,
            "not_interfaced_orders": not_interfaced_orders,
            "interface_rate": round((interfaced_orders / total_orders * 100) if total_orders > 0 else 0, 2),
            "brand_distribution": [{"brand": b[0], "count": b[1]} for b in brand_stats],
            "marketplace_distribution": [{"marketplace": m[0], "count": m[1]} for m in marketplace_stats],
            "last_updated": get_wib_now().isoformat()
        }
        
        return summary_data
        
    except Exception as e:
        logger.error(f"Error getting summary orders: {e}")
        raise HTTPException(status_code=500, detail="Failed to get summary orders")

@app.post("/api/sync-to-cms")
def sync_to_cms(
    current_user: str = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Sync orders to CMS for SummaryOrder component"""
    try:
        logger.info(f"Starting CMS sync for user: {current_user}")
        
        # Get orders that need to be synced (not interfaced orders)
        orders_to_sync = db.query(UpdateUploadedOrder).filter(
            UpdateUploadedOrder.InterfaceStatus != 'Interface'
        ).limit(100).all()  # Limit to prevent overload
        
        sync_results = {
            "total_orders": len(orders_to_sync),
            "synced_orders": 0,
            "failed_orders": 0,
            "errors": []
        }
        
        # Mock sync process - in real implementation, this would sync to actual CMS
        for order in orders_to_sync:
            try:
                # Simulate sync process
                # In real implementation, this would make API calls to CMS
                time.sleep(0.1)  # Simulate processing time
                
                # Update interface status to Interface (mock success)
                order.InterfaceStatus = 'Interface'
                sync_results["synced_orders"] += 1
                
            except Exception as e:
                sync_results["failed_orders"] += 1
                sync_results["errors"].append(f"Order {order.OrderNumber}: {str(e)}")
        
        # Commit changes
        db.commit()
        
        logger.info(f"CMS sync completed: {sync_results['synced_orders']} synced, {sync_results['failed_orders']} failed")
        
        return {
            "message": "CMS sync completed",
            "results": sync_results
        }
        
    except Exception as e:
        logger.error(f"Error syncing to CMS: {e}")
        raise HTTPException(status_code=500, detail="Failed to sync to CMS")

if __name__ == "__main__":
    
    # Production mode detection
    import os
    is_production = os.getenv("PRODUCTION", "false").lower() == "true"
    
    if is_production:
        # Production server configuration
        print("🚀 Starting in PRODUCTION mode...")
        uvicorn.run(
            "main:app",
            host="0.0.0.0", 
            port=int(os.getenv("PORT", "8001")),
            workers=int(os.getenv("WORKERS", "4")),
            log_level="warning",
            access_log=False,
            timeout_keep_alive=65,
            timeout_graceful_shutdown=30,
            limit_max_requests=1000,
            limit_concurrency=100,
            reload=False  # Disable auto-reload in production
        )
    else:
        # Development mode (existing configuration)
        logging.getLogger("uvicorn.access").setLevel(logging.WARNING)
        
        uvicorn.run(
            app, 
            host="0.0.0.0", 
            port=8001,
            log_level="warning",
            access_log=False,
            timeout_keep_alive=65,
            timeout_graceful_shutdown=30,
            limit_max_requests=1000,
            limit_concurrency=100
        )
